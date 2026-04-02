"""
utils/export_utils.py
---------------------
Shared Excel (openpyxl) export helper used by all journal widgets
and db_manager's COA export.

No Qt imports — pure Python / openpyxl only.

Usage
-----
    from utils.export_utils import export_to_xls

    rows = [
        {'date': '03/25/2026', 'reference_no': 'JV-001', ...},
        ...
    ]
    columns = [
        ('Date',         'date'),
        ('Reference No', 'reference_no'),
        ('Particulars',  'particulars'),
        ('Account Desc', 'account_description'),
        ('Account Code', 'account_code'),
        ('Debit',        'debit'),
        ('Credit',       'credit'),
    ]
    col_widths = {1: 12, 2: 16, 3: 28, 4: 28, 5: 14, 6: 14, 7: 14}

    count, err = export_to_xls(
        rows       = rows,
        path       = '/path/to/output.xlsx',
        sheet_title= 'Cash Disbursement Journal',
        columns    = columns,
        col_widths = col_widths,
    )
    if err:
        print('Export failed:', err)
    else:
        print(f'{count} rows exported.')
"""

from __future__ import annotations

from datetime import datetime
from typing import Any

# ---------------------------------------------------------------------------
# Colour / style constants  (centralised so restyling is one-place)
# ---------------------------------------------------------------------------

HEADER_BG     = '2F5496'   # dark blue  — header row background
HEADER_FG     = 'FFFFFF'   # white      — header row text
ALT_ROW_BG    = 'DCE6F1'   # light blue — alternating row fill
TITLE_SIZE    = 14         # pt — sheet title font size
SUBTITLE_SIZE = 11         # pt — "For the Year …" font size
HEADER_SIZE   = 11         # pt — column header font size
DATA_SIZE     = 10         # pt — data row font size
HEADER_ROW    = 5          # row index for column headers (1-based)
DATA_START    = 6          # first data row (1-based)
HEADER_HEIGHT = 28         # points
DATA_HEIGHT   = 18         # points
TITLE_HEIGHT  = 22         # points

# Keys whose values should be right-aligned (numeric columns)
_NUMERIC_KEYS = frozenset({'debit', 'credit', 'amount', 'net_amount',
                            'output_vat', 'input_vat', 'gross_amount',
                            'goods', 'services'})


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def export_to_xls(
    rows:        list[dict],
    path:        str,
    sheet_title: str,
    columns:     list[tuple[str, str]],
    col_widths:  dict[int, int] | None = None,
    year:        int | None = None,
    numeric_keys: frozenset[str] | None = None,
) -> tuple[int, str]:
    """
    Write *rows* to a formatted .xlsx file at *path*.

    Parameters
    ----------
    rows        : list of dicts, one per data row
    path        : output file path (must end in .xlsx)
    sheet_title : used as the sheet name and the bold title in row 2
    columns     : list of (header_label, dict_key) pairs — defines both
                  column order and the header row text
    col_widths  : optional {1-based-col-index: width} mapping
    year        : fiscal year shown in the subtitle; defaults to current year
    numeric_keys: set of dict keys whose columns get right-alignment;
                  defaults to the module-level _NUMERIC_KEYS set

    Returns
    -------
    (count, error_string) — count is 0 and error_string is non-empty on failure
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return 0, 'openpyxl is not installed.\nInstall with: pip install openpyxl'

    _year       = year if year else datetime.now().year
    _num_keys   = numeric_keys if numeric_keys is not None else _NUMERIC_KEYS
    headers     = [h for h, _ in columns]
    keys        = [k for _, k in columns]
    n_cols      = len(headers)

    # ── Styles ────────────────────────────────────────────────────────────
    hdr_font    = Font(name='Arial', bold=True, color=HEADER_FG, size=HEADER_SIZE)
    hdr_fill    = PatternFill('solid', start_color=HEADER_BG)
    hdr_align   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    title_font  = Font(name='Arial', bold=True,   size=TITLE_SIZE)
    sub_font    = Font(name='Arial', italic=True,  size=SUBTITLE_SIZE)
    data_font   = Font(name='Arial', size=DATA_SIZE)
    alt_fill    = PatternFill('solid', start_color=ALT_ROW_BG)
    thin        = Side(style='thin', color='B0B0B0')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    left_align  = Alignment(horizontal='left',   vertical='center')
    right_align = Alignment(horizontal='right',  vertical='center')
    center_align= Alignment(horizontal='center', vertical='center')

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title[:31]  # Excel sheet name max length = 31

        # ── Title rows ────────────────────────────────────────────────────
        last_col_letter = get_column_letter(n_cols)

        ws.merge_cells(f'A2:{last_col_letter}2')
        ws['A2'].value     = sheet_title.upper()
        ws['A2'].font      = title_font
        ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[2].height = TITLE_HEIGHT

        ws.merge_cells(f'A3:{last_col_letter}3')
        ws['A3'].value     = f'For the Year {_year}'
        ws['A3'].font      = sub_font
        ws['A3'].alignment = Alignment(horizontal='left', vertical='center')

        # ── Header row ────────────────────────────────────────────────────
        ws.row_dimensions[HEADER_ROW].height = HEADER_HEIGHT
        for ci, header in enumerate(headers, 1):
            cell            = ws.cell(row=HEADER_ROW, column=ci, value=header)
            cell.font       = hdr_font
            cell.fill       = hdr_fill
            cell.alignment  = hdr_align
            cell.border     = border

        # ── Data rows ─────────────────────────────────────────────────────
        for ri, row in enumerate(rows):
            row_idx = DATA_START + ri
            ws.row_dimensions[row_idx].height = DATA_HEIGHT
            fill = alt_fill if ri % 2 == 0 else None

            for ci, key in enumerate(keys, 1):
                val  = row.get(key, '') or ''
                cell = ws.cell(row=row_idx, column=ci, value=val)
                cell.font   = data_font
                cell.border = border
                cell.alignment = (
                    right_align if key in _num_keys else left_align
                )
                if fill:
                    cell.fill = fill

        # ── Column widths ─────────────────────────────────────────────────
        if col_widths:
            for ci, width in col_widths.items():
                ws.column_dimensions[get_column_letter(ci)].width = width

        # ── Freeze panes + auto-filter ────────────────────────────────────
        ws.freeze_panes = f'A{DATA_START}'
        ws.auto_filter.ref = f'A{HEADER_ROW}:{last_col_letter}{HEADER_ROW}'

        wb.save(path)
        return len(rows), ''

    except Exception as exc:
        return 0, str(exc)


# ---------------------------------------------------------------------------
# Backwards-compatible thin wrapper used by alphalist_widget
# (keeps the old call signature working without changes)
# ---------------------------------------------------------------------------

def export_alphalist_to_xls(
    entries: list[dict],
    path:    str,
    year:    int | None = None,
) -> tuple[int, str]:
    """Thin wrapper that calls export_to_xls with the alphalist column layout."""
    columns = [
        ('TIN',          'tin'),
        ('Entry Type',   'entry_type'),
        ('Company Name', 'company_name'),
        ('First Name',   'first_name'),
        ('Middle Name',  'middle_name'),
        ('Last Name',    'last_name'),
        ('Address 1',    'address1'),
        ('Address 2',    'address2'),
    ]
    col_widths = {1: 16, 2: 16, 3: 28, 4: 18, 5: 18, 6: 18, 7: 30, 8: 30}
    return export_to_xls(
        rows        = entries,
        path        = path,
        sheet_title = 'Alphalist',
        columns     = columns,
        col_widths  = col_widths,
        year        = year,
    )


__all__ = [
    'export_to_xls',
    'export_alphalist_to_xls',
    'HEADER_BG',
    'HEADER_FG',
    'ALT_ROW_BG',
    'HEADER_ROW',
    'DATA_START',
]