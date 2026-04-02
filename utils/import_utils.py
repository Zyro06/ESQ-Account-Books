"""
utils/import_utils.py
---------------------
Shared Excel (openpyxl) import helpers used by all journal widgets
and the fullbook importer.

No Qt imports — pure Python / openpyxl only.

Two layers are provided:

1.  Low-level helpers (_find_header_row, _col, _get)
    — used by fullbook_importer.py for its flexible multi-sheet logic.

2.  High-level import_from_xls()
    — used by CDJ, CRJ, and GJ widgets which all share identical logic.

Usage — journal widgets (CDJ / CRJ / GJ)
-----------------------------------------
    from utils.import_utils import import_from_xls

    COLUMNS = [
        ('Date',               'date'),
        ('Reference No',       'reference_no'),
        ('Particulars',        'particulars'),
        ('Account Description','account_description'),
        ('Account Code',       'account_code'),
        ('Debit',              'debit'),
        ('Credit',             'credit'),
    ]

    try:
        imported, skipped, errors = import_from_xls(
            path           = '/path/to/file.xlsx',
            db_manager     = self.db_manager,
            add_method_name= 'add_cash_disbursement_entry',
            columns        = COLUMNS,
        )
    except RuntimeError as exc:
        QMessageBox.critical(self, 'Import Failed', str(exc))
        return

Usage — fullbook_importer
--------------------------
    from utils.import_utils import find_header_row, col_index, get_cell

    col_map, data_start = find_header_row(ws, ['date', 'reference no'])
    ci_date = col_index(col_map, 'date')
    value   = get_cell(row_vals, ci_date)
"""

from __future__ import annotations

import os
from typing import Any

from utils.date_utils import norm_str   # no Qt dependency


# ---------------------------------------------------------------------------
# Low-level helpers  (used by fullbook_importer)
# ---------------------------------------------------------------------------

def find_header_row(
    ws,
    target_headers: list[str],
    max_scan: int = 10,
) -> tuple[dict[str, int], int | None]:
    """
    Scan the first *max_scan* rows of *ws* for a row that contains at least
    half of the *target_headers* (case-insensitive, partial match allowed).

    Returns
    -------
    (col_map, data_start_row)
        col_map          : {normalised_cell_text: 0-based column index}
        data_start_row   : 1-based row index of the first data row (header + 1)
        Both are ({}, None) if no matching header row is found.
    """
    target_set = {h.lower() for h in target_headers}

    for ri, row_vals in enumerate(
            ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        row_strs = [norm_str(v).lower() for v in row_vals]
        hits = sum(1 for h in target_set if any(h in cell for cell in row_strs))
        if hits >= max(1, len(target_set) // 2):
            col_map = {}
            for ci, val in enumerate(row_vals):
                s = norm_str(val).lower()
                if s:
                    col_map[s] = ci
            return col_map, ri + 1

    return {}, None


def col_index(col_map: dict[str, int], *aliases: str) -> int | None:
    """
    Return the first matching 0-based column index for any of *aliases*,
    trying exact match first then partial match.  Returns None if not found.
    """
    for alias in aliases:
        a = alias.lower()
        if a in col_map:
            return col_map[a]
        for key, idx in col_map.items():
            if a in key or key in a:
                return idx
    return None


def get_cell(row_vals: tuple, idx: int | None) -> Any:
    """Return the cell value at *idx*, or None if idx is None / out of range."""
    if idx is None or idx >= len(row_vals):
        return None
    return row_vals[idx]


# ---------------------------------------------------------------------------
# High-level import  (used by CDJ / CRJ / GJ widgets)
# ---------------------------------------------------------------------------

def import_from_xls(
    path:            str,
    db_manager,
    add_method_name: str,
    columns:         list[tuple[str, str]],
) -> tuple[int, int, list[str]]:
    """
    Import rows from an .xlsx file into the database.

    Parameters
    ----------
    path             : path to the Excel file
    db_manager       : DatabaseManager instance
    add_method_name  : name of the db_manager method to call per row
                       e.g. 'add_cash_disbursement_entry'
    columns          : list of (header_label, dict_key) pairs that define
                       which columns to read and what dict keys to use

    Returns
    -------
    (imported, skipped, errors)
        imported : number of rows successfully inserted
        skipped  : number of rows skipped (duplicates, missing fields, errors)
        errors   : list of human-readable error strings

    Raises
    ------
    RuntimeError  if the file cannot be opened
    ValueError    if no matching header row is found
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError(
            'openpyxl is not installed.\nInstall with: pip install openpyxl'
        )

    # Build a lowercase header → dict_key mapping
    header_map = {label.lower(): key for label, key in columns}
    all_keys   = [key for _, key in columns]

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        raise RuntimeError(f'Cannot open "{os.path.basename(path)}".')

    ws         = wb.active
    col_idx    = {}
    data_start = None

    # Find the header row
    for r_idx, row_vals in enumerate(
            ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        if not any(
            v and str(v).strip().lower() in header_map for v in row_vals
        ):
            continue
        for ci, cv in enumerate(row_vals):
            if cv is None:
                continue
            key = header_map.get(str(cv).strip().lower())
            if key:
                col_idx[key] = ci
        data_start = r_idx + 1
        break

    if not col_idx:
        wb.close()
        raise ValueError('Could not find a matching header row in the first 10 rows.')

    def _val(rv, k: str) -> str:
        i = col_idx.get(k)
        if i is None or i >= len(rv):
            return ''
        v = rv[i]
        return str(v).strip() if v is not None else ''

    imported = skipped = 0
    errors: list[str] = []
    add_fn = getattr(db_manager, add_method_name)

    for rn, rv in enumerate(
            ws.iter_rows(min_row=data_start, values_only=True), data_start):
        if all(v is None for v in rv):
            continue
        try:
            data = {k: _val(rv, k) for k in all_keys}

            # Coerce numeric fields
            data['debit']  = float((data.get('debit',  '') or 0)
                                   .replace(',', ''))
            data['credit'] = float((data.get('credit', '') or 0)
                                   .replace(',', ''))

            if not data.get('reference_no'):
                skipped += 1
                errors.append(f'Row {rn}: missing reference_no — skipped')
                continue

            if add_fn(data):
                imported += 1
            else:
                skipped += 1
                errors.append(f'Row {rn}: failed to insert — skipped')

        except Exception as exc:
            skipped += 1
            errors.append(f'Row {rn}: {exc}')

    wb.close()
    return imported, skipped, errors


# ---------------------------------------------------------------------------
# SJ / PJ grouped import  (rows share a reference_no → one header + N lines)
# ---------------------------------------------------------------------------

def import_grouped_from_xls(
    path:              str,
    db_manager,
    add_method_name:   str,
    columns:           list[tuple[str, str]],
    group_key:         str = 'reference_no',
    customer_key:      str | None = None,   # 'customer_name' for SJ
    payee_key:         str | None = None,   # 'payee_name' for PJ
) -> tuple[int, int, list[str]]:
    """
    Import grouped journal entries (SJ / PJ style) where multiple rows
    share the same reference_no and each row represents one journal line.

    Parameters
    ----------
    path             : path to the Excel file
    db_manager       : DatabaseManager instance
    add_method_name  : e.g. 'add_sales_entry'
    columns          : (header_label, dict_key) pairs
    group_key        : key used to group rows into a single entry
    customer_key     : dict key for the customer/payee name on the header
    payee_key        : alias for customer_key (PJ uses 'payee_name')

    Returns
    -------
    (imported, skipped, errors)
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError(
            'openpyxl is not installed.\nInstall with: pip install openpyxl'
        )

    header_map = {label.lower(): key for label, key in columns}

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        raise RuntimeError(f'Cannot open "{os.path.basename(path)}".')

    ws         = wb.active
    col_idx    = {}
    data_start = None

    for r_idx, row_vals in enumerate(
            ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        if not any(
            v and str(v).strip().lower() in header_map for v in row_vals
        ):
            continue
        for ci, cv in enumerate(row_vals):
            if cv is None:
                continue
            key = header_map.get(str(cv).strip().lower())
            if key:
                col_idx[key] = ci
        data_start = r_idx + 1
        break

    if not col_idx:
        wb.close()
        raise ValueError('Could not find a matching header row.')

    def _val(rv, k: str) -> str:
        i = col_idx.get(k)
        if i is None or i >= len(rv):
            return ''
        v = rv[i]
        return str(v).strip() if v is not None else ''

    # Group rows by (date, reference_no)
    from collections import defaultdict
    groups: dict[tuple, list] = defaultdict(list)
    order:  list[tuple]        = []

    for rn, rv in enumerate(
            ws.iter_rows(min_row=data_start, values_only=True), data_start):
        if all(v is None for v in rv):
            continue
        ref  = _val(rv, 'reference_no')
        date = _val(rv, 'date')
        key  = (date, ref)
        if key not in groups:
            order.append(key)
        groups[key].append(rv)

    imported = skipped = 0
    errors: list[str] = []
    add_fn = getattr(db_manager, add_method_name)

    name_key = customer_key or payee_key or 'customer_name'

    for date, ref in order:
        if not date or not ref:
            skipped += 1
            continue

        first = groups[(date, ref)][0]
        entry = {
            'date':         date,
            name_key:       _val(first, name_key),
            'reference_no': ref,
            'tin':          _val(first, 'tin'),
            'particulars':  _val(first, 'particulars'),
            'lines':        [],
        }
        # PJ extra field
        if 'branch_code' in col_idx:
            entry['branch_code'] = _val(first, 'branch_code')

        for rv in groups[(date, ref)]:
            ac = _val(rv, 'account_code')
            ad = _val(rv, 'account_description')
            try:
                dr = float(_val(rv, 'debit').replace(',', '')  or 0)
            except ValueError:
                dr = 0.0
            try:
                cr = float(_val(rv, 'credit').replace(',', '') or 0)
            except ValueError:
                cr = 0.0
            if ac and (dr > 0 or cr > 0):
                entry['lines'].append({
                    'account_code':        ac,
                    'account_description': ad,
                    'debit':               dr,
                    'credit':              cr,
                })

        if not entry['lines']:
            skipped += 1
            errors.append(f'Ref {ref}: no valid lines — skipped')
            continue

        if add_fn(entry):
            imported += 1
        else:
            skipped += 1
            errors.append(f'Ref {ref}: failed to insert — skipped')

    wb.close()
    return imported, skipped, errors


# ---------------------------------------------------------------------------
# Alphalist importer  (moved from alphalist_widget.py)
# ---------------------------------------------------------------------------

def import_alphalist_from_xls(path: str, db_manager) -> dict:
    """
    Import alphalist entries from an .xlsx file.

    Returns a summary dict:
        {'imported': int, 'skipped_duplicate': int,
         'skipped_invalid': int, 'errors': [str]}
    """
    _ALPHALIST_COLUMNS = [
        ('TIN',          'tin'),
        ('Entry Type',   'entry_type'),
        ('Company Name', 'company_name'),
        ('First Name',   'first_name'),
        ('Middle Name',  'middle_name'),
        ('Last Name',    'last_name'),
        ('Address 1',    'address1'),
        ('Address 2',    'address2'),
    ]
    _EXTRA_ALIASES = {
        'address 2': 'address2', 'address 1': 'address1',
        'company':   'company_name', 'first': 'first_name',
        'middle':    'middle_name',  'last':  'last_name',
        'entry type':'entry_type',
    }

    def _fmt_tin(raw) -> str:
        digits = ''.join(ch for ch in str(raw or '') if ch.isdigit())
        if not digits:
            return ''
        digits = digits.zfill(9)[-9:]
        return f'{digits[:3]}-{digits[3:6]}-{digits[6:]}'

    summary = {'imported': 0, 'skipped_duplicate': 0, 'skipped_invalid': 0, 'errors': []}

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError('openpyxl is required.\nInstall with: pip install openpyxl')

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        raise RuntimeError(
            f'Cannot open "{os.path.basename(path)}".\n\n'
            'If this is an old .xls file, please re-save it as .xlsx first.')

    ws = wb.active
    header_to_key = {h.lower(): k for h, k in _ALPHALIST_COLUMNS}
    header_to_key.update(_EXTRA_ALIASES)

    col_index: dict[str, int] = {}
    data_start_row = None

    for r_idx, row_vals in enumerate(
            ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if not any(v and str(v).strip().upper() == 'TIN' for v in row_vals):
            continue
        for ci, cv in enumerate(row_vals):
            if cv is None:
                continue
            norm = str(cv).strip().lower()
            key  = header_to_key.get(norm)
            if key:
                col_index[key] = ci
        data_start_row = r_idx + 1
        break

    if not col_index or 'tin' not in col_index:
        wb.close()
        raise ValueError("Could not find a header row with a 'TIN' column in the first 10 rows.")

    def _val(row_vals, key: str) -> str:
        idx = col_index.get(key)
        if idx is None or idx >= len(row_vals):
            return ''
        v = row_vals[idx]
        return str(v).strip() if v is not None else ''

    for row_num, row_vals in enumerate(
            ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        if all(v is None for v in row_vals):
            continue
        tin = _fmt_tin(_val(row_vals, 'tin'))
        if not tin:
            summary['skipped_invalid'] += 1
            summary['errors'].append(f'Row {row_num}: invalid TIN "{_val(row_vals, "tin")}"')
            continue

        entry_type = _val(row_vals, 'entry_type') or 'Customer&Vendor'
        if entry_type not in ('Customer&Vendor', 'Customer', 'Vendor'):
            entry_type = 'Customer&Vendor'

        data = {
            'tin':          tin,
            'company_name': _val(row_vals, 'company_name'),
            'first_name':   _val(row_vals, 'first_name'),
            'middle_name':  _val(row_vals, 'middle_name'),
            'last_name':    _val(row_vals, 'last_name'),
            'address1':     _val(row_vals, 'address1'),
            'address2':     _val(row_vals, 'address2'),
            'entry_type':   entry_type,
        }
        if db_manager.add_alphalist(data):
            summary['imported'] += 1
        else:
            summary['skipped_duplicate'] += 1
            summary['errors'].append(f'Row {row_num}: duplicate TIN "{tin}" — skipped')

    wb.close()
    return summary


__all__ = [
    'find_header_row',
    'col_index',
    'get_cell',
    'import_from_xls',
    'import_grouped_from_xls',
    'import_alphalist_from_xls',
]