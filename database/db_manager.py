import sqlite3
from datetime import datetime
from typing import List, Dict, Any, Optional

_AR_NUM            = '1110'
_OUTPUT_VAT_NUM    = '2210'
_INPUT_VAT_NUM     = '1320'
_AP_NUM            = '2010'
_SALES_GOODS_NUM   = '4010'
_SALES_SERVICE_NUM = '4020'


def _numeric_suffix(code: str) -> str:
    if '-' in code:
        return code.rsplit('-', 1)[-1].strip()
    return code.strip()


class DatabaseManager:
    def __init__(self, db_path: str):
        self.db_path = db_path
        self.connection = None
        self.current_year = datetime.now().year

    def get_connection(self):
        if self.connection is None:
            self.connection = sqlite3.connect(self.db_path)
            self.connection.row_factory = sqlite3.Row
            self.connection.execute("PRAGMA foreign_keys = ON")
        return self.connection

    def set_current_year(self, year: int):
        self.current_year = year

    def get_current_year(self) -> int:
        return self.current_year

    # ------------------------------------------------------------------
    # Schema + migration
    # ------------------------------------------------------------------

    def initialize_database(self, coa_xlsx_path: str = None, use_default_coa: bool = True):
        conn = self.get_connection()
        cursor = conn.cursor()

        # COA
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS chart_of_accounts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                account_code TEXT UNIQUE NOT NULL,
                account_description TEXT NOT NULL,
                normal_balance TEXT DEFAULT 'Debit',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''')

        # Alphalist
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS alphalist (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tin TEXT UNIQUE NOT NULL,
                company_name TEXT, first_name TEXT, middle_name TEXT, last_name TEXT,
                address1 TEXT, address2 TEXT,
                entry_type TEXT DEFAULT 'Customer&Vendor',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''')

        # Alphalist migrations
        cursor.execute("PRAGMA table_info(alphalist)")
        cols = [c[1] for c in cursor.fetchall()]
        if 'entry_type' not in cols:
            cursor.execute("ALTER TABLE alphalist ADD COLUMN entry_type TEXT DEFAULT 'Customer'")

        # COA migration
        cursor.execute("PRAGMA table_info(chart_of_accounts)")
        coa_cols = [c[1] for c in cursor.fetchall()]
        if 'normal_balance' not in coa_cols:
            cursor.execute("ALTER TABLE chart_of_accounts ADD COLUMN normal_balance TEXT DEFAULT 'Debit'")
            cursor.execute("UPDATE chart_of_accounts SET normal_balance='Credit' WHERE account_code LIKE 'COA-2%' OR account_code LIKE 'COA-3%' OR account_code LIKE 'COA-4%' OR account_code LIKE 'COA-7%'")
            cursor.execute("UPDATE chart_of_accounts SET normal_balance='Credit' WHERE account_description LIKE 'ACCUM%DEP%' OR account_description LIKE 'ACCUM%AMORT%'")
            cursor.execute("UPDATE chart_of_accounts SET normal_balance='Debit' WHERE account_description LIKE '%DRAWING%'")

        # ── NEW: Sales Journal header + lines ─────────────────────────────
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS sales_journal (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date DATE NOT NULL,
                customer_name TEXT NOT NULL,
                reference_no TEXT NOT NULL,
                tin TEXT,
                particulars TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS sales_journal_lines (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                journal_id INTEGER NOT NULL REFERENCES sales_journal(id) ON DELETE CASCADE,
                account_code TEXT NOT NULL,
                account_description TEXT,
                debit REAL DEFAULT 0,
                credit REAL DEFAULT 0
            )''')

        # ── NEW: Purchase Journal header + lines ──────────────────────────
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS purchase_journal (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date DATE NOT NULL,
                payee_name TEXT,
                reference_no TEXT NOT NULL,
                tin TEXT,
                branch_code TEXT,
                particulars TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS purchase_journal_lines (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                journal_id INTEGER NOT NULL REFERENCES purchase_journal(id) ON DELETE CASCADE,
                account_code TEXT NOT NULL,
                account_description TEXT,
                debit REAL DEFAULT 0,
                credit REAL DEFAULT 0
            )''')

        # CDJ, CRJ, GJ unchanged
        for tbl in ('cash_disbursement_journal', 'cash_receipts_journal', 'general_journal'):
            cursor.execute(f'''
                CREATE TABLE IF NOT EXISTS {tbl} (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date DATE NOT NULL,
                    reference_no TEXT NOT NULL,
                    particulars TEXT,
                    account_code TEXT NOT NULL,
                    account_description TEXT,
                    debit REAL DEFAULT 0,
                    credit REAL DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )''')

        conn.commit()

        # ── Migrate old flat SJ/PJ if lines tables are empty ──────────────
        self._migrate_old_sj(cursor, conn)
        self._migrate_old_pj(cursor, conn)

        # Seed COA
        cursor.execute('SELECT COUNT(*) FROM chart_of_accounts')
        if cursor.fetchone()[0] == 0:
            if coa_xlsx_path:
                self._import_coa_from_xlsx(coa_xlsx_path)
            elif use_default_coa:
                self._seed_default_coa()

    # ------------------------------------------------------------------
    # Migration helpers
    # ------------------------------------------------------------------

    def _migrate_old_sj(self, cursor, conn):
        """Convert old flat sales_journal rows to header+lines if needed."""
        # Check if old schema had net_amount column
        cursor.execute("PRAGMA table_info(sales_journal)")
        cols = {c[1] for c in cursor.fetchall()}
        if 'net_amount' not in cols:
            return  # Already new schema, nothing to do

        # Check if lines table is empty (fresh migration)
        cursor.execute("SELECT COUNT(*) FROM sales_journal_lines")
        if cursor.fetchone()[0] > 0:
            # Already migrated
            self._drop_old_sj_columns(cursor, conn)
            return

        # Read old flat rows
        cursor.execute("""
            SELECT id, date, customer_name, reference_no, tin,
                   net_amount, output_vat, gross_amount, goods, services, particulars
            FROM sales_journal
        """)
        old_rows = cursor.fetchall()
        if not old_rows:
            self._drop_old_sj_columns(cursor, conn)
            return

        # Find account codes from COA
        def _find_code(suffix):
            cursor.execute("SELECT account_code, account_description FROM chart_of_accounts")
            for row in cursor.fetchall():
                if _numeric_suffix(row[0]) == suffix:
                    return row[0], row[1]
            return suffix, suffix

        ar_code,    ar_desc    = _find_code(_AR_NUM)
        vat_code,   vat_desc   = _find_code(_OUTPUT_VAT_NUM)
        goods_code, goods_desc = _find_code(_SALES_GOODS_NUM)
        svc_code,   svc_desc   = _find_code(_SALES_SERVICE_NUM)

        # We need to rebuild the sales_journal table without old columns
        # First create temp table, copy headers, drop old, rename
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS _sj_headers_temp (
                id INTEGER PRIMARY KEY,
                date DATE NOT NULL,
                customer_name TEXT NOT NULL,
                reference_no TEXT NOT NULL,
                tin TEXT,
                particulars TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )""")

        for row in old_rows:
            old_id      = row[0]
            date        = row[1]
            customer    = row[2]
            ref         = row[3]
            tin         = row[4]
            net         = float(row[5] or 0)
            output_vat  = float(row[6] or 0)
            gross       = float(row[7] or 0)
            goods       = float(row[8] or 0)
            services    = float(row[9] or 0)
            particulars = row[10]

            # Insert header with same id
            cursor.execute("""
                INSERT INTO _sj_headers_temp
                (id, date, customer_name, reference_no, tin, particulars)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (old_id, date, customer, ref, tin, particulars))

            # Insert lines
            lines = []
            if gross > 0:
                lines.append((old_id, ar_code,    ar_desc,    gross, 0))
            if output_vat > 0:
                lines.append((old_id, vat_code,   vat_desc,   0, output_vat))
            if goods > 0:
                lines.append((old_id, goods_code, goods_desc, 0, goods))
            if services > 0:
                lines.append((old_id, svc_code,   svc_desc,   0, services))

            for jid, ac, ad, dr, cr in lines:
                cursor.execute("""
                    INSERT INTO sales_journal_lines
                    (journal_id, account_code, account_description, debit, credit)
                    VALUES (?, ?, ?, ?, ?)
                """, (jid, ac, ad, dr, cr))

        # Swap tables
        cursor.execute("DROP TABLE sales_journal")
        cursor.execute("ALTER TABLE _sj_headers_temp RENAME TO sales_journal")
        conn.commit()

    def _migrate_old_pj(self, cursor, conn):
        """Convert old flat purchase_journal rows to header+lines if needed."""
        cursor.execute("PRAGMA table_info(purchase_journal)")
        cols = {c[1] for c in cursor.fetchall()}
        if 'net_amount' not in cols:
            return

        cursor.execute("SELECT COUNT(*) FROM purchase_journal_lines")
        if cursor.fetchone()[0] > 0:
            self._drop_old_pj_columns(cursor, conn)
            return

        cursor.execute("""
            SELECT id, date, payee_name, reference_no, tin, branch_code,
                   net_amount, input_vat, gross_amount,
                   account_description, account_code, debit, credit, particulars
            FROM purchase_journal
        """)
        old_rows = cursor.fetchall()
        if not old_rows:
            self._drop_old_pj_columns(cursor, conn)
            return

        def _find_code(suffix):
            cursor.execute("SELECT account_code, account_description FROM chart_of_accounts")
            for row in cursor.fetchall():
                if _numeric_suffix(row[0]) == suffix:
                    return row[0], row[1]
            return suffix, suffix

        ap_code,  ap_desc  = _find_code(_AP_NUM)
        vat_code, vat_desc = _find_code(_INPUT_VAT_NUM)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS _pj_headers_temp (
                id INTEGER PRIMARY KEY,
                date DATE NOT NULL,
                payee_name TEXT,
                reference_no TEXT NOT NULL,
                tin TEXT,
                branch_code TEXT,
                particulars TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )""")

        # Group by reference to avoid duplicate headers
        from collections import defaultdict
        ref_groups = defaultdict(list)
        for row in old_rows:
            key = (row[1], row[3])  # date, reference_no
            ref_groups[key].append(row)

        seen_ids = set()
        for (date, ref), rows in ref_groups.items():
            first = rows[0]
            old_id      = first[0]
            payee       = first[2]
            tin         = first[4]
            branch      = first[5]
            gross       = float(first[8] or 0)
            input_vat   = float(first[7] or 0)
            particulars = first[13]

            if old_id not in seen_ids:
                cursor.execute("""
                    INSERT INTO _pj_headers_temp
                    (id, date, payee_name, reference_no, tin, branch_code, particulars)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (old_id, date, payee, ref, tin, branch, particulars))
                seen_ids.add(old_id)

                # AP credit line
                if gross > 0:
                    cursor.execute("""
                        INSERT INTO purchase_journal_lines
                        (journal_id, account_code, account_description, debit, credit)
                        VALUES (?, ?, ?, 0, ?)
                    """, (old_id, ap_code, ap_desc, gross))

                # Input VAT debit line
                if input_vat > 0:
                    cursor.execute("""
                        INSERT INTO purchase_journal_lines
                        (journal_id, account_code, account_description, debit, credit)
                        VALUES (?, ?, ?, ?, 0)
                    """, (old_id, vat_code, vat_desc, input_vat))

            # Each row = one debit line
            for row in rows:
                ac   = row[10] or ''
                ad   = row[9]  or ''
                debit = float(row[11] or row[6] or 0)
                if debit > 0 and ac:
                    cursor.execute("""
                        INSERT INTO purchase_journal_lines
                        (journal_id, account_code, account_description, debit, credit)
                        VALUES (?, ?, ?, ?, 0)
                    """, (old_id, ac, ad, debit))

        cursor.execute("DROP TABLE purchase_journal")
        cursor.execute("ALTER TABLE _pj_headers_temp RENAME TO purchase_journal")
        conn.commit()

    def _drop_old_sj_columns(self, cursor, conn):
        pass  # SQLite can't drop columns easily; old columns are harmless if present

    def _drop_old_pj_columns(self, cursor, conn):
        pass

    # ------------------------------------------------------------------
    # COA seeding
    # ------------------------------------------------------------------

    def _seed_default_coa(self):
        conn = self.get_connection()
        cursor = conn.cursor()
        default_accounts = [
            ('COA-1010','CASH IN BANK','Debit'),('COA-1020','CASH ON HAND','Debit'),
            ('COA-1021','PETTY CASH FUND','Debit'),('COA-1110','ACCOUNTS RECEIVABLE','Debit'),
            ('COA-1120','OTHER RECEIVABLES','Debit'),('COA-1130','ADVANCES TO SUPPLIERS','Debit'),
            ('COA-1140','ADVANCES TO EMPLOYEES','Debit'),('COA-1150','ADVANCES FOR LIQUIDATION','Debit'),
            ('COA-1210','INVENTORIES','Debit'),('COA-1310','PREPAID TAXES WITHHELD','Debit'),
            ('COA-1320','INPUT VAT','Debit'),('COA-1330','PREPAID EXPENSES','Debit'),
            ('COA-1610','COMMUNICATION EQUIPMENT','Debit'),('COA-1611','ACCUM DEP - COMMUNICATION EQUIPMENT','Credit'),
            ('COA-1620','OFFICE EQUIPMENT','Debit'),('COA-1621','ACCUM DEP - OFFICE EQUIPMENT','Credit'),
            ('COA-1630','TRANSPORTATION EQUIPMENT','Debit'),('COA-1631','ACCUM DEP - TRANSPORTATION EQUIPMENT','Credit'),
            ('COA-1640','COMPUTER EQUIPMENT','Debit'),('COA-1641','ACCUM DEP - COMPUTER EQUIPMENT','Credit'),
            ('COA-1650','FURNITURE AND FIXTURE','Debit'),('COA-1651','ACCUM DEP - FURNITURE AND FIXTURE','Credit'),
            ('COA-1660','LEASEHOLD IMPROVEMENT','Debit'),('COA-1661','ACCUM DEP - LEASEHOLD IMPROVEMENT','Credit'),
            ('COA-1670','COMPUTER SOFTWARE','Debit'),('COA-1671','ACCUM AMORT - COMPUTER SOFTWARE','Credit'),
            ('COA-1710','RENT DEPOSIT','Debit'),('COA-1720','SECURITY DEPOSIT','Debit'),
            ('COA-2010','ACCOUNTS PAYABLE','Credit'),('COA-2020','ACCOUNTS PAYABLE - OTHERS','Credit'),
            ('COA-2110','PHILHEALTH PREMIUM PAYABLE','Credit'),('COA-2120','SSS PREMIUM PAYABLE','Credit'),
            ('COA-2130','HDMF PREMIUM PAYABLE','Credit'),('COA-2140','SSS LOAN PAYABLE','Credit'),
            ('COA-2150','HDMF LOAN PAYABLE','Credit'),('COA-2210','OUTPUT VAT PAYABLE','Credit'),
            ('COA-2220','WITHHOLDING TAX PAYABLE - COMPENSATION','Credit'),('COA-2230','INCOME TAX PAYABLE','Credit'),
            ('COA-2310','ACCRUED EXPENSES','Credit'),('COA-2320','ACCRUED SALARIES','Credit'),
            ('COA-3000','OWNER, CAPITAL','Credit'),('COA-3100','OWNER, DRAWING','Debit'),
            ('COA-3999','INCOME SUMMARY','Credit'),('COA-4010','SALES - GOODS','Credit'),
            ('COA-4020','SALES - SERVICES','Credit'),('COA-4610','SALES DISCOUNT','Debit'),
            ('COA-4620','SALES RETURNS','Debit'),('COA-4630','SALES ALLOWANCES','Debit'),
            ('COA-5010','COST OF SALES','Debit'),('COA-5020','PURCHASES','Debit'),
            ('COA-5030','DIRECT LABOR','Debit'),('COA-5040','FACTORY OVERHEAD','Debit'),
            ('COA-6010','SALARIES EXPENSES - BASIC PAY','Debit'),('COA-6011','SALARIES EXPENSES - OVERTIME','Debit'),
            ('COA-6012','SALARIES EXPENSES - DE MINIMIS','Debit'),('COA-6020','GOVERNMENT CONTRIBUTION EXPENSES','Debit'),
            ('COA-6021','PHILHEALTH EXPENSE - EMPLOYER SHARE','Debit'),('COA-6022','SSS EXPENSE - EMPLOYER SHARE','Debit'),
            ('COA-6023','HDMF EXPENSE - EMPLOYER SHARE','Debit'),('COA-6041','UNIFORM','Debit'),
            ('COA-6060','MARKETING EXPENSES','Debit'),('COA-6110','DEPRECIATION EXPENSE','Debit'),
            ('COA-6160','TRANSPORTATION AND TRAVEL','Debit'),('COA-6161','GASOLINE AND OIL','Debit'),
            ('COA-6210','INSURANCE EXPENSE','Debit'),('COA-6260','PROFESSIONAL FEES','Debit'),
            ('COA-6310','RENTAL EXPENSES','Debit'),('COA-6320','CUSA EXPENSE','Debit'),
            ('COA-6360','COMMUNICATION EXPENSE','Debit'),('COA-6410','UTILITIES','Debit'),
            ('COA-6460','STATIONERY AND SUPPLIES','Debit'),('COA-6461','TOOLS AND EQUIPMENT','Debit'),
            ('COA-6510','TAXES AND LICENSES','Debit'),('COA-6540','FINES AND PENALTIES','Debit'),
            ('COA-6560','SUBSCRIPTION EXPENSE','Debit'),('COA-6610','REPRESENTATION EXPENSE','Debit'),
            ('COA-6660','REPAIRS AND MAINTENANCE','Debit'),('COA-6710','TRAININGS AND SEMINARS','Debit'),
            ('COA-6910','MISCELLANEOUS EXPENSE','Debit'),('COA-6911','SERVICE FEES','Debit'),
            ('COA-6912','MEALS','Debit'),('COA-6913','NOTARIAL EXPENSES','Debit'),
            ('COA-7010','OTHER INCOME','Credit'),('COA-7020','MISCELLANEOUS INCOME','Credit'),
            ('COA-7030','COMMISSION INCOME','Credit'),('COA-7040','INTEREST INCOME - BANK DEPOSIT','Credit'),
            ('COA-9010','PROVISION FOR INCOME TAX - CURRENT','Debit'),
            ('COA-9020','PROVISION FOR INCOME TAX - DEFERRED','Debit'),
        ]
        cursor.executemany(
            'INSERT OR IGNORE INTO chart_of_accounts (account_code, account_description, normal_balance) VALUES (?,?,?)',
            default_accounts)
        conn.commit()

    def _import_coa_from_xlsx(self, xlsx_path: str) -> tuple:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return 0, ["openpyxl not installed"]
        try:
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        except Exception as e:
            return 0, [f"Cannot open file: {e}"]
        ws = wb.active
        DATA_START = 2
        col_map = {}
        for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            row_strs = [str(v).strip().lower() if v else '' for v in row]
            if any('account code' in s or 'account_code' in s for s in row_strs):
                for ci, val in enumerate(row):
                    if val is None: continue
                    s = str(val).strip().lower()
                    if 'account code' in s or 'account_code' in s: col_map['code'] = ci
                    elif 'description' in s: col_map['desc'] = ci
                    elif 'debit' in s or 'credit' in s or 'normal' in s: col_map['nb'] = ci
                DATA_START = ri + 1
                break
        if 'code' not in col_map or 'desc' not in col_map:
            wb.close()
            return 0, ["Header row not found."]
        has_nb = 'nb' in col_map
        def _infer(code, desc):
            desc_upper = desc.upper()
            if any(kw in desc_upper for kw in ('ACCUM DEP','ACCUM. DEP','ACCUMULATED DEP','ACCUM AMORT')): return 'Credit'
            if any(kw in desc_upper for kw in ('DRAWING','DRAWINGS')): return 'Debit'
            digits = ''.join(ch for ch in code if ch.isdigit())
            if not digits: return 'Debit'
            first = digits[0]
            if first in ('2','3','4','7'): return 'Credit'
            return 'Debit'
        conn = self.get_connection(); cursor = conn.cursor()
        imported = 0; errors = []
        for rn, row in enumerate(ws.iter_rows(min_row=DATA_START, values_only=True), DATA_START):
            if all(v is None for v in row): continue
            try:
                code = str(row[col_map['code']]).strip() if row[col_map['code']] else ''
                desc = str(row[col_map['desc']]).strip() if row[col_map['desc']] else ''
                if not code or not desc: continue
                if code.lower() in ('none','nan','account code'): continue
                if has_nb and col_map['nb'] < len(row) and row[col_map['nb']]:
                    nb_raw = str(row[col_map['nb']]).strip().upper()
                    nb = 'Credit' if 'CREDIT' in nb_raw else 'Debit'
                else:
                    nb = _infer(code, desc)
                cursor.execute('INSERT OR IGNORE INTO chart_of_accounts (account_code, account_description, normal_balance) VALUES (?,?,?)', (code, desc, nb))
                imported += 1
            except Exception as e:
                errors.append(f"Row {rn}: {e}")
        conn.commit(); wb.close()
        return imported, errors

    # ------------------------------------------------------------------
    # COA CRUD
    # ------------------------------------------------------------------

    def get_all_accounts(self) -> List[Dict]:
        conn = self.get_connection(); cursor = conn.cursor()
        cursor.execute('SELECT * FROM chart_of_accounts ORDER BY account_code')
        return [dict(row) for row in cursor.fetchall()]

    def add_account(self, data: Dict) -> bool:
        try:
            conn = self.get_connection(); cursor = conn.cursor()
            cursor.execute('INSERT INTO chart_of_accounts (account_code, account_description, normal_balance) VALUES (?,?,?)',
                (data.get('account_code'), data.get('account_description'), data.get('normal_balance','Debit')))
            conn.commit(); return True
        except: return False

    def update_account(self, account_id: int, data: Dict) -> bool:
        try:
            conn = self.get_connection(); cursor = conn.cursor()
            cursor.execute('UPDATE chart_of_accounts SET account_code=?, account_description=?, normal_balance=? WHERE id=?',
                (data.get('account_code'), data.get('account_description'), data.get('normal_balance','Debit'), account_id))
            conn.commit(); return True
        except: return False

    def delete_account(self, account_id: int) -> bool:
        try:
            conn = self.get_connection(); cursor = conn.cursor()
            cursor.execute('DELETE FROM chart_of_accounts WHERE id=?', (account_id,))
            conn.commit(); return True
        except: return False

    def import_coa_from_xlsx(self, xlsx_path: str) -> tuple:
        return self._import_coa_from_xlsx(xlsx_path)

    def export_coa_to_xlsx(self, xlsx_path: str) -> tuple:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from datetime import datetime as _dt
        except ImportError:
            return 0, "openpyxl not installed"
        accounts = self.get_all_accounts()
        try:
            wb = Workbook(); ws = wb.active; ws.title = "Chart of Accounts"
            hdr_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
            hdr_fill  = PatternFill('solid', start_color='2F5496')
            hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_font = Font(name='Arial', size=10)
            alt_fill  = PatternFill('solid', start_color='DCE6F1')
            thin      = Side(style='thin', color='B0B0B0')
            border    = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws.merge_cells('A2:C2'); ws['A2'].value = 'CHART OF ACCOUNTS'
            ws['A2'].font = Font(name='Arial', bold=True, size=14)
            ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells('A3:C3'); ws['A3'].value = f'For the Year {_dt.now().year}'
            ws['A3'].font = Font(name='Arial', italic=True, size=11)
            ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
            HEADER_ROW = 5
            ws.row_dimensions[HEADER_ROW].height = 28
            for ci, hdr in enumerate(['Account Code','Account Description','DEBIT/CREDIT'], 1):
                cell = ws.cell(row=HEADER_ROW, column=ci, value=hdr)
                cell.font = hdr_font; cell.fill = hdr_fill
                cell.alignment = hdr_align; cell.border = border
            for ri, acct in enumerate(accounts):
                row_idx = 6 + ri; ws.row_dimensions[row_idx].height = 18
                fill = alt_fill if ri % 2 == 0 else None
                for ci, val in enumerate([acct['account_code'], acct['account_description'], acct.get('normal_balance','Debit').upper()], 1):
                    cell = ws.cell(row=row_idx, column=ci, value=val)
                    cell.font = cell_font; cell.border = border
                    cell.alignment = Alignment(horizontal='center' if ci!=2 else 'left', vertical='center')
                    if fill: cell.fill = fill
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 45
            ws.column_dimensions['C'].width = 16
            ws.freeze_panes = 'A6'
            ws.auto_filter.ref = f'A{HEADER_ROW}:C{HEADER_ROW}'
            wb.save(xlsx_path); return len(accounts), ''
        except Exception as e:
            return 0, str(e)

    # ------------------------------------------------------------------
    # Alphalist
    # ------------------------------------------------------------------

    def get_all_alphalist(self, entry_type: str = None) -> List[Dict]:
        conn = self.get_connection(); cursor = conn.cursor()
        if entry_type and entry_type != 'All List':
            cursor.execute("SELECT * FROM alphalist WHERE entry_type=? OR entry_type='Customer&Vendor' ORDER BY company_name, last_name", (entry_type,))
        else:
            cursor.execute('SELECT * FROM alphalist ORDER BY company_name, last_name')
        return [dict(row) for row in cursor.fetchall()]

    def add_alphalist(self, data: Dict) -> bool:
        try:
            conn = self.get_connection(); cursor = conn.cursor()
            cursor.execute('INSERT INTO alphalist (tin,company_name,first_name,middle_name,last_name,address1,address2,entry_type) VALUES (?,?,?,?,?,?,?,?)',
                (data.get('tin'), data.get('company_name'), data.get('first_name'), data.get('middle_name'),
                 data.get('last_name'), data.get('address1'), data.get('address2'),
                 data.get('entry_type','Customer&Vendor')))
            conn.commit(); return True
        except: return False

    def update_alphalist(self, entry_id: int, data: Dict) -> bool:
        try:
            conn = self.get_connection(); cursor = conn.cursor()
            cursor.execute('UPDATE alphalist SET tin=?,company_name=?,first_name=?,middle_name=?,last_name=?,address1=?,address2=?,entry_type=? WHERE id=?',
                (data.get('tin'), data.get('company_name'), data.get('first_name'), data.get('middle_name'),
                 data.get('last_name'), data.get('address1'), data.get('address2'),
                 data.get('entry_type','Customer&Vendor'), entry_id))
            conn.commit(); return True
        except: return False

    def delete_alphalist(self, entry_id: int) -> bool:
        try:
            conn = self.get_connection(); cursor = conn.cursor()
            cursor.execute('DELETE FROM alphalist WHERE id=?', (entry_id,))
            conn.commit(); return True
        except: return False


    # ------------------------------------------------------------------
    # Sales Journal — new header+lines schema
    # ------------------------------------------------------------------

    def get_sales_journal(self, year: int = None) -> List[Dict]:
        """Returns list of header dicts, each with a 'lines' key."""
        conn = self.get_connection(); cursor = conn.cursor()
        if year is None:
            cursor.execute('SELECT * FROM sales_journal ORDER BY date DESC, id DESC')
        else:
            cursor.execute("SELECT * FROM sales_journal WHERE strftime('%Y',date)=? ORDER BY date DESC, id DESC", (str(year),))
        headers = [dict(row) for row in cursor.fetchall()]
        for hdr in headers:
            cursor.execute('SELECT * FROM sales_journal_lines WHERE journal_id=? ORDER BY id', (hdr['id'],))
            hdr['lines'] = [dict(r) for r in cursor.fetchall()]
            # Compute summary fields from lines for compatibility
            hdr['gross_amount'] = sum(l['debit']  for l in hdr['lines'] if _numeric_suffix(l['account_code']) == _AR_NUM)
            hdr['output_vat']   = sum(l['credit'] for l in hdr['lines'] if _numeric_suffix(l['account_code']) == _OUTPUT_VAT_NUM)
            hdr['goods']        = sum(l['credit'] for l in hdr['lines'] if _numeric_suffix(l['account_code']) == _SALES_GOODS_NUM)
            hdr['services']     = sum(l['credit'] for l in hdr['lines'] if _numeric_suffix(l['account_code']) == _SALES_SERVICE_NUM)
            hdr['net_amount']   = hdr['goods'] + hdr['services']
        return headers

    def add_sales_entry(self, data: Dict) -> bool:
        """
        Accepts either:
        - New format: {'date','customer_name','reference_no','tin','particulars','lines':[...]}
        - Legacy flat format: {'date','customer_name','reference_no','tin','net_amount','output_vat','gross_amount','goods','services','particulars'}
        """
        try:
            conn = self.get_connection(); cursor = conn.cursor()
            cursor.execute('INSERT INTO sales_journal (date,customer_name,reference_no,tin,particulars) VALUES (?,?,?,?,?)',
                (data.get('date'), data.get('customer_name'), data.get('reference_no'),
                 data.get('tin'), data.get('particulars')))
            journal_id = cursor.lastrowid

            lines = data.get('lines')
            if lines:
                for ln in lines:
                    cursor.execute('INSERT INTO sales_journal_lines (journal_id,account_code,account_description,debit,credit) VALUES (?,?,?,?,?)',
                        (journal_id, ln.get('account_code',''), ln.get('account_description',''),
                         ln.get('debit',0), ln.get('credit',0)))
            else:
                # Legacy flat format — reconstruct lines
                self._insert_sj_lines_from_flat(cursor, journal_id, data)

            conn.commit(); return True
        except Exception as e:
            print(f"SJ add error: {e}"); return False

    def _insert_sj_lines_from_flat(self, cursor, journal_id, data):
        """Build lines from old flat fields."""
        def _find(suffix):
            cursor.execute("SELECT account_code, account_description FROM chart_of_accounts")
            for row in cursor.fetchall():
                if _numeric_suffix(row[0]) == suffix:
                    return row[0], row[1]
            return suffix, suffix

        ar_code,    ar_desc    = _find(_AR_NUM)
        vat_code,   vat_desc   = _find(_OUTPUT_VAT_NUM)
        goods_code, goods_desc = _find(_SALES_GOODS_NUM)
        svc_code,   svc_desc   = _find(_SALES_SERVICE_NUM)

        gross   = float(data.get('gross_amount', 0) or 0)
        vat     = float(data.get('output_vat',   0) or 0)
        goods   = float(data.get('goods',         0) or 0)
        services= float(data.get('services',      0) or 0)

        if gross   > 0: cursor.execute('INSERT INTO sales_journal_lines (journal_id,account_code,account_description,debit,credit) VALUES (?,?,?,?,0)', (journal_id, ar_code,    ar_desc,    gross))
        if vat     > 0: cursor.execute('INSERT INTO sales_journal_lines (journal_id,account_code,account_description,debit,credit) VALUES (?,?,?,0,?)', (journal_id, vat_code,   vat_desc,   vat))
        if goods   > 0: cursor.execute('INSERT INTO sales_journal_lines (journal_id,account_code,account_description,debit,credit) VALUES (?,?,?,0,?)', (journal_id, goods_code, goods_desc, goods))
        if services> 0: cursor.execute('INSERT INTO sales_journal_lines (journal_id,account_code,account_description,debit,credit) VALUES (?,?,?,0,?)', (journal_id, svc_code,   svc_desc,   services))

    def update_sales_entry(self, entry_id: int, data: Dict) -> bool:
        try:
            conn = self.get_connection(); cursor = conn.cursor()
            cursor.execute('UPDATE sales_journal SET date=?,customer_name=?,reference_no=?,tin=?,particulars=? WHERE id=?',
                (data.get('date'), data.get('customer_name'), data.get('reference_no'),
                 data.get('tin'), data.get('particulars'), entry_id))
            cursor.execute('DELETE FROM sales_journal_lines WHERE journal_id=?', (entry_id,))
            lines = data.get('lines')
            if lines:
                for ln in lines:
                    cursor.execute('INSERT INTO sales_journal_lines (journal_id,account_code,account_description,debit,credit) VALUES (?,?,?,?,?)',
                        (entry_id, ln.get('account_code',''), ln.get('account_description',''), ln.get('debit',0), ln.get('credit',0)))
            else:
                self._insert_sj_lines_from_flat(cursor, entry_id, data)
            conn.commit(); return True
        except Exception as e:
            print(f"SJ update error: {e}"); return False

    def delete_sales_entry(self, entry_id: int) -> bool:
        try:
            conn = self.get_connection(); cursor = conn.cursor()
            cursor.execute('DELETE FROM sales_journal_lines WHERE journal_id=?', (entry_id,))
            cursor.execute('DELETE FROM sales_journal WHERE id=?', (entry_id,))
            conn.commit(); return True
        except: return False

    # ------------------------------------------------------------------
    # Purchase Journal — new header+lines schema
    # ------------------------------------------------------------------

    def get_purchase_journal(self, year: int = None) -> List[Dict]:
        conn = self.get_connection(); cursor = conn.cursor()
        if year is None:
            cursor.execute('SELECT * FROM purchase_journal ORDER BY date DESC, id DESC')
        else:
            cursor.execute("SELECT * FROM purchase_journal WHERE strftime('%Y',date)=? ORDER BY date DESC, id DESC", (str(year),))
        headers = [dict(row) for row in cursor.fetchall()]
        for hdr in headers:
            cursor.execute('SELECT * FROM purchase_journal_lines WHERE journal_id=? ORDER BY id', (hdr['id'],))
            hdr['lines'] = [dict(r) for r in cursor.fetchall()]
            hdr['gross_amount'] = sum(l['credit'] for l in hdr['lines'] if _numeric_suffix(l['account_code']) == _AP_NUM)
            hdr['input_vat']    = sum(l['debit']  for l in hdr['lines'] if _numeric_suffix(l['account_code']) == _INPUT_VAT_NUM)
            hdr['net_amount']   = hdr['gross_amount'] - hdr['input_vat']
        return headers

    def add_purchase_entry(self, data: Dict) -> bool:
        try:
            conn = self.get_connection(); cursor = conn.cursor()
            cursor.execute('INSERT INTO purchase_journal (date,payee_name,reference_no,tin,branch_code,particulars) VALUES (?,?,?,?,?,?)',
                (data.get('date'), data.get('payee_name'), data.get('reference_no'),
                 data.get('tin'), data.get('branch_code'), data.get('particulars')))
            journal_id = cursor.lastrowid
            lines = data.get('lines')
            if lines:
                for ln in lines:
                    cursor.execute('INSERT INTO purchase_journal_lines (journal_id,account_code,account_description,debit,credit) VALUES (?,?,?,?,?)',
                        (journal_id, ln.get('account_code',''), ln.get('account_description',''), ln.get('debit',0), ln.get('credit',0)))
            else:
                self._insert_pj_lines_from_flat(cursor, journal_id, data)
            conn.commit(); return True
        except Exception as e:
            print(f"PJ add error: {e}"); return False

    def _insert_pj_lines_from_flat(self, cursor, journal_id, data):
        def _find(suffix):
            cursor.execute("SELECT account_code, account_description FROM chart_of_accounts")
            for row in cursor.fetchall():
                if _numeric_suffix(row[0]) == suffix:
                    return row[0], row[1]
            return suffix, suffix
        ap_code,  ap_desc  = _find(_AP_NUM)
        vat_code, vat_desc = _find(_INPUT_VAT_NUM)
        gross  = float(data.get('gross_amount', 0) or 0)
        vat    = float(data.get('input_vat',    0) or 0)
        net    = float(data.get('net_amount',   0) or 0)
        ac     = data.get('account_code',        '')
        ad     = data.get('account_description', '')
        debit  = float(data.get('debit', net)    or 0)
        if gross > 0: cursor.execute('INSERT INTO purchase_journal_lines (journal_id,account_code,account_description,debit,credit) VALUES (?,?,?,0,?)', (journal_id, ap_code,  ap_desc,  gross))
        if vat   > 0: cursor.execute('INSERT INTO purchase_journal_lines (journal_id,account_code,account_description,debit,credit) VALUES (?,?,?,?,0)', (journal_id, vat_code, vat_desc, vat))
        if debit > 0 and ac: cursor.execute('INSERT INTO purchase_journal_lines (journal_id,account_code,account_description,debit,credit) VALUES (?,?,?,?,0)', (journal_id, ac, ad, debit))

    def update_purchase_entry(self, entry_id: int, data: Dict) -> bool:
        try:
            conn = self.get_connection(); cursor = conn.cursor()
            cursor.execute('UPDATE purchase_journal SET date=?,payee_name=?,reference_no=?,tin=?,branch_code=?,particulars=? WHERE id=?',
                (data.get('date'), data.get('payee_name'), data.get('reference_no'),
                 data.get('tin'), data.get('branch_code'), data.get('particulars'), entry_id))
            cursor.execute('DELETE FROM purchase_journal_lines WHERE journal_id=?', (entry_id,))
            lines = data.get('lines')
            if lines:
                for ln in lines:
                    cursor.execute('INSERT INTO purchase_journal_lines (journal_id,account_code,account_description,debit,credit) VALUES (?,?,?,?,?)',
                        (entry_id, ln.get('account_code',''), ln.get('account_description',''), ln.get('debit',0), ln.get('credit',0)))
            else:
                self._insert_pj_lines_from_flat(cursor, entry_id, data)
            conn.commit(); return True
        except Exception as e:
            print(f"PJ update error: {e}"); return False

    def delete_purchase_entry(self, entry_id: int) -> bool:
        try:
            conn = self.get_connection(); cursor = conn.cursor()
            cursor.execute('DELETE FROM purchase_journal_lines WHERE journal_id=?', (entry_id,))
            cursor.execute('DELETE FROM purchase_journal WHERE id=?', (entry_id,))
            conn.commit(); return True
        except: return False

    # ------------------------------------------------------------------
    # CDJ / CRJ / GJ (unchanged)
    # ------------------------------------------------------------------

    def _get_journal(self, table, year):
        conn = self.get_connection(); cursor = conn.cursor()
        if year is None:
            cursor.execute(f'SELECT * FROM {table} ORDER BY date DESC')
        else:
            cursor.execute(f"SELECT * FROM {table} WHERE strftime('%Y',date)=? ORDER BY date DESC", (str(year),))
        return [dict(row) for row in cursor.fetchall()]

    def get_cash_disbursement_journal(self, year=None): return self._get_journal('cash_disbursement_journal', year)
    def get_cash_receipts_journal(self, year=None):     return self._get_journal('cash_receipts_journal', year)
    def get_general_journal(self, year=None):           return self._get_journal('general_journal', year)

    def _add_simple(self, table, data):
        try:
            conn = self.get_connection(); cursor = conn.cursor()
            cursor.execute(f'INSERT INTO {table} (date,reference_no,particulars,account_code,account_description,debit,credit) VALUES (?,?,?,?,?,?,?)',
                (data.get('date'), data.get('reference_no'), data.get('particulars'),
                 data.get('account_code'), data.get('account_description'),
                 data.get('debit',0), data.get('credit',0)))
            conn.commit(); return True
        except Exception as e:
            print(f"Error: {e}"); return False

    def _update_simple(self, table, entry_id, data):
        try:
            conn = self.get_connection(); cursor = conn.cursor()
            cursor.execute(f'UPDATE {table} SET date=?,reference_no=?,particulars=?,account_description=?,account_code=?,debit=?,credit=? WHERE id=?',
                (data.get('date'), data.get('reference_no'), data.get('particulars'),
                 data.get('account_description'), data.get('account_code'),
                 data.get('debit',0), data.get('credit',0), entry_id))
            conn.commit(); return True
        except: return False

    def _delete_simple(self, table, entry_id):
        try:
            conn = self.get_connection(); cursor = conn.cursor()
            cursor.execute(f'DELETE FROM {table} WHERE id=?', (entry_id,))
            conn.commit(); return True
        except: return False

    def add_cash_disbursement_entry(self, data):    return self._add_simple('cash_disbursement_journal', data)
    def update_cash_disbursement_entry(self, i, d): return self._update_simple('cash_disbursement_journal', i, d)
    def delete_cash_disbursement_entry(self, i):    return self._delete_simple('cash_disbursement_journal', i)
    def add_cash_receipts_entry(self, data):        return self._add_simple('cash_receipts_journal', data)
    def update_cash_receipts_entry(self, i, d):     return self._update_simple('cash_receipts_journal', i, d)
    def delete_cash_receipts_entry(self, i):        return self._delete_simple('cash_receipts_journal', i)
    def add_general_journal_entry(self, data):      return self._add_simple('general_journal', data)
    def update_general_journal_entry(self, i, d):   return self._update_simple('general_journal', i, d)
    def delete_general_journal_entry(self, i):      return self._delete_simple('general_journal', i)

    # ------------------------------------------------------------------
    # General Ledger
    # ------------------------------------------------------------------

    def get_general_ledger(self, account_code: str, date_from: str = None, date_to: str = None) -> Dict:
        conn = self.get_connection(); cursor = conn.cursor()
        entries = []

        def _to_sql(d):
            try: return datetime.strptime(d, "%m/%d/%Y").strftime("%Y-%m-%d")
            except: return d

        if date_from and date_to:
            sql_from = _to_sql(date_from); sql_to = _to_sql(date_to)
            date_filter = "AND strftime('%Y-%m-%d',substr(date,7,4)||'-'||substr(date,1,2)||'-'||substr(date,4,2)) BETWEEN ? AND ?"
            date_params = (sql_from, sql_to)
        elif date_to:
            sql_to = _to_sql(date_to)
            date_filter = "AND strftime('%Y-%m-%d',substr(date,7,4)||'-'||substr(date,1,2)||'-'||substr(date,4,2)) <= ?"
            date_params = (sql_to,)
        elif date_from:
            sql_from = _to_sql(date_from)
            date_filter = "AND strftime('%Y-%m-%d',substr(date,7,4)||'-'||substr(date,1,2)||'-'||substr(date,4,2)) >= ?"
            date_params = (sql_from,)
        else:
            date_filter = ""; date_params = ()

        # SJ lines
        cursor.execute(f"""
            SELECT sj.date, sj.reference_no, sj.customer_name as particulars,
                   sjl.debit, sjl.credit, 'SJ' as source
            FROM sales_journal sj
            JOIN sales_journal_lines sjl ON sjl.journal_id = sj.id
            WHERE sjl.account_code = ? {date_filter}
            ORDER BY sj.date
        """, (account_code,) + date_params)
        entries.extend(cursor.fetchall())

        # PJ lines
        cursor.execute(f"""
            SELECT pj.date, pj.reference_no, pj.payee_name as particulars,
                   pjl.debit, pjl.credit, 'PJ' as source
            FROM purchase_journal pj
            JOIN purchase_journal_lines pjl ON pjl.journal_id = pj.id
            WHERE pjl.account_code = ? {date_filter}
            ORDER BY pj.date
        """, (account_code,) + date_params)
        entries.extend(cursor.fetchall())

        # CDJ, CRJ, GJ
        for table, tag in [('cash_disbursement_journal','CDJ'),('cash_receipts_journal','CRJ'),('general_journal','GJ')]:
            cursor.execute(f"""
                SELECT date, reference_no, particulars, debit, credit, '{tag}' as source
                FROM {table} WHERE account_code=? {date_filter} ORDER BY date
            """, (account_code,) + date_params)
            entries.extend(cursor.fetchall())

        result = sorted([dict(row) for row in entries], key=lambda x: x.get('date',''))

        cursor.execute("SELECT normal_balance FROM chart_of_accounts WHERE account_code=?", (account_code,))
        row = cursor.fetchone()
        return {'account_code': account_code, 'normal_balance': row['normal_balance'] if row else 'Debit', 'entries': result}

    # ------------------------------------------------------------------
    # Trial Balance
    # ------------------------------------------------------------------

    def get_trial_balance(self, date_from: str = None, date_to: str = None) -> List[Dict]:
        conn = self.get_connection(); cursor = conn.cursor()
        cursor.execute('SELECT account_code, account_description, normal_balance FROM chart_of_accounts ORDER BY account_code')
        accounts = cursor.fetchall()

        def _to_sql(d):
            try: return datetime.strptime(d, "%m/%d/%Y").strftime("%Y-%m-%d")
            except: return d

        if date_from and date_to:
            sql_from = _to_sql(date_from); sql_to = _to_sql(date_to)
            df = "AND strftime('%Y-%m-%d',substr(date,7,4)||'-'||substr(date,1,2)||'-'||substr(date,4,2)) BETWEEN ? AND ?"
            dp = (sql_from, sql_to)
        elif date_to:
            sql_to = _to_sql(date_to)
            df = "AND strftime('%Y-%m-%d',substr(date,7,4)||'-'||substr(date,1,2)||'-'||substr(date,4,2)) <= ?"
            dp = (sql_to,)
        elif date_from:
            sql_from = _to_sql(date_from)
            df = "AND strftime('%Y-%m-%d',substr(date,7,4)||'-'||substr(date,1,2)||'-'||substr(date,4,2)) >= ?"
            dp = (sql_from,)
        else:
            df = ""; dp = ()

        trial_balance = []
        for account in accounts:
            ac = account['account_code']
            td = tc = 0

            # SJ lines
            cursor.execute(f"SELECT COALESCE(SUM(sjl.debit),0), COALESCE(SUM(sjl.credit),0) FROM sales_journal sj JOIN sales_journal_lines sjl ON sjl.journal_id=sj.id WHERE sjl.account_code=? {df}", (ac,)+dp)
            r = cursor.fetchone(); td += r[0]; tc += r[1]

            # PJ lines
            cursor.execute(f"SELECT COALESCE(SUM(pjl.debit),0), COALESCE(SUM(pjl.credit),0) FROM purchase_journal pj JOIN purchase_journal_lines pjl ON pjl.journal_id=pj.id WHERE pjl.account_code=? {df}", (ac,)+dp)
            r = cursor.fetchone(); td += r[0]; tc += r[1]

            # CDJ, CRJ, GJ
            for tbl in ('cash_disbursement_journal','cash_receipts_journal','general_journal'):
                cursor.execute(f"SELECT COALESCE(SUM(debit),0), COALESCE(SUM(credit),0) FROM {tbl} WHERE account_code=? {df}", (ac,)+dp)
                r = cursor.fetchone(); td += r[0]; tc += r[1]

            balance = td - tc
            if balance != 0:
                trial_balance.append({
                    'account_code':        ac,
                    'account_description': account['account_description'],
                    'normal_balance':      account['normal_balance'] or 'Debit',
                    'amount':              balance,
                })
        return trial_balance

    def close(self):
        if self.connection:
            self.connection.close()
            self.connection = None