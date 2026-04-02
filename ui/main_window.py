import sys
import os
import shutil
from datetime import datetime

from PySide6.QtWidgets import (
    QMainWindow, QStackedWidget, QStatusBar, QWidget, QVBoxLayout,
    QHBoxLayout, QLabel, QMenuBar, QMenu, QMessageBox, QFileDialog,
    QProgressDialog, QInputDialog, QLineEdit, QDialog, QTextBrowser,
    QDialogButtonBox, QApplication, QFormLayout, QSpinBox,
    QComboBox, QGroupBox,
)
from PySide6.QtCore import Qt, QObject, QEvent
from PySide6.QtGui  import QIcon, QKeySequence, QFont, QAction, QShortcut, QGuiApplication

from database.db_manager              import DatabaseManager
from resources.file_paths             import get_resource
from ui.sidebar.sidebar_widget        import (SidebarWidget,
                                              PAGE_DASHBOARD,
                                              PAGE_COA, PAGE_ALPHALIST,
                                              PAGE_SJ,  PAGE_PJ,
                                              PAGE_CDJ, PAGE_CRJ,
                                              PAGE_GJ,  PAGE_GL,
                                              PAGE_TB,  PAGE_FS,
                                              PAGE_SETTINGS)
from ui.widgets.dashboard_widget          import DashboardWidget
from ui.widgets.coa_widget                import COAWidget
from ui.widgets.alphalist_widget          import AlphalistWidget
from ui.widgets.sales_journal_widget      import SalesJournalWidget
from ui.widgets.purchase_journal_widget   import PurchaseJournalWidget
from ui.widgets.cash_disbursement_widget  import CashDisbursementWidget
from ui.widgets.cash_receipts_widget      import CashReceiptsWidget
from ui.widgets.general_journal_widget    import GeneralJournalWidget
from ui.widgets.general_ledger_widget     import GeneralLedgerWidget
from ui.widgets.trial_balance_widget      import TrialBalanceWidget
from ui.widgets.financial_statements_widget import FinancialStatementsWidget


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _saves_dir() -> str:
    if getattr(sys, 'frozen', False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = os.path.join(base, 'data', 'saves')
    os.makedirs(path, exist_ok=True)
    return path


def _backups_dir() -> str:
    if getattr(sys, 'frozen', False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = os.path.join(base, 'data', 'backups')
    os.makedirs(path, exist_ok=True)
    return path


def _recent_file() -> str:
    if getattr(sys, 'frozen', False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    data_dir = os.path.join(base, 'data')
    os.makedirs(data_dir, exist_ok=True)
    return os.path.join(data_dir, 'recent.json')


def _load_recent() -> list:
    import json
    try:
        with open(_recent_file(), 'r') as f:
            items = json.load(f)
        return [p for p in items if os.path.exists(p)]
    except Exception:
        return []


def _save_recent(path: str):
    import json
    items = _load_recent()
    if path in items:
        items.remove(path)
    items.insert(0, path)
    items = items[:8]
    try:
        with open(_recent_file(), 'w') as f:
            json.dump(items, f)
    except Exception:
        pass


def _prefs_path(db_path: str) -> str:
    """Return the sidecar JSON prefs file path for a given .db file."""
    db_dir = os.path.dirname(db_path)
    name   = os.path.splitext(os.path.basename(db_path))[0]
    return os.path.join(db_dir, f'{name}_prefs.json')


# ---------------------------------------------------------------------------
# Settings widget
# ---------------------------------------------------------------------------

class SettingsWidget(QWidget):
    def __init__(self, db_manager: DatabaseManager, apply_theme_fn=None, parent=None):
        super().__init__(parent)
        self.db_manager   = db_manager
        self._apply_theme = apply_theme_fn
        self._dark_mode   = False
        self._setup_ui()

    def _setup_ui(self):
        from PySide6.QtWidgets import QPushButton
        root = QVBoxLayout(self)
        root.setContentsMargins(32, 24, 32, 24)
        root.setSpacing(24)

        title = QLabel('Settings')
        tf = QFont(); tf.setPointSize(18); tf.setBold(True)
        title.setFont(tf)
        root.addWidget(title)

        appear_group  = QGroupBox('Appearance')
        appear_layout = QFormLayout(appear_group)
        self._theme_btn = QPushButton('Switch to Dark Theme')
        self._theme_btn.setFixedWidth(200)
        self._theme_btn.clicked.connect(self._toggle_theme)
        appear_layout.addRow('Theme:', self._theme_btn)
        root.addWidget(appear_group)

        fy_group  = QGroupBox('Fiscal Year')
        fy_layout = QFormLayout(fy_group)
        self.year_spin = QSpinBox()
        self.year_spin.setRange(2000, 2100)
        self.year_spin.setValue(self.db_manager.get_current_year())
        self.year_spin.setFixedWidth(120)
        save_year_btn = QPushButton('Apply')
        save_year_btn.setFixedWidth(80)
        save_year_btn.clicked.connect(self._save_year)
        year_row = QHBoxLayout()
        year_row.addWidget(self.year_spin)
        year_row.addWidget(save_year_btn)
        year_row.addStretch()
        fy_layout.addRow('Current Year:', year_row)
        self.fy_month_combo = QComboBox()
        self.fy_month_combo.addItems([
            'January','February','March','April','May','June',
            'July','August','September','October','November','December'])
        self.fy_month_combo.setFixedWidth(160)
        fy_layout.addRow('Fiscal Year Start:', self.fy_month_combo)
        root.addWidget(fy_group)

        co_group  = QGroupBox('Company')
        co_layout = QFormLayout(co_group)
        self.company_input = QLineEdit()
        self.company_input.setPlaceholderText('e.g. ESQ Company')
        co_layout.addRow('Company Name:', self.company_input)
        self.tin_input = QLineEdit()
        self.tin_input.setPlaceholderText('e.g. 123-456-789-000')
        co_layout.addRow('Company TIN:', self.tin_input)
        save_co_btn = QPushButton('Save Company Info')
        save_co_btn.setFixedWidth(160)
        save_co_btn.clicked.connect(self._save_company)
        co_layout.addRow('', save_co_btn)
        root.addWidget(co_group)

        root.addStretch()
        self._load_prefs()

    def _prefs_path(self) -> str:
        return _prefs_path(self.db_manager.db_path)

    def _load_prefs(self):
        import json
        try:
            with open(self._prefs_path(), 'r') as f:
                prefs = json.load(f)
            self.year_spin.setValue(prefs.get('year', self.db_manager.get_current_year()))
            self.fy_month_combo.setCurrentIndex(prefs.get('fy_start_month', 0))
            self.company_input.setText(prefs.get('company_name', ''))
            self.tin_input.setText(prefs.get('company_tin', ''))
            self._dark_mode = prefs.get('dark_mode', False)
            self._theme_btn.setText(
                'Switch to Light Theme' if self._dark_mode else 'Switch to Dark Theme')
        except Exception:
            pass

    def _save_prefs(self, extra: dict | None = None):
        import json
        prefs = {
            'year':           self.year_spin.value(),
            'fy_start_month': self.fy_month_combo.currentIndex(),
            'company_name':   self.company_input.text().strip(),
            'company_tin':    self.tin_input.text().strip(),
            'dark_mode':      self._dark_mode,
        }
        if extra:
            prefs.update(extra)
        try:
            with open(self._prefs_path(), 'w') as f:
                json.dump(prefs, f, indent=2)
        except Exception as e:
            QMessageBox.warning(self, 'Warning', f'Could not save preferences:\n{e}')
        return prefs

    def _save_year(self):
        prefs = self._save_prefs()
        self.db_manager.set_current_year(prefs['year'])
        QMessageBox.information(self, 'Saved', f"Year set to {prefs['year']}.")

    def _save_company(self):
        self._save_prefs()
        QMessageBox.information(self, 'Saved', 'Company info saved.')

    def _toggle_theme(self):
        self._dark_mode = not self._dark_mode
        self._theme_btn.setText(
            'Switch to Light Theme' if self._dark_mode else 'Switch to Dark Theme')
        if self._apply_theme:
            self._apply_theme(QApplication.instance(),
                              'dark' if self._dark_mode else 'light')
        self._save_prefs()

    def load_data(self):
        self._load_prefs()


# ---------------------------------------------------------------------------
# Main Window
# ---------------------------------------------------------------------------

class MainWindow(QMainWindow):

    def __init__(self, db_manager: DatabaseManager, apply_theme_fn=None):
        super().__init__()
        self.db_manager      = db_manager
        self._apply_theme_fn = apply_theme_fn

        self.setWindowTitle(self._make_title())

        # Clamp initial geometry to screen so window never overflows the display
        screen = QGuiApplication.primaryScreen().availableGeometry()
        w = min(1400, screen.width()  - 20)
        h = min(800,  screen.height() - 60)
        self.setMinimumSize(800, 500)   # hard floor — prevents child widgets from blowing up minimum
        self.setGeometry(
            max(0, (screen.width()  - w) // 2),
            max(0, (screen.height() - h) // 2),
            w, h,
        )

        _save_recent(db_manager.db_path)

        self._create_menu_bar()
        self._create_status_bar()   # must be before _create_layout
        self._create_layout()
        self._setup_shortcuts()
        self._apply_saved_prefs()

    # ------------------------------------------------------------------ title

    def _make_title(self) -> str:
        name = os.path.splitext(os.path.basename(self.db_manager.db_path))[0]
        return f'ESQ Accounting System — {name}'

    # ------------------------------------------------------------------ prefs

    def _prefs_path(self) -> str:
        return _prefs_path(self.db_manager.db_path)

    def _apply_saved_prefs(self):
        import json
        try:
            with open(self._prefs_path(), 'r') as f:
                prefs = json.load(f)
            year = prefs.get('year', datetime.now().year)
            self.db_manager.set_current_year(year)
        except Exception:
            pass

    # ------------------------------------------------------------------ menu

    def _create_menu_bar(self):
        menubar = self.menuBar()

        file_menu = menubar.addMenu('&File')

        load_action = QAction('&Load Account…', self)
        load_action.setShortcut('Ctrl+O')
        load_action.triggered.connect(self._load_account)
        file_menu.addAction(load_action)

        self._recent_menu = QMenu('&Recent Accounts', self)
        file_menu.addMenu(self._recent_menu)
        self._rebuild_recent_menu()
        file_menu.addSeparator()

        close_action = QAction('&Close Account', self)
        close_action.setShortcut('Ctrl+W')
        close_action.triggered.connect(self._close_account)
        file_menu.addAction(close_action)

        rename_action = QAction('Re&name Account…', self)
        rename_action.triggered.connect(self._rename_account)
        file_menu.addAction(rename_action)
        file_menu.addSeparator()

        backup_action = QAction('&Backup Account', self)
        backup_action.setShortcut('Ctrl+B')
        backup_action.triggered.connect(self._backup_account)
        file_menu.addAction(backup_action)

        restore_action = QAction('Res&tore from Backup…', self)
        restore_action.triggered.connect(self._restore_from_backup)
        file_menu.addAction(restore_action)
        file_menu.addSeparator()

        import_book_action = QAction('&Import Full Book…', self)
        import_book_action.setShortcut('Ctrl+Shift+I')
        import_book_action.triggered.connect(self._import_full_book)
        file_menu.addAction(import_book_action)

        export_book_action = QAction('E&xport Full Book…', self)
        export_book_action.setShortcut('Ctrl+Shift+X')
        export_book_action.triggered.connect(self._export_full_book)
        file_menu.addAction(export_book_action)
        file_menu.addSeparator()

        refresh_action = QAction('&Refresh All', self)
        refresh_action.setShortcut('F5')
        refresh_action.triggered.connect(self._refresh_all)
        file_menu.addAction(refresh_action)
        file_menu.addSeparator()

        exit_action = QAction('E&xit', self)
        exit_action.setShortcut('Ctrl+Q')
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        help_menu = menubar.addMenu('&Help')

        manual_action = QAction('&User Manual', self)
        manual_action.setShortcut('F1')
        manual_action.triggered.connect(self._show_user_manual)
        help_menu.addAction(manual_action)
        help_menu.addSeparator()

        about_action = QAction('&About', self)
        about_action.triggered.connect(self._show_about)
        help_menu.addAction(about_action)

    # ------------------------------------------------------------------ layout

    def _create_layout(self):
        container = QWidget()
        h_layout  = QHBoxLayout(container)
        h_layout.setContentsMargins(0, 0, 0, 0)
        h_layout.setSpacing(0)

        self.sidebar = SidebarWidget()
        self.sidebar.page_requested.connect(self._navigate_to)
        h_layout.addWidget(self.sidebar)

        self.stack = QStackedWidget()
        h_layout.addWidget(self.stack, stretch=1)

        self._dashboard_widget   = DashboardWidget(self.db_manager)
        self._coa_widget         = COAWidget(self.db_manager)
        self._alphalist_widget   = AlphalistWidget(self.db_manager)
        self._sj_widget          = SalesJournalWidget(self.db_manager)
        self._pj_widget          = PurchaseJournalWidget(self.db_manager)
        self._cdj_widget         = CashDisbursementWidget(self.db_manager)
        self._crj_widget         = CashReceiptsWidget(self.db_manager)
        self._gj_widget          = GeneralJournalWidget(self.db_manager)
        self._gl_widget          = GeneralLedgerWidget(self.db_manager)
        self._tb_widget          = TrialBalanceWidget(self.db_manager)
        self._fs_widget          = FinancialStatementsWidget(self.db_manager)
        self._settings_widget    = SettingsWidget(
            self.db_manager, apply_theme_fn=self._apply_theme_fn)

        for w in [
            self._dashboard_widget,
            self._coa_widget,
            self._alphalist_widget,
            self._sj_widget,
            self._pj_widget,
            self._cdj_widget,
            self._crj_widget,
            self._gj_widget,
            self._gl_widget,
            self._tb_widget,
            self._fs_widget,
            self._settings_widget,
        ]:
            self.stack.addWidget(w)

        self.sidebar.theme_toggle_clicked.connect(
            self._settings_widget._toggle_theme)

        self.setCentralWidget(container)
        self._navigate_to(PAGE_DASHBOARD)

    # ------------------------------------------------------------------ navigation

    def _navigate_to(self, page_index: int):
        self.stack.setCurrentIndex(page_index)
        self.sidebar.set_active_page(page_index)
        self._update_status_bar()

    # ------------------------------------------------------------------ shortcuts

    def _setup_shortcuts(self):
        page_shortcuts = {
            'Ctrl+1': PAGE_DASHBOARD,
            'Ctrl+2': PAGE_SJ,
            'Ctrl+3': PAGE_PJ,
            'Ctrl+4': PAGE_CDJ,
            'Ctrl+5': PAGE_CRJ,
            'Ctrl+6': PAGE_GJ,
            'Ctrl+7': PAGE_GL,
            'Ctrl+8': PAGE_TB,
            'Ctrl+9': PAGE_FS,
            'Ctrl+0': PAGE_SETTINGS,
        }
        for key, page in page_shortcuts.items():
            QShortcut(QKeySequence(key), self).activated.connect(
                lambda p=page: self._navigate_to(p))

        QShortcut(QKeySequence('F5'), self).activated.connect(self._refresh_all)
        QShortcut(QKeySequence('F1'), self).activated.connect(self._show_user_manual)
        QShortcut(QKeySequence('Ctrl+Tab'), self).activated.connect(self._next_page)
        QShortcut(QKeySequence('Ctrl+Shift+Tab'), self).activated.connect(self._prev_page)

    def _next_page(self):
        n = self.stack.count()
        self._navigate_to((self.stack.currentIndex() + 1) % n)

    def _prev_page(self):
        n = self.stack.count()
        self._navigate_to((self.stack.currentIndex() - 1) % n)

    # ------------------------------------------------------------------ status bar

    def _create_status_bar(self):
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self._update_status_bar()

    def _update_status_bar(self):
        db_name = os.path.splitext(os.path.basename(self.db_manager.db_path))[0]
        year    = self.db_manager.get_current_year()
        self.status_bar.showMessage(f'Ready  |  {db_name}  |  Year: {year}')

    # ------------------------------------------------------------------ refresh

    def _refresh_all(self):
        for i in range(self.stack.count()):
            w = self.stack.widget(i)
            if hasattr(w, 'load_data'):
                w.load_data()
        self._update_status_bar()
        db_name = os.path.splitext(os.path.basename(self.db_manager.db_path))[0]
        self.status_bar.showMessage(
            f'All data refreshed  |  {db_name}  |  Year: {self.db_manager.get_current_year()}', 4000)

    # ------------------------------------------------------------------ recent menu

    def _rebuild_recent_menu(self):
        self._recent_menu.clear()
        items   = _load_recent()
        current = self.db_manager.db_path if self.db_manager else ''
        items   = [p for p in items if p != current]

        if not items:
            empty = QAction('(No recent accounts)', self)
            empty.setEnabled(False)
            self._recent_menu.addAction(empty)
            return

        for path in items:
            name   = os.path.splitext(os.path.basename(path))[0]
            action = QAction(name, self)
            action.setStatusTip(path)
            action.triggered.connect(
                lambda checked=False, p=path: self._open_recent(p))
            self._recent_menu.addAction(action)

        self._recent_menu.addSeparator()
        clear_action = QAction('Clear Recent List', self)
        clear_action.triggered.connect(self._clear_recent)
        self._recent_menu.addAction(clear_action)

    def _open_recent(self, path: str):
        if not os.path.exists(path):
            QMessageBox.warning(self, 'File Not Found', f'Could not find:\n{path}')
            self._rebuild_recent_menu()
            return
        self._switch_to_db(path)

    def _clear_recent(self):
        import json
        try:
            with open(_recent_file(), 'w') as f:
                json.dump([], f)
        except Exception:
            pass
        self._rebuild_recent_menu()

    # ------------------------------------------------------------------ File actions

    def _load_account(self):
        path, _ = QFileDialog.getOpenFileName(
            self, 'Load Account', _saves_dir(),
            'Database Files (*.db);;All Files (*)')
        if path:
            self._switch_to_db(path)

    def _switch_to_db(self, path: str):
        self.db_manager.close()
        new_db = DatabaseManager(path)
        new_db.initialize_database()
        self.db_manager = new_db
        _save_recent(path)

        old = self.centralWidget()
        self._create_layout()
        if old:
            old.deleteLater()

        self.setWindowTitle(self._make_title())
        self._apply_saved_prefs()
        self._rebuild_recent_menu()
        self.status_bar.showMessage(f'Loaded: {os.path.basename(path)}', 4000)

    def _close_account(self):
        reply = QMessageBox.question(
            self, 'Close Account',
            'Close the current account and return to the startup screen?',
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        self.db_manager.close()

        from ui.startup_dialog import StartupDialog
        startup = StartupDialog()
        if startup.exec() != StartupDialog.Accepted or not startup.db_path:
            QApplication.quit()
            return

        new_db = DatabaseManager(startup.db_path)
        new_db.initialize_database(
            coa_xlsx_path=startup.coa_xlsx_path,
            use_default_coa=(startup.coa_xlsx_path is None))
        self.db_manager = new_db
        _save_recent(startup.db_path)

        old = self.centralWidget()
        self._create_layout()
        if old:
            old.deleteLater()

        self.setWindowTitle(self._make_title())
        self._apply_saved_prefs()
        self._rebuild_recent_menu()
        self.status_bar.showMessage('Account loaded.', 3000)

    def _rename_account(self):
        current_path = self.db_manager.db_path
        current_name = os.path.splitext(os.path.basename(current_path))[0]
        new_name, ok = QInputDialog.getText(
            self, 'Rename Account',
            'Enter a new name for this account:',
            QLineEdit.Normal, current_name)
        if not ok or not new_name.strip():
            return
        safe = ''.join(
            c for c in new_name.strip()
            if c.isalnum() or c in (' ', '-', '_')).strip()
        if not safe:
            QMessageBox.warning(self, 'Invalid Name', 'Please enter a valid name.')
            return
        folder   = os.path.dirname(current_path)
        new_path = os.path.join(folder, safe + '.db')
        if os.path.exists(new_path) and new_path != current_path:
            QMessageBox.warning(self, 'Name Taken', f'"{safe}.db" already exists.')
            return
        try:
            self.db_manager.close()
            os.rename(current_path, new_path)
            old_prefs = os.path.join(folder, f'{current_name}_prefs.json')
            new_prefs = os.path.join(folder, f'{safe}_prefs.json')
            if os.path.exists(old_prefs):
                os.rename(old_prefs, new_prefs)
        except OSError as e:
            QMessageBox.critical(self, 'Rename Failed', str(e))
            return
        self.db_manager = DatabaseManager(new_path)
        self.db_manager.get_connection()
        self._apply_saved_prefs()
        _save_recent(new_path)
        self._rebuild_recent_menu()
        self.setWindowTitle(self._make_title())
        self.status_bar.showMessage(f'Renamed to "{safe}".', 3000)

    # ------------------------------------------------------------------ Backup

    def _backup_account(self):
        src  = self.db_manager.db_path
        name = os.path.splitext(os.path.basename(src))[0]
        ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
        dst  = os.path.join(_backups_dir(), f'{name}_backup_{ts}.db')
        try:
            self.db_manager.close()
            shutil.copy2(src, dst)
            self.db_manager.get_connection()
        except Exception as e:
            QMessageBox.critical(self, 'Backup Failed', str(e))
            return
        self.status_bar.showMessage(f'Backup saved: {os.path.basename(dst)}', 5000)
        QMessageBox.information(self, 'Backup Successful', f'Backup saved to:\n{dst}')

    def _restore_from_backup(self):
        path, _ = QFileDialog.getOpenFileName(
            self, 'Restore from Backup', _backups_dir(),
            'Database Files (*.db);;All Files (*)')
        if not path:
            return
        reply = QMessageBox.warning(
            self, 'Confirm Restore',
            f'Replace current account with:\n{os.path.basename(path)}\n\n'
            f'All unsaved changes will be lost. Continue?',
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply != QMessageBox.Yes:
            return
        dst = self.db_manager.db_path
        try:
            self.db_manager.close()
            shutil.copy2(path, dst)
        except Exception as e:
            QMessageBox.critical(self, 'Restore Failed', str(e))
            return
        self.db_manager = DatabaseManager(dst)
        self.db_manager.initialize_database()
        self._apply_saved_prefs()
        old = self.centralWidget()
        self._create_layout()
        if old:
            old.deleteLater()
        self.status_bar.showMessage('Account restored from backup.', 4000)
        QMessageBox.information(self, 'Restore Successful', 'Account has been restored.')

    # ------------------------------------------------------------------ Import/Export

    def _import_full_book(self):
        from resources.file_paths import get_import_dir
        path, _ = QFileDialog.getOpenFileName(
            self, 'Import Full Book', get_import_dir(''),
            'Excel Files (*.xlsx)')
        if not path:
            return
        reply = QMessageBox.question(
            self, 'Import Full Book',
            f'Import all supported sheets from:\n{os.path.basename(path)}\n\n'
            f'Duplicates are skipped. Continue?',
            QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
        if reply != QMessageBox.Yes:
            return
        progress = QProgressDialog('Importing…', None, 0, 0, self)
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)
        progress.show()
        QApplication.processEvents()
        try:
            from ui.fullbook_importer import import_full_book, build_summary_message
            def _on_progress(sheet_name: str):
                progress.setLabelText(f'Importing: {sheet_name}…')
                QApplication.processEvents()
            results = import_full_book(path, self.db_manager, _on_progress)
        except Exception as exc:
            progress.close()
            QMessageBox.critical(self, 'Import Failed', str(exc))
            return
        progress.close()
        self._refresh_all()
        QMessageBox.information(self, 'Import Complete',
                                build_summary_message(results))

    def _export_full_book(self):
        from resources.file_paths import get_io_dir
        db_name = os.path.splitext(os.path.basename(self.db_manager.db_path))[0]
        year    = self.db_manager.get_current_year()
        path, _ = QFileDialog.getSaveFileName(
            self, 'Export Full Book',
            os.path.join(get_io_dir('Full Book'), f'{db_name}_{year}_books.xlsx'),
            'Excel Files (*.xlsx)')
        if not path:
            return

        try:
            from openpyxl import Workbook
        except ImportError:
            QMessageBox.critical(self, 'Missing Library',
                                 'openpyxl is required.\nInstall with: pip install openpyxl')
            return

        progress = QProgressDialog('Exporting…', None, 0, 0, self)
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.show()
        QApplication.processEvents()

        try:
            from utils.export_utils import export_to_xls

            wb = Workbook()
            total_rows = 0
            first_sheet = True

            JOURNALS = [
                ('Sales Journal',             self.db_manager.get_sales_journal,
                 [('Date','date'),('Customer','customer_name'),('Reference','reference_no'),
                  ('TIN','tin'),('Particulars','particulars'),
                  ('Account Description','account_description'),('Account Code','account_code'),
                  ('Debit','debit'),('Credit','credit')],
                 'customer_name'),
                ('Purchase Journal',          self.db_manager.get_purchase_journal,
                 [('Date','date'),('Payee','payee_name'),('Reference','reference_no'),
                  ('TIN','tin'),('Branch','branch_code'),('Particulars','particulars'),
                  ('Account Description','account_description'),('Account Code','account_code'),
                  ('Debit','debit'),('Credit','credit')],
                 'payee_name'),
                ('Cash Disbursement Journal', self.db_manager.get_cash_disbursement_journal,
                 [('Date','date'),('Reference','reference_no'),('Particulars','particulars'),
                  ('Account Description','account_description'),('Account Code','account_code'),
                  ('Debit','debit'),('Credit','credit')],
                 None),
                ('Cash Receipts Journal',     self.db_manager.get_cash_receipts_journal,
                 [('Date','date'),('Reference','reference_no'),('Particulars','particulars'),
                  ('Account Description','account_description'),('Account Code','account_code'),
                  ('Debit','debit'),('Credit','credit')],
                 None),
                ('General Journal',           self.db_manager.get_general_journal,
                 [('Date','date'),('Reference','reference_no'),('Particulars','particulars'),
                  ('Account Description','account_description'),('Account Code','account_code'),
                  ('Debit','debit'),('Credit','credit')],
                 None),
            ]

            for sheet_title, get_fn, columns, name_key in JOURNALS:
                progress.setLabelText(f'Exporting: {sheet_title}…')
                QApplication.processEvents()

                entries = get_fn()
                rows = []
                for entry in entries:
                    lines = entry.get('lines') or [entry]
                    for ln in lines:
                        row = {k: entry.get(k, '') for _, k in columns}
                        row.update({k: ln.get(k, '') for _, k in columns if k in ln})
                        # Format numeric fields
                        for fld in ('debit', 'credit'):
                            v = ln.get(fld, 0) or 0
                            row[fld] = f'{float(v):,.2f}' if v else ''
                        rows.append(row)

                if not rows:
                    continue

                # Write to a temp file, then copy the sheet into the workbook
                import tempfile, shutil
                tmp = tempfile.mktemp(suffix='.xlsx')
                n, err = export_to_xls(
                    rows=rows, path=tmp,
                    sheet_title=sheet_title, columns=columns)

                if err:
                    raise RuntimeError(f'{sheet_title}: {err}')

                # Merge sheet into the master workbook
                from openpyxl import load_workbook as _lw
                src_wb = _lw(tmp)
                src_ws = src_wb.active
                if first_sheet:
                    ws = wb.active
                    ws.title = sheet_title[:31]
                    first_sheet = False
                else:
                    ws = wb.create_sheet(title=sheet_title[:31])

                for row in src_ws.iter_rows(values_only=True):
                    ws.append(row)

                src_wb.close()
                os.unlink(tmp)
                total_rows += n

            wb.save(path)

        except Exception as exc:
            progress.close()
            QMessageBox.critical(self, 'Export Failed', str(exc))
            return

        progress.close()
        QMessageBox.information(
            self, 'Export Complete',
            f'Full book exported to:\n{path}\n\n{total_rows} total rows across {len(JOURNALS)} sheets.')

    # ------------------------------------------------------------------ Help

    def _show_user_manual(self):
        dlg = QDialog(self)
        dlg.setWindowTitle('ESQ Accounting System — User Manual')
        dlg.resize(820, 680)
        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(0, 0, 0, 8)
        header = QLabel(' 📖  ESQ Accounting System — User Manual')
        header.setStyleSheet(
            'background:#1a3a5c; color:#ffffff; font-size:16px; '
            'font-weight:bold; padding:10px 14px;')
        layout.addWidget(header)
        browser = QTextBrowser()
        browser.setOpenExternalLinks(False)
        browser.setFont(QFont('Segoe UI', 10))
        manual_path = get_resource('user_manual.html')
        try:
            with open(manual_path, 'r', encoding='utf-8') as f:
                html = f.read()
        except FileNotFoundError:
            html = f'<p><b>User manual not found.</b><br>{manual_path}</p>'
        browser.setHtml(html)
        layout.addWidget(browser)
        btn_box = QDialogButtonBox(QDialogButtonBox.Close)
        btn_box.rejected.connect(dlg.accept)
        btn_layout = QHBoxLayout()
        btn_layout.setContentsMargins(8, 0, 8, 0)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_box)
        layout.addLayout(btn_layout)
        dlg.exec()

    def _show_about(self):
        QMessageBox.about(
            self, 'About ESQ Company',
            '<h2>ESQ Company Accounting System</h2>'
            '<p>Version 2.0</p>'
            '<p>Built with Python and PySide6.</p>'
            '<p>© 2025 ESQ Company</p>')