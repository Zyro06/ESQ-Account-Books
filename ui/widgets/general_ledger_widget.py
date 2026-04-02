from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget,
                             QTableWidgetItem, QPushButton, QLabel, QComboBox,
                             QHeaderView, QGroupBox, QDateEdit, QLineEdit, QFileDialog, QMessageBox)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QKeySequence, QFont, QColor, QShortcut
from database.db_manager import DatabaseManager
from ui.utils.search_utils import SearchFilter, add_month_combo
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

MONTHS = ["January","February","March","April","May","June",
          "July","August","September","October","November","December"]

# ---------------------------------------------------------------------------
# Column width defaults (pixels) — change these to adjust default widths
# ---------------------------------------------------------------------------

# Single-account view: Date, Reference No., Particulars, Debit, Credit, Balance
SINGLE_COL_WIDTHS = {
    0: 100,   # Date
    1: 145,   # Reference No.
    2: None,  # Particulars — stretches to fill remaining space
    3: 130,   # Debit
    4: 130,   # Credit
    5: 140,   # Balance
}

# All-accounts view: Account Code, Account Description, Debit, Credit, Balance
ALL_COL_WIDTHS = {
    0: 150,   # Account Code
    1: None,  # Account Description — stretches to fill remaining space
    2: 130,   # Debit
    3: 130,   # Credit
    4: 140,   # Balance
}


def _apply_col_widths(header: QHeaderView, widths: dict):
    """
    Apply column widths from a dict {col_index: pixels | None}.
    None  → Interactive (user-resizable stretch column).
    int   → Fixed pixel width, still user-resizable (Interactive mode).
    """
    for col, width in widths.items():
        if width is None:
            header.setSectionResizeMode(col, QHeaderView.Stretch)
        else:
            header.setSectionResizeMode(col, QHeaderView.Interactive)
            header.resizeSection(col, width)


class GeneralLedgerWidget(QWidget):

    def __init__(self, db_manager: DatabaseManager):
        super().__init__()
        self.db_manager = db_manager
        self._current_mode    = None   # 'single' or 'all'
        self._current_account = None
        self._current_entries = []
        self._current_normal  = 'Debit'
        self._setup_ui()
        self._setup_shortcuts()

    # ------------------------------------------------------------------ UI

    def _setup_ui(self):
        layout = QVBoxLayout()

        title = QLabel("GENERAL LEDGER")
        title.setProperty("class", "title")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # ── Row 1: Search & Filter ─────────────────────────────────────────
        search_group = QGroupBox("Search && Filter")
        search_layout = QHBoxLayout()

        self.month_combo = add_month_combo(search_layout)

        search_layout.addWidget(QLabel("Search:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search by date, reference, or particulars...")
        self.search_input.setClearButtonEnabled(True)
        search_layout.addWidget(self.search_input)

        search_layout.addWidget(QLabel("From:"))
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDisplayFormat("MM/dd/yyyy")
        self.date_from.setDate(QDate(2000, 1, 1))
        search_layout.addWidget(self.date_from)

        search_layout.addWidget(QLabel("To:"))
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDisplayFormat("MM/dd/yyyy")
        self.date_to.setDate(QDate.currentDate())
        search_layout.addWidget(self.date_to)

        self.clear_search_btn = QPushButton("Clear Filter")
        self.clear_search_btn.clicked.connect(self._clear_filter)
        search_layout.addWidget(self.clear_search_btn)

        self.results_label = QLabel("Showing: 0 of 0")
        search_layout.addWidget(self.results_label)

        search_group.setLayout(search_layout)
        layout.addWidget(search_group)

        # ── Row 2: Account selector + Export ──────────────────────────────
        account_group = QGroupBox("Select Account")
        account_layout = QHBoxLayout()

        account_layout.addWidget(QLabel("Account:"))
        self.account_combo = QComboBox()
        self.account_combo.setEditable(True)
        self.account_combo.setMinimumWidth(380)
        self.account_combo.setInsertPolicy(QComboBox.NoInsert)
        account_layout.addWidget(self.account_combo)

        self.load_btn = QPushButton("Load")
        self.load_btn.clicked.connect(self._on_load)
        account_layout.addWidget(self.load_btn)

        self.export_btn = QPushButton("Export")
        self.export_btn.clicked.connect(self._export_xls)
        account_layout.addWidget(self.export_btn)

        account_layout.addStretch()
        account_group.setLayout(account_layout)
        layout.addWidget(account_group)

        # ── Account info label ─────────────────────────────────────────────
        self.account_info_label = QLabel("")
        self.account_info_label.setProperty("class", "bold")
        layout.addWidget(self.account_info_label)

        # ── Table ──────────────────────────────────────────────────────────
        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setAlternatingRowColors(True)
        # Allow user to drag column borders to resize
        self.table.horizontalHeader().setSectionsMovable(False)
        self.table.horizontalHeader().setStretchLastSection(False)
        layout.addWidget(self.table)

        # ── Totals ─────────────────────────────────────────────────────────
        self.totals_label = QLabel("")
        self.totals_label.setProperty("class", "total")
        layout.addWidget(self.totals_label)

        self.setLayout(layout)

        # SearchFilter + hook totals update
        self._search = SearchFilter(
            table         = self.table,
            search_input  = self.search_input,
            results_label = self.results_label,
            date_from     = self.date_from,
            date_to       = self.date_to,
            month_combo   = self.month_combo,
            date_col      = 0,
        )
        orig_run = self._search._run
        def _run_with_totals():
            orig_run()
            if self._current_mode == 'single':
                self._update_single_totals_from_visible()
            elif self._current_mode == 'all':
                self._update_all_totals_from_visible()
        self._search._timer.timeout.disconnect()
        self._search._timer.timeout.connect(_run_with_totals)
        self.date_from.dateChanged.connect(_run_with_totals)
        self.date_to.dateChanged.connect(_run_with_totals)
        self.month_combo.currentIndexChanged.connect(_run_with_totals)
        self._run_with_totals = _run_with_totals

    def _setup_shortcuts(self):
        QShortcut(QKeySequence("Ctrl+F"), self).activated.connect(self.search_input.setFocus)
        QShortcut(QKeySequence("Return"), self).activated.connect(self._on_load)
        QShortcut(QKeySequence("Ctrl+R"), self).activated.connect(self._clear_filter)

    # ------------------------------------------------------------------ account list

    def load_data(self):
        self._load_accounts()

    def showEvent(self, event):
        super().showEvent(event)
        self._load_accounts()

    def _load_accounts(self):
        prev_code = self.account_combo.currentData()
        self.account_combo.blockSignals(True)
        self.account_combo.clear()
        self.account_combo.addItem("All Accounts", "__ALL__")
        self.account_combo.addItem("-- Select Account --", "")
        accounts = self.db_manager.get_all_accounts()
        for account in accounts:
            self.account_combo.addItem(
                f"{account['account_code']}  {account['account_description']}",
                account['account_code'])
        if prev_code:
            for i in range(self.account_combo.count()):
                if self.account_combo.itemData(i) == prev_code:
                    self.account_combo.setCurrentIndex(i)
                    break
        self.account_combo.blockSignals(False)

    # ------------------------------------------------------------------ helpers

    def _clear_filter(self):
        self.search_input.clear()
        self.date_from.setDate(QDate(2000, 1, 1))
        self.date_to.setDate(QDate.currentDate())
        self.month_combo.setCurrentIndex(0)

    def _on_load(self):
        account_code = self.account_combo.currentData()
        date_from    = self.date_from.date().toString("MM/dd/yyyy")
        date_to      = self.date_to.date().toString("MM/dd/yyyy")
        if account_code == "__ALL__":
            self._load_all_accounts(date_from, date_to)
        elif account_code:
            self._load_single_account(account_code, date_from, date_to)
        else:
            self.table.setRowCount(0)
            self.account_info_label.setText("")
            self.totals_label.setText("")

    # ------------------------------------------------------------------ single account

    def _load_single_account(self, account_code=None, date_from=None, date_to=None):
        if not account_code:
            return

        ledger         = self.db_manager.get_general_ledger(account_code, date_from, date_to)
        entries        = ledger['entries']
        normal_balance = ledger.get('normal_balance', 'Debit')

        self._current_mode    = 'single'
        self._current_account = account_code
        self._current_entries = entries
        self._current_normal  = normal_balance

        accounts     = self.db_manager.get_all_accounts()
        account_desc = next(
            (a['account_description'] for a in accounts
             if a['account_code'] == account_code), '')
        self.account_info_label.setText(f"Account: {account_code}  |  {account_desc}")

        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(
            ["Date", "Reference No.", "Particulars", "Debit", "Credit", "Balance"])

        # Apply user-configurable column widths for single-account view
        _apply_col_widths(self.table.horizontalHeader(), SINGLE_COL_WIDTHS)

        self.table.setRowCount(len(entries))

        running_balance = total_debit = total_credit = 0

        for row, entry in enumerate(entries):
            debit  = float(entry.get('debit',  0) or 0)
            credit = float(entry.get('credit', 0) or 0)
            running_balance += (credit - debit) if normal_balance == 'Credit' \
                               else (debit - credit)
            total_debit  += debit
            total_credit += credit

            row_data = [
                entry.get('date', ''),
                entry.get('reference_no', ''),
                entry.get('particulars', ''),
                f"{debit:,.2f}"          if debit  > 0 else "",
                f"{credit:,.2f}"         if credit > 0 else "",
                f"{running_balance:,.2f}",
            ]
            for col, text in enumerate(row_data):
                item = QTableWidgetItem(str(text))
                item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                if col >= 3:
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table.setItem(row, col, item)

        self._search.refresh()
        self._update_single_totals_from_visible()

    def _update_single_totals_from_visible(self):
        normal_balance = self._current_normal
        running = total_dr = total_cr = 0.0

        for row in range(self.table.rowCount()):
            if self.table.isRowHidden(row):
                continue
            try:
                dr_text = self.table.item(row, 3).text().replace(',', '')
                cr_text = self.table.item(row, 4).text().replace(',', '')
                dr = float(dr_text) if dr_text else 0.0
                cr = float(cr_text) if cr_text else 0.0
                running += (cr - dr) if normal_balance == 'Credit' else (dr - cr)
                total_dr += dr
                total_cr += cr
                bal_item = QTableWidgetItem(f"{running:,.2f}")
                bal_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                bal_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table.setItem(row, 5, bal_item)
            except (AttributeError, ValueError):
                pass

        self.totals_label.setText(
            f"Totals: Debit: {total_dr:,.2f} | "
            f"Credit: {total_cr:,.2f} | "
            f"Balance: {running:,.2f}")

    # ------------------------------------------------------------------ all accounts

    def _load_all_accounts(self, date_from=None, date_to=None):
        accounts = self.db_manager.get_all_accounts()
        self._current_mode = 'all'
        self.account_info_label.setText("All Accounts")

        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(
            ["Account Code", "Account Description", "Debit", "Credit", "Balance"])

        # Apply user-configurable column widths for all-accounts view
        _apply_col_widths(self.table.horizontalHeader(), ALL_COL_WIDTHS)

        rows = []
        grand_debit = grand_credit = grand_balance = 0

        for account in accounts:
            code    = account['account_code']
            desc    = account['account_description']
            ledger  = self.db_manager.get_general_ledger(code, date_from, date_to)
            entries = ledger['entries']
            total_debit  = sum(float(e.get('debit',  0) or 0) for e in entries)
            total_credit = sum(float(e.get('credit', 0) or 0) for e in entries)
            nb      = ledger.get('normal_balance', 'Debit')
            balance = (total_credit - total_debit) if nb == 'Credit' \
                      else (total_debit - total_credit)
            if total_debit == 0 and total_credit == 0:
                continue
            rows.append((code, desc, total_debit, total_credit, balance))
            grand_debit   += total_debit
            grand_credit  += total_credit
            grand_balance += balance

        self.table.setRowCount(len(rows))
        for row, (code, desc, debit, credit, balance) in enumerate(rows):
            for col, (text, right) in enumerate([
                (code,              False),
                (desc,              False),
                (f"{debit:,.2f}",   True),
                (f"{credit:,.2f}",  True),
                (f"{balance:,.2f}", True),
            ]):
                item = QTableWidgetItem(text)
                item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                if right:
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table.setItem(row, col, item)

        self._search.refresh()
        self._update_all_totals_from_visible()

    def _update_all_totals_from_visible(self):
        td = tc = tb = 0.0
        for row in range(self.table.rowCount()):
            if self.table.isRowHidden(row):
                continue
            try:
                td += float(self.table.item(row, 2).text().replace(',', ''))
                tc += float(self.table.item(row, 3).text().replace(',', ''))
                tb += float(self.table.item(row, 4).text().replace(',', ''))
            except (AttributeError, ValueError):
                pass
        self.totals_label.setText(
            f"Totals: Debit: {td:,.2f} | Credit: {tc:,.2f} | Balance: {tb:,.2f}")

    # ------------------------------------------------------------------ Export

    def _export_xls(self):
        if not _OPENPYXL_OK:
            QMessageBox.critical(self, "Missing Library",
                                 "openpyxl is required.\nInstall with: pip install openpyxl")
            return
        if self._current_mode is None:
            QMessageBox.information(self, "Nothing to Export",
                                    "Please load an account first.")
            return

        from resources.file_paths import get_io_dir
        default_name = "general_ledger.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "Export General Ledger",
            os.path.join(get_io_dir("General Ledger"), default_name),
            "Excel Files (*.xlsx)")
        if not path:
            return

        if self._current_mode == 'single':
            err = self._export_single(path)
        else:
            err = self._export_all(path)

        if err:
            QMessageBox.critical(self, "Export Failed", err)
        else:
            QMessageBox.information(self, "Export Successful", f"Exported to:\n{path}")

    def _export_single(self, path: str) -> str:
        try:
            from datetime import datetime as _dt

            visible_entries = []
            for row in range(self.table.rowCount()):
                if self.table.isRowHidden(row):
                    continue
                entry = self._current_entries[row] if row < len(self._current_entries) else None
                if entry:
                    visible_entries.append(entry)

            if not visible_entries:
                return "No visible entries to export."

            accounts = self.db_manager.get_all_accounts()
            account_code = self._current_account
            account_desc = next(
                (a['account_description'] for a in accounts
                 if a['account_code'] == account_code), '')
            normal_balance = self._current_normal

            wb = Workbook()
            ws = wb.active
            ws.title = "General Ledger"

            hdr_font    = XLFont(name='Arial', bold=True, color='FFFFFF', size=11)
            hdr_fill    = PatternFill('solid', start_color='2F5496')
            hdr_align   = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_font   = XLFont(name='Arial', size=10)
            total_font  = XLFont(name='Arial', bold=True, size=10)
            total_fill  = PatternFill('solid', start_color='D9E1F2')
            bal_fill    = PatternFill('solid', start_color='E2EFDA')
            beg_fill    = PatternFill('solid', start_color='FFF2CC')
            thin        = Side(style='thin', color='B0B0B0')
            border      = Border(left=thin, right=thin, top=thin, bottom=thin)
            r_align     = Alignment(horizontal='right',  vertical='center')
            l_align     = Alignment(horizontal='left',   vertical='center')

            COLS = ['DATE','ACCOUNT CODE','ACCOUNT DESCRIPTION',
                    'REFERENCE NO','PARTICULARS','DEBIT','CREDIT','BALANCE']
            WIDTHS = {1:14, 2:14, 3:28, 4:16, 5:36, 6:14, 7:14, 8:14}

            ws.merge_cells('A2:H2')
            ws['A2'].value = 'GENERAL LEDGER'
            ws['A2'].font  = XLFont(name='Arial', bold=True, size=14)
            ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[2].height = 22

            ws.merge_cells('A3:H3')
            ws['A3'].value = f'For the Year {_dt.now().year}'
            ws['A3'].font  = XLFont(name='Arial', italic=True, size=11)
            ws['A3'].alignment = Alignment(horizontal='left', vertical='center')

            HR = 5
            ws.row_dimensions[HR].height = 28
            for ci, h in enumerate(COLS, 1):
                cell = ws.cell(row=HR, column=ci, value=h)
                cell.font = hdr_font; cell.fill = hdr_fill
                cell.alignment = hdr_align; cell.border = border

            for ci, w in WIDTHS.items():
                ws.column_dimensions[get_column_letter(ci)].width = w
            ws.freeze_panes = 'A6'

            from collections import defaultdict
            months = defaultdict(list)
            for e in visible_entries:
                date_str = e.get('date', '')
                try:
                    from datetime import datetime
                    d = datetime.strptime(date_str, "%m/%d/%Y")
                    key = (d.year, d.month)
                except:
                    key = (0, 0)
                months[key].append(e)

            current_row = 6
            running_balance = 0.0

            for (year, month) in sorted(months.keys()):
                month_entries = months[(year, month)]
                import calendar
                last_day = calendar.monthrange(year, month)[1]
                beg_bal = running_balance

                beg_date = f"{month:02d}/01/{year}"
                row_vals = [beg_date, account_code, account_desc,
                            'Beginning Balance', '', '', '', f"{beg_bal:,.2f}"]
                ws.row_dimensions[current_row].height = 18
                for ci, val in enumerate(row_vals, 1):
                    cell = ws.cell(row=current_row, column=ci, value=val)
                    cell.font = total_font; cell.fill = beg_fill; cell.border = border
                    cell.alignment = r_align if ci >= 6 else l_align
                current_row += 1

                month_dr = month_cr = 0.0
                for e in month_entries:
                    debit  = float(e.get('debit',  0) or 0)
                    credit = float(e.get('credit', 0) or 0)
                    running_balance += (credit - debit) if normal_balance == 'Credit' \
                                       else (debit - credit)
                    month_dr += debit
                    month_cr += credit

                    row_vals = [
                        e.get('date', ''), account_code, account_desc,
                        e.get('reference_no', ''), e.get('particulars', '') or '',
                        f"{debit:,.2f}" if debit > 0 else '',
                        f"{credit:,.2f}" if credit > 0 else '',
                        '',
                    ]
                    ws.row_dimensions[current_row].height = 18
                    for ci, val in enumerate(row_vals, 1):
                        cell = ws.cell(row=current_row, column=ci, value=val)
                        cell.font = cell_font; cell.border = border
                        cell.alignment = r_align if ci >= 6 else l_align
                    current_row += 1

                end_date = f"{month:02d}/{last_day:02d}/{year}"
                row_vals = [end_date, account_code, account_desc,
                            'MONTH TOTAL', '',
                            f"{month_dr:,.2f}", f"{month_cr:,.2f}",
                            f"{month_dr - month_cr:,.2f}"]
                ws.row_dimensions[current_row].height = 18
                for ci, val in enumerate(row_vals, 1):
                    cell = ws.cell(row=current_row, column=ci, value=val)
                    cell.font = total_font; cell.fill = total_fill; cell.border = border
                    cell.alignment = r_align if ci >= 6 else l_align
                current_row += 1

                row_vals = [end_date, account_code, account_desc,
                            'ENDING BALANCE', '', '', '', f"{running_balance:,.2f}"]
                ws.row_dimensions[current_row].height = 18
                for ci, val in enumerate(row_vals, 1):
                    cell = ws.cell(row=current_row, column=ci, value=val)
                    cell.font = total_font; cell.fill = bal_fill; cell.border = border
                    cell.alignment = r_align if ci >= 6 else l_align
                current_row += 1
                current_row += 1

            ws.auto_filter.ref = f'A{HR}:H{HR}'
            wb.save(path)
            return ''
        except Exception as e:
            return str(e)

    def _export_all(self, path: str) -> str:
        try:
            from datetime import datetime as _dt
            wb = Workbook()
            ws = wb.active
            ws.title = "General Ledger - All Accounts"

            hdr_font  = XLFont(name='Arial', bold=True, color='FFFFFF', size=11)
            hdr_fill  = PatternFill('solid', start_color='2F5496')
            hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_font = XLFont(name='Arial', size=10)
            alt_fill  = PatternFill('solid', start_color='DCE6F1')
            thin      = Side(style='thin', color='B0B0B0')
            border    = Border(left=thin, right=thin, top=thin, bottom=thin)
            r_align   = Alignment(horizontal='right', vertical='center')
            l_align   = Alignment(horizontal='left',  vertical='center')

            ws.merge_cells('A2:E2')
            ws['A2'].value = 'GENERAL LEDGER — ALL ACCOUNTS'
            ws['A2'].font  = XLFont(name='Arial', bold=True, size=14)
            ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[2].height = 22

            ws.merge_cells('A3:E3')
            ws['A3'].value = f'For the Year {_dt.now().year}'
            ws['A3'].font  = XLFont(name='Arial', italic=True, size=11)
            ws['A3'].alignment = Alignment(horizontal='left', vertical='center')

            HR = 5
            headers = ['Account Code', 'Account Description', 'Debit', 'Credit', 'Balance']
            ws.row_dimensions[HR].height = 28
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=HR, column=ci, value=h)
                cell.font = hdr_font; cell.fill = hdr_fill
                cell.alignment = hdr_align; cell.border = border

            for ci, w in {1:16, 2:40, 3:16, 4:16, 5:16}.items():
                ws.column_dimensions[get_column_letter(ci)].width = w
            ws.freeze_panes = 'A6'

            ri = 0
            for row in range(self.table.rowCount()):
                if self.table.isRowHidden(row):
                    continue
                row_idx = 6 + ri
                fill = alt_fill if ri % 2 == 0 else None
                ws.row_dimensions[row_idx].height = 18
                for ci in range(5):
                    item = self.table.item(row, ci)
                    val  = item.text() if item else ''
                    cell = ws.cell(row=row_idx, column=ci + 1, value=val)
                    cell.font = cell_font; cell.border = border
                    cell.alignment = r_align if ci >= 2 else l_align
                    if fill:
                        cell.fill = fill
                ri += 1

            ws.auto_filter.ref = f'A{HR}:E{HR}'
            wb.save(path)
            return ''
        except Exception as e:
            return str(e)