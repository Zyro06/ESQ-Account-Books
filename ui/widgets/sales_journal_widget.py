import os
from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget,
                             QTableWidgetItem, QPushButton, QLabel, QLineEdit,
                             QDialog, QFormLayout, QDialogButtonBox, QMessageBox,
                             QHeaderView, QDateEdit, QDoubleSpinBox, QTextEdit,
                             QComboBox, QGroupBox, QFileDialog, QFrame)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QKeySequence, QFont, QColor, QShortcut
from database.db_manager import DatabaseManager, _numeric_suffix
from ui.utils.search_utils import SearchFilter, add_month_combo
from resources.file_paths import get_import_dir, get_io_dir
from utils.date_utils   import DateItem as _DateItem
from utils.export_utils import export_to_xls as _export_to_xls
from utils.import_utils import import_grouped_from_xls
from ui.dialogs.line_dialog         import LineDialog       as _LineDialog
from ui.dialogs.view_details_dialog import ViewEntryDialog  as _ViewSJDialog

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

_AR_NUM    = '1110'
_VAT_NUM   = '2210'
_GOODS_NUM = '4010'
_SVC_NUM   = '4020'

_XLS_COLS = [
    ('Date','date'), ('Customer Name','customer_name'), ('Reference No','reference_no'),
    ('TIN','tin'), ('Particulars','particulars'),
    ('Account Description','account_description'), ('Account Code','account_code'),
    ('Debit','debit'), ('Credit','credit'),
]
_XLS_HEADERS = [h for h,_ in _XLS_COLS]
_XLS_KEYS    = [k for _,k in _XLS_COLS]


class SalesJournalDialog(QDialog):
    def __init__(self, db_manager, parent=None, entry_data=None, is_copy=False):
        super().__init__(parent)
        self.db_manager    = db_manager
        self.user_lines    = []
        self.alphalist_map = {}
        self._ar_code = ''; self._ar_desc = 'Accounts Receivable'

        title = ("Copy Entry" if is_copy else
                 "Add Sales Entry" if entry_data is None else "Edit Sales Entry")
        self.setWindowTitle(title)
        self.setModal(True); self.setMinimumWidth(820); self.setMinimumHeight(580)

        root = QVBoxLayout(self); root.setSpacing(10)

        hdr = QFormLayout(); hdr.setLabelAlignment(Qt.AlignRight)
        self.date_input = QDateEdit(); self.date_input.setCalendarPopup(True)
        self.date_input.setDisplayFormat("MM/dd/yyyy")
        self.date_input.setDate(
            QDate.fromString(entry_data["date"], "MM/dd/yyyy")
            if entry_data else QDate.currentDate())
        hdr.addRow("Date:", self.date_input)

        self.customer_input = QComboBox()
        self.customer_input.setEditable(True)
        self.customer_input.setMinimumWidth(300)
        self._load_customers()
        hdr.addRow("Customer Name:", self.customer_input)

        self.tin_input = QLineEdit()
        self.tin_input.setReadOnly(True)
        self.tin_input.setObjectName("readOnlyTin")
        hdr.addRow("TIN:", self.tin_input)

        self.reference_input = QLineEdit()
        if entry_data:
            self.reference_input.setText(entry_data.get("reference_no", ""))
        hdr.addRow("Reference No:", self.reference_input)

        self.particulars_input = QLineEdit()
        if entry_data:
            self.particulars_input.setText(entry_data.get("particulars", "") or "")
        hdr.addRow("Particulars:", self.particulars_input)

        root.addLayout(hdr)
        sep = QFrame(); sep.setFrameShape(QFrame.HLine); sep.setFrameShadow(QFrame.Sunken)
        root.addWidget(sep)

        lines_hdr = QHBoxLayout()
        lbl = QLabel("Journal Lines:"); lbl.setStyleSheet("font-weight:bold;")
        ar_info = QLabel("  (AR will be auto-computed as the balancing debit)")
        ar_info.setStyleSheet("color:#888; font-style:italic;")
        lines_hdr.addWidget(lbl); lines_hdr.addWidget(ar_info); lines_hdr.addStretch()
        root.addLayout(lines_hdr)

        self.lines_table = QTableWidget()
        self.lines_table.setColumnCount(4)
        self.lines_table.setHorizontalHeaderLabels(["Account", "Code", "Debit", "Credit"])
        self.lines_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.lines_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.lines_table.verticalHeader().setVisible(False)
        self.lines_table.setAlternatingRowColors(True)
        lh = self.lines_table.horizontalHeader()
        lh.setSectionResizeMode(0, QHeaderView.Stretch)
        lh.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        lh.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        lh.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        root.addWidget(self.lines_table)

        btn_row = QHBoxLayout()
        add_btn  = QPushButton("Add Line");    add_btn.clicked.connect(self._add_line)
        edit_btn = QPushButton("Edit Line");   edit_btn.clicked.connect(self._edit_line)
        del_btn  = QPushButton("Remove Line"); del_btn.clicked.connect(self._remove_line)
        btn_row.addWidget(add_btn); btn_row.addWidget(edit_btn)
        btn_row.addWidget(del_btn); btn_row.addStretch()
        root.addLayout(btn_row)

        tot_row = QHBoxLayout(); tot_row.addStretch()
        bf = QFont(); bf.setBold(True)
        self.total_dr_lbl = QLabel("Total Debit:  0.00")
        self.total_cr_lbl = QLabel("Total Credit: 0.00")
        self.balance_lbl  = QLabel("")
        self.ar_lbl       = QLabel("")
        self.total_dr_lbl.setFont(bf); self.total_cr_lbl.setFont(bf)
        tot_row.addWidget(self.total_dr_lbl); tot_row.addWidget(QLabel("  |  "))
        tot_row.addWidget(self.total_cr_lbl); tot_row.addWidget(QLabel("  "))
        tot_row.addWidget(self.balance_lbl)
        root.addLayout(tot_row)
        root.addWidget(self.ar_lbl)

        btns = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        btns.accepted.connect(self._on_save); btns.rejected.connect(self.reject)
        root.addWidget(btns)

        self.customer_input.currentTextChanged.connect(self._on_customer_changed)
        self._find_ar_account()

        if entry_data:
            self.customer_input.setCurrentText(entry_data.get("customer_name", "") or "")
            self.tin_input.setText(str(entry_data.get("tin", "") or ""))
            for ln in entry_data.get("lines", []):
                if _numeric_suffix(ln.get("account_code", "")) != _AR_NUM:
                    self.user_lines.append(dict(ln))
        self._refresh_lines()

    def _find_ar_account(self):
        for acct in self.db_manager.get_all_accounts():
            if _numeric_suffix(acct["account_code"]) == _AR_NUM:
                self._ar_code = acct["account_code"]
                self._ar_desc = acct["account_description"]
                break

    def _load_customers(self):
        self.customer_input.addItem("")
        for entry in self.db_manager.get_all_alphalist():
            if entry.get("entry_type", "Customer") not in ("Customer", "Customer&Vendor"):
                continue
            name = (entry.get("company_name") or
                    f"{entry.get('first_name','')} {entry.get('last_name','')}".strip())
            if name:
                self.customer_input.addItem(name)
                self.alphalist_map[name] = {"tin": entry.get("tin", "")}

    def _on_customer_changed(self, name):
        info = self.alphalist_map.get(name, {})
        self.tin_input.setText(str(info.get("tin", "")) if info else "")

    def _add_line(self):
        dlg = _LineDialog(self.db_manager, self)
        if dlg.exec():
            self.user_lines.append(dlg.get_data()); self._refresh_lines()

    def _edit_line(self):
        row = self.lines_table.currentRow()
        if row < 0 or row >= len(self.user_lines):
            QMessageBox.warning(self, "No Selection", "Select a user line to edit.")
            return
        dlg = _LineDialog(self.db_manager, self, self.user_lines[row])
        if dlg.exec():
            self.user_lines[row] = dlg.get_data(); self._refresh_lines()

    def _remove_line(self):
        row = self.lines_table.currentRow()
        if row < 0 or row >= len(self.user_lines):
            QMessageBox.warning(self, "No Selection", "Select a user line to remove.")
            return
        del self.user_lines[row]; self._refresh_lines()

    def _refresh_lines(self):
        total_cr = sum(l.get("credit", 0) for l in self.user_lines)
        total_dr = sum(l.get("debit",  0) for l in self.user_lines)
        ar_amount = total_cr - total_dr

        all_lines = list(self.user_lines)
        if ar_amount > 0:
            all_lines.append({
                "account_description": self._ar_desc,
                "account_code": self._ar_code,
                "debit": ar_amount, "credit": 0, "_auto": True})

        self.lines_table.setRowCount(len(all_lines))
        for r, ln in enumerate(all_lines):
            is_auto = ln.get("_auto", False)
            color   = QColor(230, 240, 255) if is_auto else QColor(240, 255, 240)
            for c, val in enumerate([
                ln.get("account_description", ""), ln.get("account_code", ""),
                f"{ln.get('debit',0):,.2f}"  if ln.get("debit",  0) else "",
                f"{ln.get('credit',0):,.2f}" if ln.get("credit", 0) else "",
            ]):
                item = QTableWidgetItem(val)
                item.setBackground(color)
                item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                item.setTextAlignment(
                    Qt.AlignRight | Qt.AlignVCenter if c >= 2
                    else Qt.AlignLeft  | Qt.AlignVCenter)
                self.lines_table.setItem(r, c, item)

        final_dr = total_dr + (ar_amount if ar_amount > 0 else 0)
        self.total_dr_lbl.setText(f"Total Debit:  {final_dr:,.2f}")
        self.total_cr_lbl.setText(f"Total Credit: {total_cr:,.2f}")
        balanced = abs(final_dr - total_cr) < 0.005 and final_dr > 0
        if balanced:
            self.balance_lbl.setText("Balanced")
            self.balance_lbl.setStyleSheet("color:green;font-weight:bold;")
        elif all_lines:
            self.balance_lbl.setText("Check lines")
            self.balance_lbl.setStyleSheet("color:orange;font-weight:bold;")
        else:
            self.balance_lbl.setText("")
        if ar_amount > 0:
            self.ar_lbl.setText(f"  AR (Balancing Debit): {ar_amount:,.2f}")
            self.ar_lbl.setStyleSheet("color:#2244aa; font-style:italic;")
        else:
            self.ar_lbl.setText("")

    def _on_save(self):
        if not self.customer_input.currentText().strip():
            QMessageBox.warning(self, "Validation", "Customer Name is required.")
            return
        if not self.reference_input.text().strip():
            QMessageBox.warning(self, "Validation", "Reference No. is required.")
            return
        if not self.user_lines:
            QMessageBox.warning(self, "Validation", "Add at least one journal line.")
            return
        self.accept()

    def get_data(self):
        total_cr  = sum(l.get("credit", 0) for l in self.user_lines)
        total_dr  = sum(l.get("debit",  0) for l in self.user_lines)
        ar_amount = total_cr - total_dr
        lines = list(self.user_lines)
        if ar_amount > 0:
            lines.append({"account_description": self._ar_desc,
                          "account_code": self._ar_code,
                          "debit": ar_amount, "credit": 0})
        return {
            "date":          self.date_input.date().toString("MM/dd/yyyy"),
            "customer_name": self.customer_input.currentText().strip(),
            "reference_no":  self.reference_input.text().strip(),
            "tin":           self.tin_input.text().strip(),
            "particulars":   self.particulars_input.text().strip(),
            "lines":         lines,
        }


class SalesJournalWidget(QWidget):
    def __init__(self, db_manager: DatabaseManager):
        super().__init__()
        self.db_manager  = db_manager
        self.all_entries = []
        self._setup_ui(); self._setup_shortcuts(); self.load_data()

    def _setup_ui(self):
        layout = QVBoxLayout()
        title = QLabel("SALES JOURNAL")
        title.setProperty("class", "title")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        sg = QGroupBox("Search && Filter"); sl = QHBoxLayout()
        self.month_combo = add_month_combo(sl)
        sl.addWidget(QLabel("Search:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText(
            "Search by customer, reference, TIN, or particulars...")
        self.search_input.setClearButtonEnabled(True)
        sl.addWidget(self.search_input)
        sl.addWidget(QLabel("From:"))
        self.date_from = QDateEdit(); self.date_from.setCalendarPopup(True)
        self.date_from.setDisplayFormat("MM/dd/yyyy")
        self.date_from.setDate(QDate(2000, 1, 1))
        sl.addWidget(self.date_from)
        sl.addWidget(QLabel("To:"))
        self.date_to = QDateEdit(); self.date_to.setCalendarPopup(True)
        self.date_to.setDisplayFormat("MM/dd/yyyy")
        self.date_to.setDate(QDate.currentDate())
        sl.addWidget(self.date_to)
        self.clear_btn = QPushButton("Clear Filter")
        self.clear_btn.clicked.connect(self._clear_filters)
        sl.addWidget(self.clear_btn)
        self.results_label = QLabel("Showing: 0 of 0")
        sl.addWidget(self.results_label)
        sg.setLayout(sl); layout.addWidget(sg)

        # ── Action buttons (no totals here) ───────────────────────────
        br = QHBoxLayout()
        for label, slot in [
            ("Add Entry",    self._add_entry),
            ("Edit Entry",   self._edit_entry),
            ("Copy Entry",   self._copy_entry),
            ("View Details", self._view_details),
        ]:
            btn = QPushButton(label); btn.clicked.connect(slot); br.addWidget(btn)
        self.delete_btn = QPushButton("Delete Entry")
        self.delete_btn.setProperty("class", "danger")
        self.delete_btn.clicked.connect(self._delete_entry)
        br.addWidget(self.delete_btn)
        self.import_btn = QPushButton("Import")
        self.import_btn.clicked.connect(self._import_xls)
        br.addWidget(self.import_btn)
        self.export_btn = QPushButton("Export")
        self.export_btn.clicked.connect(self._export_xls)
        br.addWidget(self.export_btn)
        br.addStretch()
        layout.addLayout(br)

        # ── Table ──────────────────────────────────────────────────────
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(
            ["Date", "Customer Name", "Reference No.", "TIN", "Lines", "Total Gross"])
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.Interactive)
        hh.setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.setColumnWidth(0, 100)
        self.table.setColumnWidth(2, 150)
        self.table.setColumnWidth(3, 130)
        self.table.setColumnWidth(4, 100)
        self.table.setColumnWidth(5, 150)
        hh.setMinimumSectionSize(50)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.ExtendedSelection)   # ← multi-select
        self.table.setSortingEnabled(True)
        self.table.setAlternatingRowColors(True)
        self.table.doubleClicked.connect(self._view_details)
        self.table.selectionModel().selectionChanged.connect(self._on_selection_changed)
        layout.addWidget(self.table)

        # ── Totals — below the table ───────────────────────────────────
        self.totals_label = QLabel(
            "Totals  |  Gross: 0.00  |  VAT: 0.00  |  Goods: 0.00  |  Services: 0.00")
        self.totals_label.setProperty("class", "total")
        self.totals_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        layout.addWidget(self.totals_label)

        self._search = SearchFilter(
            table=self.table, search_input=self.search_input,
            results_label=self.results_label,
            date_from=self.date_from, date_to=self.date_to,
            month_combo=self.month_combo, date_col=0)
        orig = self._search._run
        def _run_with_totals(): orig(); self._update_totals()
        self._search._timer.timeout.disconnect()
        self._search._timer.timeout.connect(_run_with_totals)
        self.date_from.dateChanged.connect(_run_with_totals)
        self.date_to.dateChanged.connect(_run_with_totals)
        self.month_combo.currentIndexChanged.connect(_run_with_totals)
        self.setLayout(layout)

    def _setup_shortcuts(self):
        QShortcut(QKeySequence("Ctrl+N"),       self).activated.connect(self._add_entry)
        QShortcut(QKeySequence("Ctrl+E"),       self).activated.connect(self._edit_entry)
        QShortcut(QKeySequence("Ctrl+Shift+C"), self).activated.connect(self._copy_entry)
        QShortcut(QKeySequence("Ctrl+V"),       self).activated.connect(self._view_details)
        QShortcut(QKeySequence("Ctrl+D"),       self).activated.connect(self._delete_entry)
        QShortcut(QKeySequence("Ctrl+F"),       self).activated.connect(self.search_input.setFocus)
        QShortcut(QKeySequence("Ctrl+I"),       self).activated.connect(self._import_xls)
        QShortcut(QKeySequence("Ctrl+Shift+E"), self).activated.connect(self._export_xls)

    def _on_selection_changed(self):
        count = len(self._get_selected_entries())
        self.delete_btn.setText(f"Delete ({count})" if count > 1 else "Delete Entry")

    def _get_selected_entries(self) -> list:
        """Return all selected entry dicts (for multi-delete)."""
        seen, entries = set(), []
        for index in self.table.selectionModel().selectedRows():
            item = self.table.item(index.row(), 0)
            if item is None:
                continue
            entry = item.data(Qt.UserRole)
            if entry and entry.get("id") not in seen:
                seen.add(entry.get("id"))
                entries.append(entry)
        return entries

    def load_data(self):
        self.all_entries = self.db_manager.get_sales_journal()
        self._populate_table()
        self._search.refresh(); self._update_totals()

    def _populate_table(self):
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(self.all_entries))
        for r, entry in enumerate(self.all_entries):
            gross = entry.get("gross_amount", 0)
            date_item = _DateItem(entry.get("date", ""))
            date_item.setData(Qt.UserRole, entry)
            self.table.setItem(r, 0, date_item)
            for c, text in [
                (1, entry.get("customer_name", "")),
                (2, entry.get("reference_no", "")),
                (3, str(entry.get("tin", "") or "")),
            ]:
                self.table.setItem(r, c, QTableWidgetItem(text))
            lines_item = QTableWidgetItem(str(len(entry.get("lines", []))))
            lines_item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.table.setItem(r, 4, lines_item)
            gross_item = QTableWidgetItem(f"{gross:,.2f}")
            gross_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            gross_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
            self.table.setItem(r, 5, gross_item)
        self.table.setSortingEnabled(True)

    def _update_totals(self):
        gross = vat = goods = svc = 0.0
        for r in range(self.table.rowCount()):
            if self.table.isRowHidden(r): continue
            entry = self.table.item(r, 0).data(Qt.UserRole)
            if entry:
                gross += entry.get("gross_amount", 0)
                vat   += entry.get("output_vat",   0)
                goods += entry.get("goods",         0)
                svc   += entry.get("services",      0)
        self.totals_label.setText(
            f"Totals  |  Gross: {gross:,.2f}  |  VAT: {vat:,.2f}  |  "
            f"Goods: {goods:,.2f}  |  Services: {svc:,.2f}")

    def _clear_filters(self):
        self.search_input.clear()
        self.date_from.setDate(QDate(2000, 1, 1))
        self.date_to.setDate(QDate.currentDate())
        self.month_combo.setCurrentIndex(0)

    def _get_selected(self):
        row = self.table.currentRow()
        if row < 0: return None
        item = self.table.item(row, 0)
        return item.data(Qt.UserRole) if item else None

    def _add_entry(self):
        dlg = SalesJournalDialog(self.db_manager, self)
        if dlg.exec():
            if self.db_manager.add_sales_entry(dlg.get_data()):
                self.load_data()
                QMessageBox.information(self, "Success", "Sales entry added successfully!")
            else:
                QMessageBox.warning(self, "Error", "Failed to add entry!")

    def _edit_entry(self):
        entry = self._get_selected()
        if not entry:
            QMessageBox.warning(self, "Warning", "Please select an entry to edit.")
            return
        dlg = SalesJournalDialog(self.db_manager, self, entry)
        if dlg.exec():
            if self.db_manager.update_sales_entry(entry["id"], dlg.get_data()):
                self.load_data()
                QMessageBox.information(self, "Success", "Entry updated successfully!")
            else:
                QMessageBox.warning(self, "Error", "Failed to update entry!")

    def _copy_entry(self):
        entry = self._get_selected()
        if not entry:
            QMessageBox.warning(self, "Warning", "Please select an entry to copy.")
            return
        dlg = SalesJournalDialog(self.db_manager, self, entry, is_copy=True)
        if dlg.exec():
            if self.db_manager.add_sales_entry(dlg.get_data()):
                self.load_data()
                QMessageBox.information(self, "Success", "Entry copied successfully!")
            else:
                QMessageBox.warning(self, "Error", "Failed to copy entry!")

    def _view_details(self):
        entry = self._get_selected()
        if not entry:
            QMessageBox.warning(self, "Warning", "Please select an entry to view.")
            return
        _ViewSJDialog(self, entry, name_label='Customer').exec()

    def _delete_entry(self):
        entries = self._get_selected_entries()
        if not entries:
            QMessageBox.warning(self, "Warning", "Please select one or more entries to delete.")
            return
        count = len(entries)
        if count == 1:
            msg = f"Delete entry '{entries[0]['reference_no']}'?"
        else:
            refs = ', '.join(e['reference_no'] for e in entries[:5])
            if count > 5:
                refs += f' … and {count - 5} more'
            msg = f"Delete {count} selected entries?\n\n{refs}"
        reply = QMessageBox.question(self, "Confirm Delete", msg,
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            for entry in entries:
                self.db_manager.delete_sales_entry(entry["id"])
            self.load_data()
            QMessageBox.information(self, "Success",
                f'{count} entr{"y" if count == 1 else "ies"} deleted successfully!')

    def _export_xls(self):
        if not _OPENPYXL_OK:
            QMessageBox.critical(self, "Missing Library", "openpyxl required.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Sales Journal",
            os.path.join(get_io_dir("Sales Journal"), "sales_journal.xlsx"),
            "Excel Files (*.xlsx)")
        if not path: return
        rows = []
        for r in range(self.table.rowCount()):
            if self.table.isRowHidden(r): continue
            entry = self.table.item(r, 0).data(Qt.UserRole)
            for ln in entry.get("lines", []):
                rows.append({
                    "date": entry["date"], "customer_name": entry["customer_name"],
                    "reference_no": entry["reference_no"], "tin": entry.get("tin", ""),
                    "particulars": entry.get("particulars", ""),
                    "account_description": ln.get("account_description", ""),
                    "account_code": ln.get("account_code", ""),
                    "debit":  f"{ln.get('debit',  0):,.2f}" if ln.get("debit",  0) else "",
                    "credit": f"{ln.get('credit', 0):,.2f}" if ln.get("credit", 0) else "",
                })
        n, err = _export_to_xls(
            rows, path, "Sales Journal", columns=_XLS_COLS,
            col_widths={1: 12, 2: 28, 3: 16, 4: 16, 5: 28, 6: 28, 7: 14, 8: 14, 9: 14})
        if err:   QMessageBox.critical(self, "Export Failed", err)
        else:     QMessageBox.information(self, "Export Successful",
                                          f"{n} line(s) exported to:\n{path}")

    def _import_xls(self):
        if not _OPENPYXL_OK:
            QMessageBox.critical(self, "Missing Library", "openpyxl required.")
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "Import Sales Journal", get_import_dir(""),
            "Excel Files (*.xlsx *.xls)")
        if not path: return
        try:
            imported, skipped, errors = import_grouped_from_xls(
                path=path, db_manager=self.db_manager,
                add_method_name='add_sales_entry',
                columns=_XLS_COLS, customer_key='customer_name')
        except Exception as exc:
            QMessageBox.critical(self, "Import Failed", str(exc))
            return
        self.load_data()
        msg = f"Import complete.\n  Imported: {imported}\n  Skipped: {skipped}"
        if errors: msg += "\n\nDetails:\n" + "\n".join(errors[:20])
        QMessageBox.information(self, "Import Summary", msg)