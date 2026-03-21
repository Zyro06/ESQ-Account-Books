import os
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget,
                             QTableWidgetItem, QPushButton, QLabel, QLineEdit,
                             QDialog, QFormLayout, QDialogButtonBox, QMessageBox,
                             QHeaderView, QDateEdit, QDoubleSpinBox,
                             QComboBox, QGroupBox, QShortcut, QFileDialog)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QKeySequence
from database.db_manager import DatabaseManager
from ui.search_utils import SearchFilter, add_month_combo
from resources.file_paths import get_import_dir, get_io_dir

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

_XLS_COLUMNS = [
    ('Date',               'date'),
    ('Payee Name',         'payee_name'),
    ('Reference No',       'reference_no'),
    ('TIN',                'tin'),
    ('Branch Code',        'branch_code'),
    ('Net Amount',         'net_amount'),
    ('Input VAT',          'input_vat'),
    ('Gross Amount',       'gross_amount'),
    ('Account Description','account_description'),
    ('Account Code',       'account_code'),
    ('Debit',              'debit'),
    ('Particulars',        'particulars'),
]
_XLS_HEADERS = [h for h, _ in _XLS_COLUMNS]
_XLS_KEYS    = [k for _, k in _XLS_COLUMNS]

def _parse_date(date_str: str):
    for fmt in ("MM/dd/yyyy", "M/d/yyyy", "M/dd/yyyy", "MM/d/yyyy", "yyyy-MM-dd"):
        d = QDate.fromString(date_str, fmt)
        if d.isValid():
            return d
    return QDate()

class _DateItem(QTableWidgetItem):
    def __init__(self, display_text: str):
        super().__init__(display_text)
        qdate = _parse_date(display_text)
        self._sort_key = qdate.toString("yyyy-MM-dd") if qdate.isValid() else display_text
    def __lt__(self, other):
        if isinstance(other, _DateItem):
            return self._sort_key < other._sort_key
        return super().__lt__(other)


class PurchaseJournalDialog(QDialog):
    def __init__(self, db_manager, parent=None, entry_data=None, is_copy=False):
        super().__init__(parent)
        self.db_manager = db_manager
        self.entry_data = entry_data
        self.is_copy = is_copy
        self.alphalist_map = {}
        if is_copy:
            self.setWindowTitle("Copy Entry (Create New)")
        elif entry_data is None:
            self.setWindowTitle("Add Purchase Entry")
        else:
            self.setWindowTitle("Edit Purchase Entry")
        self.setModal(True)
        self.resize(500, 620)
        layout = QFormLayout()
        self.date_input = QDateEdit()
        self.date_input.setCalendarPopup(True)
        self.date_input.setDisplayFormat("MM/dd/yyyy")
        self.date_input.setDate(
            QDate.fromString(entry_data['date'], "MM/dd/yyyy") if entry_data else QDate.currentDate())
        layout.addRow("Date:", self.date_input)
        self.payee_input = QComboBox()
        self.payee_input.setEditable(True)
        self._load_payees()
        if entry_data:
            self.payee_input.setCurrentText(entry_data.get('payee_name', '') or '')
        layout.addRow("Payee Name:", self.payee_input)
        self.tin_input = QLineEdit()
        self.tin_input.setReadOnly(True)
        if entry_data:
            self.tin_input.setText(str(entry_data.get('tin', '') or ''))
        layout.addRow("TIN:", self.tin_input)
        self.payee_input.currentTextChanged.connect(self._on_payee_changed)
        self.branch_input = QLineEdit()
        if entry_data:
            self.branch_input.setText(str(entry_data.get('branch_code', '') or ''))
        layout.addRow("Branch Code:", self.branch_input)
        self.reference_input = QLineEdit()
        if entry_data:
            self.reference_input.setText(entry_data.get('reference_no', ''))
        layout.addRow("Reference No:", self.reference_input)
        self.account_des_input = QComboBox()
        self.account_des_input.setEditable(True)
        self.account_map = {}
        self._load_accounts()
        if entry_data:
            self.account_des_input.setCurrentText(entry_data.get('account_description', ''))
        self.account_des_input.currentTextChanged.connect(self._on_account_changed)
        layout.addRow("Account Description:", self.account_des_input)
        self.account_code_input = QLineEdit()
        self.account_code_input.setReadOnly(True)
        if entry_data:
            self.account_code_input.setText(entry_data.get('account_code', '') or '')
        layout.addRow("Account Code:", self.account_code_input)
        self.net_amount_input = QDoubleSpinBox()
        self.net_amount_input.setMaximum(99999999.99)
        self.net_amount_input.setDecimals(2)
        self.net_amount_input.setGroupSeparatorShown(True)
        if entry_data:
            self.net_amount_input.setValue(entry_data.get('net_amount', 0))
        self.net_amount_input.valueChanged.connect(self._calculate_totals)
        layout.addRow("Net Amount (Debit):", self.net_amount_input)
        self.input_vat_input = QDoubleSpinBox()
        self.input_vat_input.setMaximum(99999999.99)
        self.input_vat_input.setDecimals(2)
        self.input_vat_input.setGroupSeparatorShown(True)
        if entry_data:
            self.input_vat_input.setValue(entry_data.get('input_vat', 0))
        self.input_vat_input.valueChanged.connect(self._calculate_gross)
        layout.addRow("Input VAT:", self.input_vat_input)
        self.gross_amount_input = QDoubleSpinBox()
        self.gross_amount_input.setMaximum(99999999.99)
        self.gross_amount_input.setDecimals(2)
        self.gross_amount_input.setGroupSeparatorShown(True)
        self.gross_amount_input.setReadOnly(True)
        if entry_data:
            self.gross_amount_input.setValue(entry_data.get('gross_amount', 0))
        layout.addRow("Gross Amount:", self.gross_amount_input)
        self.particulars_input = QLineEdit()
        if entry_data:
            self.particulars_input.setText(entry_data.get('particulars', '') or '')
        layout.addRow("Particulars:", self.particulars_input)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)
        self.setLayout(layout)
        self._calculate_totals()
        self._on_account_changed()

    def _load_payees(self):
        self.payee_input.addItem("")
        alphalist = self.db_manager.get_all_alphalist()
        for entry in alphalist:
            if entry.get('entry_type', 'Customer') not in ('Vendor', 'Customer&Vendor'):
                continue
            name = entry.get('company_name') or \
                   f"{entry.get('first_name','')} {entry.get('last_name','')}".strip()
            if name:
                self.payee_input.addItem(name)
                self.alphalist_map[name] = {
                    'tin': entry.get('tin', ''),
                    'vat': entry.get('vat', 'VAT Regular'),
                }

    def _load_accounts(self):
        self.account_des_input.addItem("")
        accounts = self.db_manager.get_all_accounts()
        for account in accounts:
            desc = account['account_description']
            code = account['account_code']
            self.account_des_input.addItem(desc)
            self.account_map[desc] = code

    def _on_payee_changed(self, payee_name):
        info = self.alphalist_map.get(payee_name, {})
        self.tin_input.setText(str(info.get('tin', '')) if info else '')
        self._calculate_totals()

    def _on_account_changed(self):
        desc = self.account_des_input.currentText()
        self.account_code_input.setText(self.account_map.get(desc, ''))

    def _get_vat_rate(self) -> float:
        from ui.alphalist_widget import VAT_TYPES
        payee    = self.payee_input.currentText()
        vat_type = self.alphalist_map.get(payee, {}).get('vat', 'VAT Regular')
        return VAT_TYPES.get(vat_type, 0.12)

    def _calculate_totals(self):
        net      = self.net_amount_input.value()
        vat_rate = self._get_vat_rate()
        self.input_vat_input.blockSignals(True)
        self.input_vat_input.setValue(round(net * vat_rate, 2))
        self.input_vat_input.blockSignals(False)
        self._calculate_gross()

    def _calculate_gross(self):
        net = self.net_amount_input.value()
        vat = self.input_vat_input.value()
        self.gross_amount_input.setValue(round(net + vat, 2))

    def get_data(self):
        net = self.net_amount_input.value()
        return {
            'date':                self.date_input.date().toString("MM/dd/yyyy"),
            'payee_name':          self.payee_input.currentText().strip(),
            'reference_no':        self.reference_input.text().strip(),
            'tin':                 self.tin_input.text().strip(),
            'branch_code':         self.branch_input.text().strip(),
            'net_amount':          net,
            'input_vat':           self.input_vat_input.value(),
            'gross_amount':        self.gross_amount_input.value(),
            'account_description': self.account_des_input.currentText().strip(),
            'account_code':        self.account_code_input.text().strip(),
            'debit':               net,
            'credit':              0,
            'particulars':         self.particulars_input.text().strip(),
        }


class PurchaseJournalWidget(QWidget):
    def __init__(self, db_manager: DatabaseManager):
        super().__init__()
        self.db_manager = db_manager
        self.all_entries = []
        self._setup_ui()
        self._setup_shortcuts()
        self.load_data()

    def _setup_ui(self):
        layout = QVBoxLayout()
        title = QLabel("PURCHASE JOURNAL")
        title.setProperty("class", "title")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        search_group = QGroupBox("Search & Filter")
        search_layout = QHBoxLayout()
        self.month_combo = add_month_combo(search_layout)
        search_layout.addWidget(QLabel("Search:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search by payee, reference, TIN, account, or particulars...")
        self.search_input.setClearButtonEnabled(True)
        search_layout.addWidget(self.search_input)
        search_layout.addWidget(QLabel("From:"))
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDisplayFormat("MM/dd/yyyy")
        self.date_from.setDate(QDate(2000, 1, 1))
        search_layout.addWidget(self.date_from)
        search_layout.addWidget(QLabel("To:"))
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDisplayFormat("MM/dd/yyyy")
        self.date_to.setDate(QDate.currentDate())
        search_layout.addWidget(self.date_to)
        self.clear_filter_btn = QPushButton("Clear Filter")
        self.clear_filter_btn.clicked.connect(self._clear_filters)
        search_layout.addWidget(self.clear_filter_btn)
        self.results_label = QLabel("Showing: 0 of 0")
        search_layout.addWidget(self.results_label)
        search_group.setLayout(search_layout)
        layout.addWidget(search_group)

        button_layout = QHBoxLayout()
        self.add_btn = QPushButton("Add Entry")
        self.add_btn.clicked.connect(self._add_entry)
        button_layout.addWidget(self.add_btn)
        self.edit_btn = QPushButton("Edit Entry")
        self.edit_btn.clicked.connect(self._edit_entry)
        button_layout.addWidget(self.edit_btn)
        self.copy_btn = QPushButton("Copy Entry")
        self.copy_btn.clicked.connect(self._copy_entry)
        button_layout.addWidget(self.copy_btn)
        self.delete_btn = QPushButton("Delete Entry")
        self.delete_btn.setProperty("class", "danger")
        self.delete_btn.clicked.connect(self._delete_entry)
        button_layout.addWidget(self.delete_btn)
        self.import_btn = QPushButton("Import")
        self.import_btn.clicked.connect(self._import_xls)
        button_layout.addWidget(self.import_btn)
        self.export_btn = QPushButton("Export")
        self.export_btn.clicked.connect(self._export_xls)
        button_layout.addWidget(self.export_btn)
        button_layout.addStretch()
        self.totals_label = QLabel("Totals: Net: 0.00 | VAT: 0.00 | Gross: 0.00")
        self.totals_label.setProperty("class", "total")
        button_layout.addWidget(self.totals_label)
        layout.addLayout(button_layout)

        self.table = QTableWidget()
        self.table.setColumnCount(12)
        self.table.setHorizontalHeaderLabels([
            "Date", "Payee Name", "Ref No.", "TIN", "Branch",
            "Net Amt", "Input VAT", "Gross Amt",
            "Account Description", "Account Code", "Debit", "Particulars"])
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        self.table.setSortingEnabled(True)
        self.table.setAlternatingRowColors(True)
        layout.addWidget(self.table)

        # ── SearchFilter ─────────────────────────────────────────────
        self._search = SearchFilter(
            table         = self.table,
            search_input  = self.search_input,
            results_label = self.results_label,
            date_from     = self.date_from,
            date_to       = self.date_to,
            month_combo   = self.month_combo,
            date_col      = 0,
        )
        # Hook totals update
        orig_run = self._search._run
        def _run_with_totals():
            orig_run()
            self._update_totals_from_visible()
        self._search._timer.timeout.disconnect()
        self._search._timer.timeout.connect(_run_with_totals)
        self.date_from.dateChanged.connect(_run_with_totals)
        self.date_to.dateChanged.connect(_run_with_totals)
        self.month_combo.currentIndexChanged.connect(_run_with_totals)
        self._run_with_totals = _run_with_totals

        self.setLayout(layout)

    def _setup_shortcuts(self):
        QShortcut(QKeySequence("Ctrl+N"),       self).activated.connect(self._add_entry)
        QShortcut(QKeySequence("Ctrl+E"),       self).activated.connect(self._edit_entry)
        QShortcut(QKeySequence("Ctrl+Shift+C"), self).activated.connect(self._copy_entry)
        QShortcut(QKeySequence("Ctrl+D"),       self).activated.connect(self._delete_entry)
        QShortcut(QKeySequence("Ctrl+F"),       self).activated.connect(self.search_input.setFocus)
        QShortcut(QKeySequence("Ctrl+I"),       self).activated.connect(self._import_xls)
        QShortcut(QKeySequence("Ctrl+Shift+E"), self).activated.connect(self._export_xls)

    def load_data(self):
        self.all_entries = self.db_manager.get_purchase_journal()
        self._populate_table(self.all_entries)
        self._search.refresh()
        self._update_totals_from_visible()

    def _populate_table(self, entries):
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(entries))
        num_fields = ['net_amount', 'input_vat', 'gross_amount']
        for row, entry in enumerate(entries):
            date_item = _DateItem(entry.get('date', ''))
            date_item.setData(Qt.UserRole, entry['id'])
            self.table.setItem(row, 0, date_item)
            for text, col in [
                (entry.get('payee_name', '') or '',        1),
                (entry.get('reference_no', ''),            2),
                (str(entry.get('tin', '') or ''),          3),
                (str(entry.get('branch_code', '') or ''), 4),
            ]:
                self.table.setItem(row, col, QTableWidgetItem(text))
            for offset, field in enumerate(num_fields):
                val  = entry.get(field, 0)
                item = QTableWidgetItem(f"{val:,.2f}")
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                self.table.setItem(row, 5 + offset, item)
            self.table.setItem(row, 8,  QTableWidgetItem(entry.get('account_description', '') or ''))
            self.table.setItem(row, 9,  QTableWidgetItem(entry.get('account_code', '') or ''))
            debit_val = entry.get('debit', entry.get('net_amount', 0))
            debit_item = QTableWidgetItem(f"{debit_val:,.2f}")
            debit_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            debit_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
            self.table.setItem(row, 10, debit_item)
            self.table.setItem(row, 11, QTableWidgetItem(entry.get('particulars', '') or ''))
        self.table.setSortingEnabled(True)

    def _update_totals_from_visible(self):
        net = vat = gross = 0.0
        for row in range(self.table.rowCount()):
            if self.table.isRowHidden(row):
                continue
            try:
                net   += float(self.table.item(row, 5).text().replace(',', ''))
                vat   += float(self.table.item(row, 6).text().replace(',', ''))
                gross += float(self.table.item(row, 7).text().replace(',', ''))
            except (AttributeError, ValueError):
                pass
        self.totals_label.setText(
            f"Totals: Net: {net:,.2f} | VAT: {vat:,.2f} | Gross: {gross:,.2f}")

    def _clear_filters(self):
        self.search_input.clear()
        self.date_from.setDate(QDate(2000, 1, 1))
        self.date_to.setDate(QDate.currentDate())
        self.month_combo.setCurrentIndex(0)

    def _get_selected_entry_data(self):
        row = self.table.currentRow()
        if row < 0:
            return None
        return {
            'id':                  self.table.item(row, 0).data(Qt.UserRole),
            'date':                self.table.item(row, 0).text(),
            'payee_name':          self.table.item(row, 1).text(),
            'reference_no':        self.table.item(row, 2).text(),
            'tin':                 self.table.item(row, 3).text(),
            'branch_code':         self.table.item(row, 4).text(),
            'net_amount':          float(self.table.item(row, 5).text().replace(',', '')),
            'input_vat':           float(self.table.item(row, 6).text().replace(',', '')),
            'gross_amount':        float(self.table.item(row, 7).text().replace(',', '')),
            'account_description': self.table.item(row, 8).text(),
            'account_code':        self.table.item(row, 9).text(),
            'debit':               float(self.table.item(row, 10).text().replace(',', '')),
            'credit':              0,
            'particulars':         self.table.item(row, 11).text(),
        }

    def _add_entry(self):
        dialog = PurchaseJournalDialog(self.db_manager, self)
        if dialog.exec_():
            data = dialog.get_data()
            if data['reference_no'] and data['account_description']:
                if self.db_manager.add_purchase_entry(data):
                    self.load_data()
                    QMessageBox.information(self, "Success", "Purchase entry added successfully!")
                else:
                    QMessageBox.warning(self, "Error", "Failed to add entry!")

    def _edit_entry(self):
        entry_data = self._get_selected_entry_data()
        if not entry_data:
            QMessageBox.warning(self, "Warning", "Please select an entry to edit")
            return
        dialog = PurchaseJournalDialog(self.db_manager, self, entry_data)
        if dialog.exec_():
            if self.db_manager.update_purchase_entry(entry_data['id'], dialog.get_data()):
                self.load_data()
                QMessageBox.information(self, "Success", "Entry updated successfully!")
            else:
                QMessageBox.warning(self, "Error", "Failed to update entry!")

    def _copy_entry(self):
        entry_data = self._get_selected_entry_data()
        if not entry_data:
            QMessageBox.warning(self, "Warning", "Please select an entry to copy")
            return
        entry_data.pop('id', None)
        dialog = PurchaseJournalDialog(self.db_manager, self, entry_data, is_copy=True)
        if dialog.exec_():
            data = dialog.get_data()
            if data['reference_no'] and data['account_description']:
                if self.db_manager.add_purchase_entry(data):
                    self.load_data()
                    QMessageBox.information(self, "Success", "Entry copied successfully!")
                else:
                    QMessageBox.warning(self, "Error", "Failed to add copied entry!")

    def _delete_entry(self):
        entry_data = self._get_selected_entry_data()
        if not entry_data:
            QMessageBox.warning(self, "Warning", "Please select an entry to delete")
            return
        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Delete entry '{entry_data['reference_no']}'?",
            QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            if self.db_manager.delete_purchase_entry(entry_data['id']):
                self.load_data()
                QMessageBox.information(self, "Success", "Entry deleted successfully!")
            else:
                QMessageBox.warning(self, "Error", "Failed to delete entry!")

    def _export_xls(self):
        if not _OPENPYXL_OK:
            QMessageBox.critical(self, "Missing Library",
                                 "openpyxl is required.\nInstall with: pip install openpyxl")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Purchase Journal",
            os.path.join(get_io_dir("Purchase Journal"), "purchase_journal_report.xlsx"),
            "Excel Files (*.xlsx)")
        if not path:
            return
        rows = []
        for r in range(self.table.rowCount()):
            if not self.table.isRowHidden(r):
                row = {}
                for ci, (_, k) in enumerate(_XLS_COLUMNS):
                    item = self.table.item(r, ci)
                    row[k] = item.text() if item else ''
                rows.append(row)
        n, err = _export_to_xls(rows, path, "Purchase Journal")
        if err:
            QMessageBox.critical(self, "Export Failed", err)
        else:
            QMessageBox.information(self, "Export Successful",
                                    f"{n} record(s) exported to:\n{path}")

    def _import_xls(self):
        if not _OPENPYXL_OK:
            QMessageBox.critical(self, "Missing Library",
                                 "openpyxl is required.\nInstall with: pip install openpyxl")
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "Import Purchase Journal", get_import_dir(""),
            "Excel Files (*.xlsx *.xls)")
        if not path:
            return
        try:
            imported, skipped, errors = _import_from_xls(path, self.db_manager)
        except Exception as exc:
            QMessageBox.critical(self, "Import Failed", str(exc))
            return
        self.load_data()
        msg = f"Import complete.\n  Imported: {imported}\n  Skipped: {skipped}"
        if errors:
            msg += "\n\nDetails:\n" + "\n".join(errors[:20])
        QMessageBox.information(self, "Import Summary", msg)


def _export_to_xls(rows: list, path: str, sheet_title: str) -> tuple:
    from datetime import datetime as _dt
    try:
        wb  = Workbook()
        ws  = wb.active
        ws.title = sheet_title[:31]
        hdr_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        hdr_fill  = PatternFill('solid', start_color='2F5496')
        hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_font = Font(name='Arial', size=10)
        alt_fill  = PatternFill('solid', start_color='DCE6F1')
        thin      = Side(style='thin', color='B0B0B0')
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)
        t_font    = Font(name='Arial', bold=True, size=14)
        t_align   = Alignment(horizontal='left', vertical='center')
        HEADER_ROW = 5
        ws.row_dimensions[2].height = 22
        ws.merge_cells(f'A2:{get_column_letter(len(_XLS_HEADERS))}2')
        tc = ws['A2']
        tc.value = sheet_title.upper(); tc.font = t_font; tc.alignment = t_align
        ws.row_dimensions[3].height = 18
        ws.merge_cells(f'A3:{get_column_letter(len(_XLS_HEADERS))}3')
        sc = ws['A3']
        sc.value = f'For the Year {_dt.now().year}'
        sc.font = Font(name='Arial', italic=True, size=11); sc.alignment = t_align
        ws.row_dimensions[HEADER_ROW].height = 28
        for ci, hdr in enumerate(_XLS_HEADERS, 1):
            cell = ws.cell(row=HEADER_ROW, column=ci, value=hdr)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = hdr_align; cell.border = border
        NUM_KEYS = {'net_amount', 'input_vat', 'gross_amount', 'debit'}
        for ri, entry in enumerate(rows):
            row_idx = 6 + ri
            ws.row_dimensions[row_idx].height = 18
            fill = alt_fill if ri % 2 == 0 else None
            for ci, key in enumerate(_XLS_KEYS, 1):
                val  = entry.get(key, '') or ''
                cell = ws.cell(row=row_idx, column=ci, value=val)
                cell.font = cell_font; cell.border = border
                cell.alignment = Alignment(
                    horizontal='right' if key in NUM_KEYS else 'left', vertical='center')
                if fill:
                    cell.fill = fill
        widths = {1:12, 2:28, 3:16, 4:16, 5:12, 6:14, 7:14, 8:14, 9:28, 10:14, 11:14, 12:28}
        for ci, w in widths.items():
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = 'A6'
        last_col = get_column_letter(len(_XLS_HEADERS))
        ws.auto_filter.ref = f'A{HEADER_ROW}:{last_col}{HEADER_ROW}'
        wb.save(path)
        return len(rows), ''
    except Exception as exc:
        return 0, str(exc)


def _import_from_xls(path: str, db_manager) -> tuple:
    HEADER_MAP = {h.lower(): k for h, k in _XLS_COLUMNS}
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        raise RuntimeError(
            f'Cannot open "{os.path.basename(path)}".\n'
            'If this is a legacy .xls file, re-save as .xlsx first.')
    ws = wb.active
    col_index  = {}
    data_start = None
    for r_idx, row_vals in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        if not any(v and str(v).strip().lower() in HEADER_MAP for v in row_vals):
            continue
        for ci, cv in enumerate(row_vals):
            if cv is None:
                continue
            key = HEADER_MAP.get(str(cv).strip().lower())
            if key:
                col_index[key] = ci
        data_start = r_idx + 1
        break
    if not col_index:
        raise ValueError("Could not find a matching header row in the first 10 rows.")
    def _val(row_vals, key):
        idx = col_index.get(key)
        if idx is None or idx >= len(row_vals):
            return ''
        v = row_vals[idx]
        return str(v).strip() if v is not None else ''
    imported = skipped = 0
    errors   = []
    for rn, row_vals in enumerate(ws.iter_rows(min_row=data_start, values_only=True), data_start):
        if all(v is None for v in row_vals):
            continue
        try:
            def _f(key): return float(_val(row_vals, key).replace(',', '') or 0)
            data = {k: _val(row_vals, k) for k in _XLS_KEYS}
            data['net_amount']   = _f('net_amount')
            data['input_vat']    = _f('input_vat')
            data['gross_amount'] = _f('gross_amount')
            data['debit']        = _f('debit') or data['net_amount']
            data['credit']       = 0
            if not data.get('reference_no'):
                skipped += 1
                errors.append(f"Row {rn}: missing reference_no skipped")
                continue
            if db_manager.add_purchase_entry(data):
                imported += 1
            else:
                skipped += 1
                errors.append(f"Row {rn}: failed to insert skipped")
        except Exception as e:
            skipped += 1
            errors.append(f"Row {rn}: {e}")
    wb.close()
    return imported, skipped, errors