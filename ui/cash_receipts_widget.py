import os
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget,
                             QTableWidgetItem, QPushButton, QLabel, QLineEdit,
                             QDialog, QFormLayout, QDialogButtonBox, QMessageBox,
                             QHeaderView, QDateEdit, QDoubleSpinBox, QTextEdit,
                             QComboBox, QGroupBox, QShortcut, QFileDialog, QFrame)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QKeySequence, QFont
from database.db_manager import DatabaseManager
from ui.search_utils import SearchFilter, add_month_combo
from resources.file_paths import get_import_dir, get_io_dir

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

_XLS_COLUMNS = [
    ('Date',               'date'),
    ('Reference No',       'reference_no'),
    ('Particulars',        'particulars'),
    ('Account Description','account_description'),
    ('Account Code',       'account_code'),
    ('Debit',              'debit'),
    ('Credit',             'credit'),
]
_XLS_HEADERS = [h for h, _ in _XLS_COLUMNS]
_XLS_KEYS    = [k for _, k in _XLS_COLUMNS]

def _parse_date(date_str: str):
    for fmt in ("MM/dd/yyyy", "M/d/yyyy", "M/dd/yyyy", "MM/d/yyyy", "yyyy-MM-dd"):
        d = QDate.fromString(date_str, fmt)
        if d.isValid():
            return d
    return QDate()

class _DateItem(QTableWidgetItem):
    def __init__(self, display_text: str):
        super().__init__(display_text)
        qdate = _parse_date(display_text)
        self._sort_key = qdate.toString("yyyy-MM-dd") if qdate.isValid() else display_text
    def __lt__(self, other):
        if isinstance(other, _DateItem):
            return self._sort_key < other._sort_key
        return super().__lt__(other)


class _LineDialog(QDialog):
    def __init__(self, db_manager, parent=None, line_data=None):
        super().__init__(parent)
        self.db_manager  = db_manager
        self.account_map = {}
        self.setWindowTitle("Journal Line")
        self.setModal(True)
        self.setMinimumWidth(480)
        layout = QFormLayout()
        layout.setLabelAlignment(Qt.AlignRight)
        self.account_combo = QComboBox()
        self.account_combo.setEditable(True)
        self._load_accounts()
        self.account_combo.currentTextChanged.connect(self._on_account_changed)
        layout.addRow("Account Description:", self.account_combo)
        self.account_code_input = QLineEdit()
        self.account_code_input.setReadOnly(True)
        layout.addRow("Account Code:", self.account_code_input)
        self.debit_input = QDoubleSpinBox()
        self.debit_input.setMaximum(99_999_999.99)
        self.debit_input.setDecimals(2)
        self.debit_input.setGroupSeparatorShown(True)
        self.debit_input.valueChanged.connect(self._debit_changed)
        layout.addRow("Debit:", self.debit_input)
        self.credit_input = QDoubleSpinBox()
        self.credit_input.setMaximum(99_999_999.99)
        self.credit_input.setDecimals(2)
        self.credit_input.setGroupSeparatorShown(True)
        self.credit_input.valueChanged.connect(self._credit_changed)
        layout.addRow("Credit:", self.credit_input)
        if line_data:
            self.account_combo.setCurrentText(line_data.get('account_description', ''))
            self.debit_input.setValue(line_data.get('debit', 0))
            self.credit_input.setValue(line_data.get('credit', 0))
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self._validate_and_accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)
        self.setLayout(layout)
        self._on_account_changed(self.account_combo.currentText())

    def _load_accounts(self):
        self.account_combo.addItem("")
        for acct in self.db_manager.get_all_accounts():
            desc = acct['account_description']
            self.account_combo.addItem(desc)
            self.account_map[desc] = acct['account_code']

    def _on_account_changed(self, text):
        self.account_code_input.setText(self.account_map.get(text, ''))

    def _debit_changed(self, value):
        if value > 0:
            self.credit_input.blockSignals(True)
            self.credit_input.setValue(0)
            self.credit_input.blockSignals(False)

    def _credit_changed(self, value):
        if value > 0:
            self.debit_input.blockSignals(True)
            self.debit_input.setValue(0)
            self.debit_input.blockSignals(False)

    def _validate_and_accept(self):
        if not self.account_combo.currentText().strip():
            QMessageBox.warning(self, "Validation", "Please select an account.")
            return
        if self.debit_input.value() == 0 and self.credit_input.value() == 0:
            QMessageBox.warning(self, "Validation", "Enter a debit or credit amount.")
            return
        self.accept()

    def get_data(self):
        return {
            'account_description': self.account_combo.currentText().strip(),
            'account_code':        self.account_code_input.text().strip(),
            'debit':               self.debit_input.value(),
            'credit':              self.credit_input.value(),
        }


class CashReceiptsDialog(QDialog):
    def __init__(self, db_manager, parent=None, entry_data=None, is_copy=False):
        super().__init__(parent)
        self.db_manager = db_manager
        self.lines      = []
        if is_copy:
            self.setWindowTitle("Copy Entry Cash Receipts Journal")
        elif entry_data is None:
            self.setWindowTitle("New Journal Entry Cash Receipts Journal")
        else:
            self.setWindowTitle("Edit Entry Cash Receipts Journal")
        self.setModal(True)
        self.setMinimumWidth(820)
        self.setMinimumHeight(560)
        root = QVBoxLayout()
        root.setSpacing(10)
        hdr = QFormLayout()
        hdr.setLabelAlignment(Qt.AlignRight)
        self.date_input = QDateEdit()
        self.date_input.setCalendarPopup(True)
        self.date_input.setDisplayFormat("MM/dd/yyyy")
        self.date_input.setDate(
            QDate.fromString(entry_data['date'], "MM/dd/yyyy")
            if entry_data else QDate.currentDate())
        hdr.addRow("Date:", self.date_input)
        self.reference_input = QLineEdit()
        if entry_data:
            self.reference_input.setText(entry_data.get('reference_no', ''))
        hdr.addRow("Reference No.:", self.reference_input)
        self.particulars_input = QLineEdit()
        if entry_data:
            self.particulars_input.setText(entry_data.get('particulars', '') or '')
        hdr.addRow("Particulars:", self.particulars_input)
        root.addLayout(hdr)
        sep = QFrame(); sep.setFrameShape(QFrame.HLine); sep.setFrameShadow(QFrame.Sunken)
        root.addWidget(sep)
        lbl = QLabel("Journal Lines:"); lbl.setStyleSheet("font-weight: bold;")
        root.addWidget(lbl)
        self.lines_table = QTableWidget()
        self.lines_table.setColumnCount(4)
        self.lines_table.setHorizontalHeaderLabels(
            ["Account Description", "Account Code", "Debit", "Credit"])
        self.lines_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.lines_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.lines_table.setSelectionMode(QTableWidget.SingleSelection)
        self.lines_table.setAlternatingRowColors(True)
        self.lines_table.verticalHeader().setVisible(False)
        lhdr = self.lines_table.horizontalHeader()
        lhdr.setSectionResizeMode(0, QHeaderView.Stretch)
        lhdr.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        lhdr.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        lhdr.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        root.addWidget(self.lines_table)
        lbtn_row = QHBoxLayout()
        add_line_btn    = QPushButton("Add Line");    add_line_btn.setShortcut(QKeySequence("Ctrl+N"))
        edit_line_btn   = QPushButton("Edit Line");   edit_line_btn.setShortcut(QKeySequence("Ctrl+E"))
        remove_line_btn = QPushButton("Remove Line"); remove_line_btn.setShortcut(QKeySequence("Ctrl+D"))
        add_line_btn.clicked.connect(self._add_line)
        edit_line_btn.clicked.connect(self._edit_line)
        remove_line_btn.clicked.connect(self._remove_line)
        lbtn_row.addWidget(add_line_btn); lbtn_row.addWidget(edit_line_btn)
        lbtn_row.addWidget(remove_line_btn); lbtn_row.addStretch()
        root.addLayout(lbtn_row)
        tot_row = QHBoxLayout()
        bf = QFont(); bf.setBold(True)
        self.total_debit_lbl  = QLabel("Total Debit:  0.00")
        self.total_credit_lbl = QLabel("Total Credit: 0.00")
        self.balance_lbl      = QLabel("")
        self.total_debit_lbl.setFont(bf); self.total_credit_lbl.setFont(bf)
        tot_row.addStretch()
        tot_row.addWidget(self.total_debit_lbl); tot_row.addWidget(QLabel("  |  "))
        tot_row.addWidget(self.total_credit_lbl); tot_row.addWidget(QLabel("  "))
        tot_row.addWidget(self.balance_lbl)
        root.addLayout(tot_row)
        dlg_btns = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        dlg_btns.accepted.connect(self._save)
        dlg_btns.rejected.connect(self.reject)
        root.addWidget(dlg_btns)
        self.setLayout(root)
        if entry_data and entry_data.get('lines'):
            for ld in entry_data['lines']:
                self.lines.append(dict(ld))
            self._refresh_lines_table()

    def _add_line(self):
        dlg = _LineDialog(self.db_manager, self)
        if dlg.exec_():
            self.lines.append(dlg.get_data())
            self._refresh_lines_table()

    def _edit_line(self):
        row = self.lines_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "No Selection", "Select a line to edit."); return
        dlg = _LineDialog(self.db_manager, self, self.lines[row])
        if dlg.exec_():
            self.lines[row] = dlg.get_data()
            self._refresh_lines_table()

    def _remove_line(self):
        row = self.lines_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "No Selection", "Select a line to remove."); return
        del self.lines[row]
        self._refresh_lines_table()

    def _refresh_lines_table(self):
        self.lines_table.setRowCount(len(self.lines))
        td = tc = 0.0
        for r, ld in enumerate(self.lines):
            self.lines_table.setItem(r, 0, QTableWidgetItem(ld.get('account_description', '')))
            self.lines_table.setItem(r, 1, QTableWidgetItem(ld.get('account_code', '')))
            d = QTableWidgetItem(f"{ld.get('debit', 0):,.2f}")
            c = QTableWidgetItem(f"{ld.get('credit', 0):,.2f}")
            d.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            c.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.lines_table.setItem(r, 2, d); self.lines_table.setItem(r, 3, c)
            td += ld.get('debit', 0); tc += ld.get('credit', 0)
        self.total_debit_lbl.setText(f"Total Debit:  {td:,.2f}")
        self.total_credit_lbl.setText(f"Total Credit: {tc:,.2f}")
        balanced = abs(td - tc) < 0.005 and td > 0
        if balanced:
            self.balance_lbl.setText("Balanced"); self.balance_lbl.setStyleSheet("color: green; font-weight: bold;")
        elif self.lines:
            self.balance_lbl.setText("Unbalanced"); self.balance_lbl.setStyleSheet("color: red; font-weight: bold;")
        else:
            self.balance_lbl.setText("")

    def _save(self):
        if not self.reference_input.text().strip():
            QMessageBox.warning(self, "Validation", "Reference No. is required."); return
        if not self.lines:
            QMessageBox.warning(self, "Validation", "Add at least one journal line."); return
        td = sum(l.get('debit', 0)  for l in self.lines)
        tc = sum(l.get('credit', 0) for l in self.lines)
        if abs(td - tc) >= 0.005:
            reply = QMessageBox.question(
                self, "Unbalanced Entry",
                f"Debit ({td:,.2f}) ≠ Credit ({tc:,.2f}).\nSave anyway?",
                QMessageBox.Yes | QMessageBox.No)
            if reply != QMessageBox.Yes: return
        self.accept()

    def get_data(self):
        date    = self.date_input.date().toString("MM/dd/yyyy")
        ref     = self.reference_input.text().strip()
        particu = self.particulars_input.text().strip()
        return [{
            'date': date, 'reference_no': ref, 'particulars': particu,
            'account_description': ld.get('account_description', ''),
            'account_code':        ld.get('account_code', ''),
            'debit':               ld.get('debit', 0),
            'credit':              ld.get('credit', 0),
        } for ld in self.lines]


class _ViewDetailsDialog(QDialog):
    def __init__(self, parent, date, reference_no, particulars, lines):
        super().__init__(parent)
        self.setWindowTitle(f"View Details {reference_no}")
        self.setModal(True); self.setMinimumWidth(720); self.setMinimumHeight(720)
        root = QVBoxLayout()
        info = QLabel(
            f"<b>Date:</b> {date}    <b>Reference No.:</b> {reference_no}<br>"
            f"<b>Particulars:</b> {particulars or ''}")
        info.setStyleSheet("font-size: 12px; margin-bottom: 6px;")
        root.addWidget(info)
        sep = QFrame(); sep.setFrameShape(QFrame.HLine); sep.setFrameShadow(QFrame.Sunken)
        root.addWidget(sep)
        tbl = QTableWidget()
        tbl.setColumnCount(4)
        tbl.setHorizontalHeaderLabels(["Account Description", "Account Code", "Debit", "Credit"])
        tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        tbl.setAlternatingRowColors(True); tbl.verticalHeader().setVisible(False)
        th = tbl.horizontalHeader()
        th.setSectionResizeMode(0, QHeaderView.Stretch)
        th.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        th.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        th.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        tbl.setRowCount(len(lines))
        td = tc = 0.0
        for r, ld in enumerate(lines):
            tbl.setItem(r, 0, QTableWidgetItem(ld.get('account_description', '')))
            tbl.setItem(r, 1, QTableWidgetItem(ld.get('account_code', '')))
            d = QTableWidgetItem(f"{ld.get('debit', 0):,.2f}")
            c = QTableWidgetItem(f"{ld.get('credit', 0):,.2f}")
            d.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            c.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            tbl.setItem(r, 2, d); tbl.setItem(r, 3, c)
            td += ld.get('debit', 0); tc += ld.get('credit', 0)
        root.addWidget(tbl)
        tot_row = QHBoxLayout()
        bf = QFont(); bf.setBold(True)
        tdl = QLabel(f"Total Debit:  {td:,.2f}")
        tcl = QLabel(f"Total Credit: {tc:,.2f}")
        tdl.setFont(bf); tcl.setFont(bf)
        tot_row.addStretch()
        tot_row.addWidget(tdl); tot_row.addWidget(QLabel("  |  ")); tot_row.addWidget(tcl)
        root.addLayout(tot_row)
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        root.addWidget(close_btn, alignment=Qt.AlignRight)
        self.setLayout(root)


class CashReceiptsWidget(QWidget):
    def __init__(self, db_manager: DatabaseManager):
        super().__init__()
        self.db_manager  = db_manager
        self.all_entries = []
        self._setup_ui()
        self._setup_shortcuts()
        self.load_data()

    def _setup_ui(self):
        layout = QVBoxLayout()
        title = QLabel("CASH RECEIPTS JOURNAL")
        title.setProperty("class", "title")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        sg = QGroupBox("Search & Filter")
        sl = QHBoxLayout()
        self.month_combo = add_month_combo(sl)
        sl.addWidget(QLabel("Search:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search by reference, account, or particulars")
        self.search_input.setClearButtonEnabled(True)
        sl.addWidget(self.search_input)
        sl.addWidget(QLabel("From:"))
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDisplayFormat("MM/dd/yyyy")
        self.date_from.setDate(QDate(2000, 1, 1))
        sl.addWidget(self.date_from)
        sl.addWidget(QLabel("To:"))
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDisplayFormat("MM/dd/yyyy")
        self.date_to.setDate(QDate.currentDate())
        sl.addWidget(self.date_to)
        self.clear_filter_btn = QPushButton("Clear Filter")
        self.clear_filter_btn.clicked.connect(self._clear_filters)
        sl.addWidget(self.clear_filter_btn)
        self.results_label = QLabel("Showing: 0 of 0")
        sl.addWidget(self.results_label)
        sg.setLayout(sl)
        layout.addWidget(sg)

        br = QHBoxLayout()
        self.add_btn    = QPushButton("Add Entry");    self.add_btn.clicked.connect(self._add_entry);    br.addWidget(self.add_btn)
        self.edit_btn   = QPushButton("Edit Entry");   self.edit_btn.clicked.connect(self._edit_entry);   br.addWidget(self.edit_btn)
        self.copy_btn   = QPushButton("Copy Entry");   self.copy_btn.clicked.connect(self._copy_entry);   br.addWidget(self.copy_btn)
        self.view_btn   = QPushButton("View Details"); self.view_btn.clicked.connect(self._view_details); br.addWidget(self.view_btn)
        self.delete_btn = QPushButton("Delete Entry"); self.delete_btn.setProperty("class", "danger"); self.delete_btn.clicked.connect(self._delete_entry); br.addWidget(self.delete_btn)
        self.import_btn = QPushButton("Import");       self.import_btn.clicked.connect(self._import_xls); br.addWidget(self.import_btn)
        self.export_btn = QPushButton("Export");       self.export_btn.clicked.connect(self._export_xls); br.addWidget(self.export_btn)
        br.addStretch()
        self.totals_label = QLabel("Totals Debit: 0.00 | Credit: 0.00")
        self.totals_label.setProperty("class", "total")
        br.addWidget(self.totals_label)
        layout.addLayout(br)

        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels([
            "Date", "Reference No.", "Particulars", "Lines", "Total Debit", "Total Credit"])
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        self.table.setSortingEnabled(True)
        self.table.setAlternatingRowColors(True)
        self.table.doubleClicked.connect(self._view_details)
        layout.addWidget(self.table)

        # ── SearchFilter ─────────────────────────────────────────────
        self._search = SearchFilter(
            table         = self.table,
            search_input  = self.search_input,
            results_label = self.results_label,
            date_from     = self.date_from,
            date_to       = self.date_to,
            month_combo   = self.month_combo,
            date_col      = 0,
        )
        orig_run = self._search._run
        def _run_with_totals():
            orig_run()
            self._update_totals_from_visible()
        self._search._timer.timeout.disconnect()
        self._search._timer.timeout.connect(_run_with_totals)
        self.date_from.dateChanged.connect(_run_with_totals)
        self.date_to.dateChanged.connect(_run_with_totals)
        self.month_combo.currentIndexChanged.connect(_run_with_totals)
        self._run_with_totals = _run_with_totals

        self.setLayout(layout)

    def _setup_shortcuts(self):
        QShortcut(QKeySequence("Ctrl+N"),       self).activated.connect(self._add_entry)
        QShortcut(QKeySequence("Ctrl+E"),       self).activated.connect(self._edit_entry)
        QShortcut(QKeySequence("Ctrl+Shift+C"), self).activated.connect(self._copy_entry)
        QShortcut(QKeySequence("Ctrl+V"),       self).activated.connect(self._view_details)
        QShortcut(QKeySequence("Ctrl+D"),       self).activated.connect(self._delete_entry)
        QShortcut(QKeySequence("Ctrl+F"),       self).activated.connect(self.search_input.setFocus)
        QShortcut(QKeySequence("Ctrl+I"),       self).activated.connect(self._import_xls)
        QShortcut(QKeySequence("Ctrl+Shift+E"), self).activated.connect(self._export_xls)

    def load_data(self):
        self.all_entries = self.db_manager.get_cash_receipts_journal()
        self._populate_table(self._group_entries(self.all_entries))
        self._search.refresh()
        self._update_totals_from_visible()

    def _group_entries(self, entries):
        groups = {}; order = []
        for e in entries:
            key = (e.get('date', ''), e.get('reference_no', ''))
            if key not in groups:
                groups[key] = {
                    'date': e.get('date', ''), 'reference_no': e.get('reference_no', ''),
                    'particulars': e.get('particulars', '') or '', 'lines': [], 'ids': []
                }
                order.append(key)
            groups[key]['lines'].append({
                'account_description': e.get('account_description', ''),
                'account_code':        e.get('account_code', '') or '',
                'debit':               e.get('debit', 0),
                'credit':              e.get('credit', 0),
            })
            groups[key]['ids'].append(e['id'])
        return [groups[k] for k in order]

    def _populate_table(self, groups):
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(groups))
        for r, g in enumerate(groups):
            td = sum(l['debit']  for l in g['lines'])
            tc = sum(l['credit'] for l in g['lines'])
            date_item = _DateItem(g['date'])
            date_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
            date_item.setData(Qt.UserRole, g)
            self.table.setItem(r, 0, date_item)
            for c, text in enumerate([
                g['reference_no'], g['particulars'],
                str(len(g['lines'])), f"{td:,.2f}", f"{tc:,.2f}"
            ], start=1):
                item = QTableWidgetItem(text)
                item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                if c >= 3:
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table.setItem(r, c, item)
        self.table.setSortingEnabled(True)

    def _update_totals_from_visible(self):
        td = tc = 0.0
        for row in range(self.table.rowCount()):
            if self.table.isRowHidden(row): continue
            try:
                td += float(self.table.item(row, 4).text().replace(',', ''))
                tc += float(self.table.item(row, 5).text().replace(',', ''))
            except (AttributeError, ValueError): pass
        self.totals_label.setText(f"Totals Debit: {td:,.2f} | Credit: {tc:,.2f}")

    def _clear_filters(self):
        self.search_input.clear()
        self.date_from.setDate(QDate(2000, 1, 1))
        self.date_to.setDate(QDate.currentDate())
        self.month_combo.setCurrentIndex(0)

    def _get_selected_group(self):
        row = self.table.currentRow()
        if row < 0: return None
        item = self.table.item(row, 0)
        return item.data(Qt.UserRole) if item else None

    def _add_entry(self):
        dialog = CashReceiptsDialog(self.db_manager, self)
        if dialog.exec_():
            rows = dialog.get_data()
            ok = all(self.db_manager.add_cash_receipts_entry(r) for r in rows)
            self.load_data()
            (QMessageBox.information if ok else QMessageBox.warning)(
                self, "Success" if ok else "Error",
                "Entry added successfully!" if ok else "Some lines failed to save.")

    def _edit_entry(self):
        group = self._get_selected_group()
        if not group:
            QMessageBox.warning(self, "Warning", "Please select an entry to edit."); return
        dialog = CashReceiptsDialog(self.db_manager, self, group)
        if not dialog.exec_(): return
        new_rows = dialog.get_data()
        for old_id in group['ids']:
            self.db_manager.delete_cash_receipts_entry(old_id)
        ok = all(self.db_manager.add_cash_receipts_entry(r) for r in new_rows)
        self.load_data()
        (QMessageBox.information if ok else QMessageBox.warning)(
            self, "Success" if ok else "Error",
            "Entry updated successfully!" if ok else "Some lines failed to save.")

    def _copy_entry(self):
        group = self._get_selected_group()
        if not group:
            QMessageBox.warning(self, "Warning", "Please select an entry to copy."); return
        copy = {k: v for k, v in group.items() if k != 'ids'}
        dialog = CashReceiptsDialog(self.db_manager, self, copy, is_copy=True)
        if not dialog.exec_(): return
        rows = dialog.get_data()
        ok = all(self.db_manager.add_cash_receipts_entry(r) for r in rows)
        self.load_data()
        (QMessageBox.information if ok else QMessageBox.warning)(
            self, "Success" if ok else "Error",
            "Entry copied successfully!" if ok else "Some lines failed to save.")

    def _view_details(self):
        group = self._get_selected_group()
        if not group:
            QMessageBox.warning(self, "Warning", "Please select an entry to view."); return
        _ViewDetailsDialog(self, group['date'], group['reference_no'],
                           group['particulars'], group['lines']).exec_()

    def _delete_entry(self):
        group = self._get_selected_group()
        if not group:
            QMessageBox.warning(self, "Warning", "Please select an entry to delete."); return
        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Delete all {len(group['ids'])} line(s) for '{group['reference_no']}' ({group['date']})?"
            , QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            for old_id in group['ids']:
                self.db_manager.delete_cash_receipts_entry(old_id)
            self.load_data()
            QMessageBox.information(self, "Success", "Entry deleted successfully!")

    def _export_xls(self):
        if not _OPENPYXL_OK:
            QMessageBox.critical(self, "Missing Library", "openpyxl is required.\nInstall with: pip install openpyxl")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Cash Receipts Journal",
            os.path.join(get_io_dir("Cash Receipts"), "cash_receipts_journal_report.xlsx"),
            "Excel Files (*.xlsx)")
        if not path: return
        rows = []
        for r in range(self.table.rowCount()):
            if self.table.isRowHidden(r): continue
            g = self.table.item(r, 0).data(Qt.UserRole)
            for ld in g['lines']:
                rows.append({'date': g['date'], 'reference_no': g['reference_no'],
                             'particulars': g['particulars'],
                             'account_description': ld['account_description'],
                             'account_code': ld['account_code'],
                             'debit': f"{ld['debit']:,.2f}", 'credit': f"{ld['credit']:,.2f}"})
        n, err = _export_to_xls(rows, path, "Cash Receipts Journal")
        if err: QMessageBox.critical(self, "Export Failed", err)
        else:   QMessageBox.information(self, "Export Successful", f"{n} line(s) exported to:\n{path}")

    def _import_xls(self):
        if not _OPENPYXL_OK:
            QMessageBox.critical(self, "Missing Library", "openpyxl is required.\nInstall with: pip install openpyxl")
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "Import Cash Receipts Journal", get_import_dir(""), "Excel Files (*.xlsx *.xls)")
        if not path: return
        try:
            imported, skipped, errors = _import_from_xls(path, self.db_manager, 'add_cash_receipts_entry')
        except Exception as exc:
            QMessageBox.critical(self, "Import Failed", str(exc)); return
        self.load_data()
        msg = f"Import complete.\n  Imported: {imported}\n  Skipped: {skipped}"
        if errors: msg += "\n\nDetails:\n" + "\n".join(errors[:20])
        QMessageBox.information(self, "Import Summary", msg)


def _export_to_xls(rows, path, sheet_title):
    from datetime import datetime as _dt
    try:
        wb = Workbook(); ws = wb.active; ws.title = sheet_title[:31]
        hf  = XLFont(name='Arial', bold=True, color='FFFFFF', size=11)
        hfl = PatternFill('solid', start_color='2F5496')
        ha  = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cf  = XLFont(name='Arial', size=10)
        af  = PatternFill('solid', start_color='DCE6F1')
        th  = Side(style='thin', color='B0B0B0')
        bd  = Border(left=th, right=th, top=th, bottom=th)
        HR  = 5
        ws.merge_cells(f'A2:{get_column_letter(len(_XLS_HEADERS))}2')
        tc = ws['A2']; tc.value = sheet_title.upper()
        tc.font = XLFont(name='Arial', bold=True, size=14)
        tc.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f'A3:{get_column_letter(len(_XLS_HEADERS))}3')
        sc = ws['A3']; sc.value = f'For the Year {_dt.now().year}'
        sc.font = XLFont(name='Arial', italic=True, size=11)
        sc.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[HR].height = 28
        for ci, h in enumerate(_XLS_HEADERS, 1):
            cell = ws.cell(row=HR, column=ci, value=h)
            cell.font = hf; cell.fill = hfl; cell.alignment = ha; cell.border = bd
        NUM = {'debit', 'credit'}
        for ri, entry in enumerate(rows):
            ri2 = 6 + ri; ws.row_dimensions[ri2].height = 18
            fill = af if ri % 2 == 0 else None
            for ci, key in enumerate(_XLS_KEYS, 1):
                val = entry.get(key, '') or ''
                cell = ws.cell(row=ri2, column=ci, value=val)
                cell.font = cf; cell.border = bd
                cell.alignment = Alignment(horizontal='right' if key in NUM else 'left', vertical='center')
                if fill: cell.fill = fill
        for ci, w in {1:12, 2:16, 3:28, 4:28, 5:14, 6:14, 7:14}.items():
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = 'A6'
        ws.auto_filter.ref = f'A{HR}:{get_column_letter(len(_XLS_HEADERS))}{HR}'
        wb.save(path); return len(rows), ''
    except Exception as exc:
        return 0, str(exc)


def _import_from_xls(path, db_manager, add_method_name):
    import os as _os
    HEADER_MAP = {h.lower(): k for h, k in _XLS_COLUMNS}
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        raise RuntimeError(f'Cannot open "{_os.path.basename(path)}".')
    ws = wb.active; col_index = {}; data_start = None
    for r_idx, row_vals in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        if not any(v and str(v).strip().lower() in HEADER_MAP for v in row_vals): continue
        for ci, cv in enumerate(row_vals):
            if cv is None: continue
            key = HEADER_MAP.get(str(cv).strip().lower())
            if key: col_index[key] = ci
        data_start = r_idx + 1; break
    if not col_index: raise ValueError("Could not find a matching header row.")
    def _val(rv, k):
        i = col_index.get(k)
        if i is None or i >= len(rv): return ''
        v = rv[i]; return str(v).strip() if v is not None else ''
    imported = skipped = 0; errors = []
    add_fn = getattr(db_manager, add_method_name)
    for rn, rv in enumerate(ws.iter_rows(min_row=data_start, values_only=True), data_start):
        if all(v is None for v in rv): continue
        try:
            data = {k: _val(rv, k) for k in _XLS_KEYS}
            data['debit']  = float((data.get('debit', '')  or 0).replace(',', ''))
            data['credit'] = float((data.get('credit', '') or 0).replace(',', ''))
            if not data.get('reference_no'):
                skipped += 1; errors.append(f"Row {rn}: missing reference_no skipped"); continue
            if add_fn(data): imported += 1
            else: skipped += 1; errors.append(f"Row {rn}: failed to insert skipped")
        except Exception as e:
            skipped += 1; errors.append(f"Row {rn}: {e}")
    wb.close(); return imported, skipped, errors