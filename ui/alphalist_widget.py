import os
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget,
                             QTableWidgetItem, QPushButton, QLabel, QLineEdit,
                             QDialog, QFormLayout, QDialogButtonBox, QMessageBox,
                             QHeaderView, QGroupBox, QShortcut, QComboBox,
                             QFileDialog)
from PyQt5.QtCore import Qt, QRegExp
from PyQt5.QtGui import QKeySequence, QRegExpValidator
from resources.file_paths import get_import_dir, get_io_dir
from ui.search_utils import SearchFilter
from database.db_manager import DatabaseManager

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

ENTRY_TYPES = ['All List', 'Customer&Vendor', 'Customer', 'Vendor']

XLS_COLUMNS = [
    ('TIN',          'tin'),
    ('Entry Type',   'entry_type'),
    ('Company Name', 'company_name'),
    ('First Name',   'first_name'),
    ('Middle Name',  'middle_name'),
    ('Last Name',    'last_name'),
    ('Address 1',    'address1'),
    ('Address 2',    'address2'),
]
XLS_KEYS    = [k for _, k in XLS_COLUMNS]
XLS_HEADERS = [h for h, _ in XLS_COLUMNS]


def format_tin(raw: str) -> str:
    digits = ''.join(ch for ch in raw if ch.isdigit())
    if not digits:
        return ''
    digits = digits.zfill(9)[-9:]
    return f'{digits[:3]}-{digits[3:6]}-{digits[6:]}'


class AlphalistDialog(QDialog):
    def __init__(self, parent=None, entry_data=None, is_copy=False):
        super().__init__(parent)
        self.entry_data = entry_data
        self.is_copy    = is_copy
        if is_copy:
            self.setWindowTitle('Copy Entry (Create New)')
        elif entry_data is None:
            self.setWindowTitle('Add Entry')
        else:
            self.setWindowTitle('Edit Entry')
        self.setModal(True)
        self.resize(520, 420)
        layout = QFormLayout()
        layout.setLabelAlignment(Qt.AlignRight)

        self.entry_type_combo = QComboBox()
        self.entry_type_combo.addItems(['Customer&Vendor', 'Customer', 'Vendor'])
        if entry_data:
            idx = self.entry_type_combo.findText(entry_data.get('entry_type', 'Customer&Vendor'))
            if idx >= 0:
                self.entry_type_combo.setCurrentIndex(idx)
        layout.addRow('Entry Type:', self.entry_type_combo)

        self.tin_input = QLineEdit()
        self.tin_input.setMaxLength(9)
        tin_validator = QRegExpValidator(QRegExp(r'^\d{1,9}$'), self.tin_input)
        self.tin_input.setValidator(tin_validator)
        if entry_data:
            raw = entry_data.get('tin', '') or ''
            self.tin_input.setText(''.join(ch for ch in raw if ch.isdigit()))
        self.tin_input.editingFinished.connect(self._format_tin)
        layout.addRow('TIN (9 digits):', self.tin_input)

        self.company_input = QLineEdit()
        if entry_data:
            self.company_input.setText(entry_data.get('company_name', '') or '')
        self.company_input.textChanged.connect(self._on_company_changed)
        layout.addRow('Company Name:', self.company_input)

        self.first_name_input  = QLineEdit()
        self.middle_name_input = QLineEdit()
        self.last_name_input   = QLineEdit()
        if entry_data:
            self.first_name_input.setText(entry_data.get('first_name',  '') or '')
            self.middle_name_input.setText(entry_data.get('middle_name', '') or '')
            self.last_name_input.setText(entry_data.get('last_name',   '') or '')
        self.first_name_input.textChanged.connect(self._on_personal_changed)
        self.middle_name_input.textChanged.connect(self._on_personal_changed)
        self.last_name_input.textChanged.connect(self._on_personal_changed)
        layout.addRow('First Name:',  self.first_name_input)
        layout.addRow('Middle Name:', self.middle_name_input)
        layout.addRow('Last Name:',   self.last_name_input)

        self.address1_input = QLineEdit()
        self.address2_input = QLineEdit()
        if entry_data:
            self.address1_input.setText(entry_data.get('address1', '') or '')
            self.address2_input.setText(entry_data.get('address2', '') or '')
        layout.addRow('Address 1:', self.address1_input)
        layout.addRow('Address 2:', self.address2_input)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self._on_accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)
        self.setLayout(layout)
        self._apply_field_states()

    def _format_tin(self):
        raw = self.tin_input.text().strip()
        formatted = format_tin(raw)
        if formatted:
            self.tin_input.setValidator(None)
            self.tin_input.setMaxLength(11)
            self.tin_input.setText(formatted)

    def _on_company_changed(self):
        self._apply_field_states()

    def _on_personal_changed(self):
        first  = self.first_name_input.text().strip()
        middle = self.middle_name_input.text().strip()
        last   = self.last_name_input.text().strip()
        # Auto-compose display in company field when personal names are being used
        if first or middle or last:
            parts = [p for p in [first, middle, last] if p]
            composed = ' '.join(parts)
            self.company_input.blockSignals(True)
            self.company_input.setText(composed)
            self.company_input.blockSignals(False)
        self._apply_field_states()

    def _apply_field_states(self):
        has_company  = bool(self.company_input.text().strip())
        has_personal = bool(
            self.first_name_input.text().strip()
            or self.middle_name_input.text().strip()
            or self.last_name_input.text().strip())
        # Lock personal fields when company is manually typed (not auto-composed)
        # Lock company field whenever personal names are in use
        self._set_readonly(self.first_name_input,  has_company and not has_personal)
        self._set_readonly(self.middle_name_input, has_company and not has_personal)
        self._set_readonly(self.last_name_input,   has_company and not has_personal)
        self._set_readonly(self.company_input,     has_personal)

    @staticmethod
    def _set_readonly(field: QLineEdit, locked: bool):
        field.setReadOnly(locked)
        field.setStyleSheet(
            'QLineEdit { background-color: rgba(128,128,128,0.15); color: rgba(128,128,128,0.7); }'
            if locked else '')
        field.setToolTip('Clear the other name section first to edit this field.' if locked else '')

    def _on_accept(self):
        raw = self.tin_input.text().strip()
        formatted = format_tin(raw)
        if not formatted:
            QMessageBox.warning(self, 'Validation', 'TIN is required and must contain digits.')
            self.tin_input.setFocus()
            return
        self.tin_input.setValidator(None)
        self.tin_input.setMaxLength(11)
        self.tin_input.setText(formatted)
        self.accept()

    def get_data(self) -> dict:
        return {
            'tin':          self.tin_input.text().strip(),
            'company_name': self.company_input.text().strip(),
            'first_name':   self.first_name_input.text().strip(),
            'middle_name':  self.middle_name_input.text().strip(),
            'last_name':    self.last_name_input.text().strip(),
            'address1':     self.address1_input.text().strip(),
            'address2':     self.address2_input.text().strip(),
            'entry_type':   self.entry_type_combo.currentText(),
        }


class AlphalistWidget(QWidget):
    def __init__(self, db_manager: DatabaseManager):
        super().__init__()
        self.db_manager  = db_manager
        self.all_entries = []
        self._setup_ui()
        self._setup_shortcuts()
        self.load_data()

    def _setup_ui(self):
        layout = QVBoxLayout()
        title = QLabel('ALPHALIST\nVendors & Customers Database')
        title.setProperty('class', 'title')
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        search_group  = QGroupBox('Search & Filter')
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel('Show:'))
        self.list_type_combo = QComboBox()
        self.list_type_combo.addItems(ENTRY_TYPES)
        search_layout.addWidget(self.list_type_combo)
        search_layout.addWidget(QLabel('Search:'))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText('Search by TIN, name, or address...')
        self.search_input.setClearButtonEnabled(True)
        search_layout.addWidget(self.search_input)
        self.clear_search_btn = QPushButton('Clear')
        self.clear_search_btn.clicked.connect(self._clear_search)
        search_layout.addWidget(self.clear_search_btn)
        self.results_label = QLabel('Showing: 0 of 0')
        search_layout.addWidget(self.results_label)
        search_group.setLayout(search_layout)
        layout.addWidget(search_group)

        button_layout = QHBoxLayout()
        self.add_btn = QPushButton('Add Entry')
        self.add_btn.clicked.connect(self._add_entry)
        button_layout.addWidget(self.add_btn)
        self.edit_btn = QPushButton('Edit Entry')
        self.edit_btn.clicked.connect(self._edit_entry)
        button_layout.addWidget(self.edit_btn)
        self.copy_btn = QPushButton('Copy Entry')
        self.copy_btn.clicked.connect(self._copy_entry)
        button_layout.addWidget(self.copy_btn)
        self.delete_btn = QPushButton('Delete Entry')
        self.delete_btn.setProperty('class', 'danger')
        self.delete_btn.clicked.connect(self._delete_entry)
        button_layout.addWidget(self.delete_btn)
        self.import_btn = QPushButton('Import File')
        self.import_btn.clicked.connect(self._import_xls)
        button_layout.addWidget(self.import_btn)
        self.export_btn = QPushButton('Export Report')
        self.export_btn.clicked.connect(self._export_xls)
        button_layout.addWidget(self.export_btn)
        button_layout.addStretch()
        layout.addLayout(button_layout)

        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels([
            'TIN', 'Entry Type', 'Company Name', 'First Name',
            'Last Name', 'Address 1', 'Address 2'])
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        self.table.setSortingEnabled(True)
        self.table.setAlternatingRowColors(True)
        layout.addWidget(self.table)
        self.setLayout(layout)

        self._search = SearchFilter(
            table         = self.table,
            search_input  = self.search_input,
            results_label = self.results_label,
        )
        self.list_type_combo.currentTextChanged.connect(self.load_data)

    def _setup_shortcuts(self):
        QShortcut(QKeySequence('Ctrl+N'),       self).activated.connect(self._add_entry)
        QShortcut(QKeySequence('Ctrl+E'),       self).activated.connect(self._edit_entry)
        QShortcut(QKeySequence('Ctrl+Shift+C'), self).activated.connect(self._copy_entry)
        QShortcut(QKeySequence('Ctrl+D'),       self).activated.connect(self._delete_entry)
        QShortcut(QKeySequence('Ctrl+F'),       self).activated.connect(lambda: self.search_input.setFocus())
        QShortcut(QKeySequence('Ctrl+L'),       self).activated.connect(self._toggle_list_type_filter)
        QShortcut(QKeySequence('Ctrl+Shift+I'), self).activated.connect(self._import_xls)
        QShortcut(QKeySequence('Ctrl+Shift+E'), self).activated.connect(self._export_xls)

    def _toggle_list_type_filter(self):
        current = self.list_type_combo.currentText()
        cycle = {'All List': 'Customer&Vendor', 'Customer&Vendor': 'Customer',
                 'Customer': 'Vendor', 'Vendor': 'All List'}
        self.list_type_combo.setCurrentText(cycle.get(current, 'All List'))

    def load_data(self):
        list_type = self.list_type_combo.currentText()
        self.all_entries = (self.db_manager.get_all_alphalist()
                            if list_type == 'All List'
                            else self.db_manager.get_all_alphalist(list_type))
        self._populate_table(self.all_entries)
        self._search.refresh()

    def _populate_table(self, entries):
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(entries))
        for row, entry in enumerate(entries):
            cols = [
                entry.get('tin',          '') or '',
                entry.get('entry_type',   'Customer&Vendor') or 'Customer&Vendor',
                entry.get('company_name', '') or '',
                entry.get('first_name',   '') or '',
                entry.get('last_name',    '') or '',
                entry.get('address1',     '') or '',
                entry.get('address2',     '') or '',
            ]
            for col, text in enumerate(cols):
                item = QTableWidgetItem(text)
                item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                if col == 0:
                    item.setData(Qt.UserRole,     entry['id'])
                    item.setData(Qt.UserRole + 1, entry.get('entry_type', 'Customer&Vendor'))
                    item.setData(Qt.UserRole + 2, entry.get('middle_name', '') or '')
                self.table.setItem(row, col, item)
        self.table.setSortingEnabled(True)

    def _clear_search(self):
        self.search_input.clear()
        self.list_type_combo.setCurrentIndex(0)

    def _add_entry(self):
        dialog = AlphalistDialog(self)
        if dialog.exec_():
            data = dialog.get_data()
            if data['tin']:
                if self.db_manager.add_alphalist(data):
                    self.load_data()
                    self.search_input.clear()
                    QMessageBox.information(self, 'Success', 'Entry added successfully!')
                else:
                    QMessageBox.warning(self, 'Error', 'TIN already exists!')

    def _edit_entry(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.warning(self, 'Warning', 'Please select an entry to edit.')
            return
        entry_id    = self.table.item(row, 0).data(Qt.UserRole)
        entry_type  = self.table.item(row, 0).data(Qt.UserRole + 1)
        middle_name = self.table.item(row, 0).data(Qt.UserRole + 2)
        entry_data  = {
            'id':           entry_id,
            'tin':          self.table.item(row, 0).text(),
            'entry_type':   entry_type,
            'company_name': self.table.item(row, 2).text(),
            'first_name':   self.table.item(row, 3).text(),
            'middle_name':  middle_name,
            'last_name':    self.table.item(row, 4).text(),
            'address1':     self.table.item(row, 5).text(),
            'address2':     self.table.item(row, 6).text(),
        }
        dialog = AlphalistDialog(self, entry_data)
        if dialog.exec_():
            data = dialog.get_data()
            if data['tin']:
                if self.db_manager.update_alphalist(entry_id, data):
                    self.load_data()
                    QMessageBox.information(self, 'Success', 'Entry updated successfully!')
                else:
                    QMessageBox.warning(self, 'Error', 'Failed to update entry!')

    def _copy_entry(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.warning(self, 'Warning', 'Please select an entry to copy.')
            return
        entry_data = {
            'tin':          self.table.item(row, 0).text(),
            'entry_type':   self.table.item(row, 0).data(Qt.UserRole + 1),
            'company_name': self.table.item(row, 2).text(),
            'first_name':   self.table.item(row, 3).text(),
            'middle_name':  self.table.item(row, 0).data(Qt.UserRole + 2),
            'last_name':    self.table.item(row, 4).text(),
            'address1':     self.table.item(row, 5).text(),
            'address2':     self.table.item(row, 6).text(),
        }
        dialog = AlphalistDialog(self, entry_data, is_copy=True)
        if dialog.exec_():
            data = dialog.get_data()
            if data['tin']:
                if self.db_manager.add_alphalist(data):
                    self.load_data()
                    self.search_input.clear()
                    QMessageBox.information(self, 'Success', 'Entry copied successfully!')
                else:
                    QMessageBox.warning(self, 'Error', 'TIN already exists!')

    def _delete_entry(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.warning(self, 'Warning', 'Please select an entry to delete.')
            return
        tin   = self.table.item(row, 0).text()
        reply = QMessageBox.question(
            self, 'Confirm Delete',
            f"Are you sure you want to delete entry with TIN '{tin}'?",
            QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            entry_id = self.table.item(row, 0).data(Qt.UserRole)
            if self.db_manager.delete_alphalist(entry_id):
                self.load_data()
                QMessageBox.information(self, 'Success', 'Entry deleted successfully!')
            else:
                QMessageBox.warning(self, 'Error', 'Failed to delete entry!')

    def _export_xls(self):
        if not _OPENPYXL_OK:
            QMessageBox.critical(self, 'Missing Library',
                                 'openpyxl is required.\nInstall with: pip install openpyxl')
            return
        path, _ = QFileDialog.getSaveFileName(
            self, 'Export Report to Excel',
            os.path.join(get_io_dir("Alphalist"), 'alphalist_report.xlsx'),
            'Excel Files (*.xlsx)')
        if not path:
            return
        rows = []
        for row in range(self.table.rowCount()):
            if self.table.isRowHidden(row):
                continue
            rows.append({
                'tin':          self.table.item(row, 0).text(),
                'entry_type':   self.table.item(row, 0).data(Qt.UserRole + 1) or 'Customer&Vendor',
                'company_name': self.table.item(row, 2).text(),
                'first_name':   self.table.item(row, 3).text(),
                'middle_name':  self.table.item(row, 0).data(Qt.UserRole + 2) or '',
                'last_name':    self.table.item(row, 4).text(),
                'address1':     self.table.item(row, 5).text(),
                'address2':     self.table.item(row, 6).text(),
            })
        count, err = export_alphalist_to_xls(rows, path)
        if err:
            QMessageBox.critical(self, 'Export Failed', err)
        else:
            QMessageBox.information(self, 'Export Successful',
                                    f'{count} record(s) exported to:\n{path}')

    def _import_xls(self):
        if not _OPENPYXL_OK:
            QMessageBox.critical(self, 'Missing Library',
                                 'openpyxl is required.\nInstall with: pip install openpyxl')
            return
        path, _ = QFileDialog.getOpenFileName(
            self, 'Import File', get_import_dir(""), 'Excel Files (*.xlsx *.xls)')
        if not path:
            return
        try:
            result = import_alphalist_from_xls(path, self.db_manager)
        except Exception as exc:
            QMessageBox.critical(self, 'Import Failed', str(exc))
            return
        self.load_data()
        msg = (f'Import complete.\n\n'
               f'  Imported:                {result["imported"]}\n'
               f'  Duplicate TINs skipped:  {result["skipped_duplicate"]}\n'
               f'  Invalid TINs skipped:    {result["skipped_invalid"]}')
        if result['errors']:
            detail = '\n'.join(result['errors'][:20])
            if len(result['errors']) > 20:
                detail += f'\n...and {len(result["errors"]) - 20} more.'
            msg += f'\n\nDetails:\n{detail}'
        if result['skipped_duplicate'] or result['skipped_invalid']:
            QMessageBox.warning(self, 'Import Summary', msg)
        else:
            QMessageBox.information(self, 'Import Summary', msg)


def export_alphalist_to_xls(entries: list, path: str, year: int = None):
    from datetime import datetime as _dt
    _year = year if year else _dt.now().year
    HEADER_ROW = 5
    DATA_START  = 6
    try:
        wb = Workbook(); ws = wb.active; ws.title = 'Alphalist'
        hf  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        hfl = PatternFill('solid', start_color='2F5496')
        ha  = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cf  = Font(name='Arial', size=10)
        cl  = Alignment(horizontal='left',   vertical='center')
        cc  = Alignment(horizontal='center', vertical='center')
        af  = PatternFill('solid', start_color='DCE6F1')
        th  = Side(style='thin', color='B0B0B0')
        bd  = Border(left=th, right=th, top=th, bottom=th)

        ws.row_dimensions[2].height = 22
        ws.merge_cells(f'A2:{get_column_letter(len(XLS_HEADERS))}2')
        ws['A2'].value = 'ALPHALIST'
        ws['A2'].font  = Font(name='Arial', bold=True, size=14)
        ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[3].height = 18
        ws.merge_cells(f'A3:{get_column_letter(len(XLS_HEADERS))}3')
        ws['A3'].value = f'For the Year {_year}'
        ws['A3'].font  = Font(name='Arial', italic=True, size=11)
        ws['A3'].alignment = Alignment(horizontal='left', vertical='center')

        ws.row_dimensions[HEADER_ROW].height = 28
        for ci, hdr in enumerate(XLS_HEADERS, 1):
            cell = ws.cell(row=HEADER_ROW, column=ci, value=hdr)
            cell.font = hf; cell.fill = hfl; cell.alignment = ha; cell.border = bd

        for ri, entry in enumerate(entries):
            row_idx = DATA_START + ri
            ws.row_dimensions[row_idx].height = 18
            fill = af if ri % 2 == 0 else None
            for ci, key in enumerate(XLS_KEYS, 1):
                val  = entry.get(key, '') or ''
                cell = ws.cell(row=row_idx, column=ci, value=val)
                cell.font      = cf
                cell.alignment = cc if ci in (1, 2) else cl
                cell.border    = bd
                if fill: cell.fill = fill

        widths = {1:16, 2:16, 3:28, 4:18, 5:18, 6:18, 7:30, 8:30}
        for ci, w in widths.items():
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = f'A{DATA_START}'
        ws.auto_filter.ref = f'A{HEADER_ROW}:{get_column_letter(len(XLS_HEADERS))}{HEADER_ROW}'
        wb.save(path)
        return len(entries), ''
    except Exception as exc:
        return 0, str(exc)


def import_alphalist_from_xls(path: str, db_manager) -> dict:
    summary = {'imported': 0, 'skipped_duplicate': 0, 'skipped_invalid': 0, 'errors': []}
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        raise RuntimeError(
            f'Cannot open "{os.path.basename(path)}".\n\n'
            'If this is an old .xls file, please re-save it as .xlsx first.')
    ws = wb.active
    HEADER_TO_KEY = {h.lower(): k for h, k in XLS_COLUMNS}
    HEADER_TO_KEY.update({
        'address 2': 'address2', 'address 1': 'address1',
        'company': 'company_name', 'first': 'first_name',
        'middle': 'middle_name', 'last': 'last_name',
        'entry type': 'entry_type',
    })
    col_index = {}; data_start_row = None
    for r_idx, row_vals in enumerate(
            ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if not any(v and str(v).strip().upper() == 'TIN' for v in row_vals):
            continue
        for col_i, cell_val in enumerate(row_vals):
            if cell_val is None: continue
            norm = str(cell_val).strip().lower()
            key  = HEADER_TO_KEY.get(norm)
            if key: col_index[key] = col_i
        data_start_row = r_idx + 1
        break
    if not col_index or 'tin' not in col_index:
        raise ValueError("Could not find a header row with a 'TIN' column in the first 10 rows.")

    def _val(row_vals, key: str) -> str:
        idx = col_index.get(key)
        if idx is None or idx >= len(row_vals): return ''
        v = row_vals[idx]
        return str(v).strip() if v is not None else ''

    for row_num, row_vals in enumerate(
            ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        if all(v is None for v in row_vals): continue
        raw_tin = _val(row_vals, 'tin')
        tin     = format_tin(raw_tin)
        if not tin:
            summary['skipped_invalid'] += 1
            summary['errors'].append(f'Row {row_num}: invalid TIN "{raw_tin}"')
            continue

        # Default to Customer&Vendor if no entry_type column or unrecognized value
        entry_type = _val(row_vals, 'entry_type') or 'Customer&Vendor'
        if entry_type not in ('Customer&Vendor', 'Customer', 'Vendor'):
            entry_type = 'Customer&Vendor'

        data = {
            'tin':          tin,
            'company_name': _val(row_vals, 'company_name'),
            'first_name':   _val(row_vals, 'first_name'),
            'middle_name':  _val(row_vals, 'middle_name'),
            'last_name':    _val(row_vals, 'last_name'),
            'address1':     _val(row_vals, 'address1'),
            'address2':     _val(row_vals, 'address2'),
            'entry_type':   entry_type,
        }
        if db_manager.add_alphalist(data):
            summary['imported'] += 1
        else:
            summary['skipped_duplicate'] += 1
            summary['errors'].append(f'Row {row_num}: duplicate TIN "{tin}" — skipped')
    wb.close()
    return summary