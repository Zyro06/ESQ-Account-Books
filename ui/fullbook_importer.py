"""
fullbook_importer.py
--------------------
Imports a multi-sheet Excel workbook into the current database in one pass.

Supported sheets:
  COA / Chart of Accounts  → chart_of_accounts
  Alphalist                → alphalist
  SJ_xx  / Sales           → sales_journal
  PJ_xx  / Purchase        → purchase_journal
  CDJ_xx / Cash Disbursement → cash_disbursement_journal
  CRJ_xx / Cash Receipt(s) → cash_receipts_journal
"""

from __future__ import annotations
from utils.date_utils   import norm_date  as _norm_date
from utils.date_utils   import norm_float as _norm_float
from utils.date_utils   import norm_str   as _norm_str
from utils.import_utils import find_header_row as _find_header_row
from utils.import_utils import col_index       as _col
from utils.import_utils import get_cell        as _get

import os
import re
from datetime import datetime, date as _date
from typing import Any

try:
    from openpyxl import load_workbook
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False


# ---------------------------------------------------------------------------
# TIN normalisation
# ---------------------------------------------------------------------------

def _normalize_tin(raw) -> str:
    """
    Convert any TIN-like value to xxx-xxx-xxx[-xxx] format,
    zero-padded to 9 digits (+ optional 3-digit branch = 12 total).

    Examples:
      1            → 000-000-001
      1111111      → 001-111-111
      123456789    → 123-456-789
      123456789000 → 123-456-789-000
      None / ''    → ''
    """
    if raw is None:
        return ''
    s = str(raw).strip()
    digits = re.sub(r'[^0-9]', '', s)
    if not digits:
        return ''

    if len(digits) <= 9:
        digits = digits.zfill(9)
        return f'{digits[:3]}-{digits[3:6]}-{digits[6:]}'
    if len(digits) <= 12:
        digits = digits.zfill(12)
        return f'{digits[:3]}-{digits[3:6]}-{digits[6:9]}-{digits[9:]}'
    # Longer than 12 — take last 12 digits as best effort
    digits = digits[-12:]
    return f'{digits[:3]}-{digits[3:6]}-{digits[6:9]}-{digits[9:]}'

# ---------------------------------------------------------------------------
# Public result dataclass (plain dict for simplicity)
# ---------------------------------------------------------------------------

def _empty_result(sheet: str) -> dict:
    return {'sheet': sheet, 'imported': 0, 'skipped': 0, 'errors': []}

# ---------------------------------------------------------------------------
# Sheet name matcher
# ---------------------------------------------------------------------------

def _classify_sheet(name: str) -> str | None:
    n = name.strip().lower()
    if n in ('coa', 'chart of accounts', 'chart_of_accounts'):
        return 'coa'
    if 'alphalist' in n:
        return 'alphalist'
    if n.startswith('sj') or n in ('sales', 'sales journal', 'sales_journal'):
        return 'sj'
    if n.startswith('pj') or n in ('purchase', 'purchase journal', 'purchase_journal'):
        return 'pj'
    if n.startswith('cdj') or 'disbursement' in n:
        return 'cdj'
    if n.startswith('crj') or 'receipt' in n:
        return 'crj'
    return None

# ---------------------------------------------------------------------------
# Per-sheet importers
# ---------------------------------------------------------------------------

def _import_coa(ws, db_manager) -> dict:
    res = _empty_result('COA')
    col_map, data_start = _find_header_row(
        ws, ['account code', 'account description'])
    if data_start is None:
        res['errors'].append('Header row not found.')
        return res

    ci_code = _col(col_map, 'account code', 'code')
    ci_desc = _col(col_map, 'account description', 'description')
    ci_nb   = _col(col_map, 'debit/credit', 'normal balance', 'debit', 'credit')

    for rn, rv in enumerate(
            ws.iter_rows(min_row=data_start, values_only=True), data_start):
        if all(v is None for v in rv):
            continue
        code = _norm_str(_get(rv, ci_code))
        desc = _norm_str(_get(rv, ci_desc))
        if not code or not desc:
            continue
        if code.lower() in ('account code', 'nan', 'none'):
            continue

        if ci_nb is not None:
            nb_raw = _norm_str(_get(rv, ci_nb)).upper()
            nb = 'Credit' if 'CREDIT' in nb_raw else 'Debit'
        else:
            nb = _infer_nb(code, desc)

        ok = db_manager.add_account(
            {'account_code': code, 'account_description': desc, 'normal_balance': nb}
        )
        if ok:
            res['imported'] += 1
        else:
            res['skipped'] += 1

    return res


def _infer_nb(code: str, desc: str) -> str:
    desc_upper = desc.upper()
    if any(kw in desc_upper for kw in ('ACCUM DEP', 'ACCUM. DEP',
                                        'ACCUMULATED DEP', 'ACCUM AMORT')):
        return 'Credit'
    if any(kw in desc_upper for kw in ('DRAWING', 'DRAWINGS')):
        return 'Debit'
    digits = ''.join(ch for ch in code if ch.isdigit())
    if not digits:
        return 'Debit'
    first = digits[0]
    if first in ('2', '3', '4', '7'):
        return 'Credit'
    if first == '4' and any(kw in desc_upper for kw in ('DISCOUNT', 'RETURN', 'ALLOWANCE')):
        return 'Debit'
    return 'Debit'


def _import_alphalist(ws, db_manager) -> dict:
    res = _empty_result('Alphalist')
    col_map, data_start = _find_header_row(ws, ['tin'])
    if data_start is None:
        res['errors'].append('Header row not found.')
        return res

    ci_tin     = _col(col_map, 'tin')
    ci_company = _col(col_map, 'company name', 'company')
    ci_first   = _col(col_map, 'first name', 'first')
    ci_middle  = _col(col_map, 'middle name', 'middle')
    ci_last    = _col(col_map, 'last name', 'last')
    ci_addr1   = _col(col_map, 'address1', 'address 1', 'address_1')
    ci_addr2   = _col(col_map, 'address2', 'address 2', 'address_2')
    ci_etype   = _col(col_map, 'entry type', 'entry_type')

    ETYPE_VALID = {'Customer&Vendor', 'Customer', 'Vendor'}

    for rn, rv in enumerate(
            ws.iter_rows(min_row=data_start, values_only=True), data_start):
        if all(v is None for v in rv):
            continue

        raw_tin = _norm_str(_get(rv, ci_tin))
        tin     = _normalize_tin(raw_tin)
        if not tin:
            res['skipped'] += 1
            res['errors'].append(f'Row {rn}: invalid TIN "{raw_tin}" — skipped')
            continue

        etype = _norm_str(_get(rv, ci_etype)) or 'Customer&Vendor'
        if etype not in ETYPE_VALID:
            etype = 'Customer&Vendor'

        data = {
            'tin':          tin,
            'company_name': _norm_str(_get(rv, ci_company)),
            'first_name':   _norm_str(_get(rv, ci_first)),
            'middle_name':  _norm_str(_get(rv, ci_middle)),
            'last_name':    _norm_str(_get(rv, ci_last)),
            'address1':     _norm_str(_get(rv, ci_addr1)),
            'address2':     _norm_str(_get(rv, ci_addr2)),
            'entry_type':   etype,
        }

        if db_manager.add_alphalist(data):
            res['imported'] += 1
        else:
            res['skipped'] += 1
            res['errors'].append(f'Row {rn}: duplicate TIN "{tin}" — skipped')

    return res


def _import_sj(ws, db_manager) -> dict:
    res = _empty_result('Sales Journal')
    col_map, data_start = _find_header_row(
        ws, ['date', 'customer name', 'reference no'])
    if data_start is None:
        res['errors'].append('Header row not found.')
        return res

    ci_date  = _col(col_map, 'date')
    ci_cust  = _col(col_map, 'customer name', 'customer')
    ci_ref   = _col(col_map, 'reference no', 'reference_no', 'ref no', 'ref')
    ci_tin   = _col(col_map, 'tin')
    ci_net   = _col(col_map, 'net amount', 'net')
    ci_vat   = _col(col_map, 'output vat', 'output_vat', 'vat')
    ci_gross = _col(col_map, 'gross amount', 'gross')
    ci_goods = _col(col_map, 'goods')
    ci_svc   = _col(col_map, 'services', 'service')
    ci_part  = _col(col_map, 'particulars')

    for rn, rv in enumerate(
            ws.iter_rows(min_row=data_start, values_only=True), data_start):
        if all(v is None for v in rv):
            continue

        dt  = _norm_date(_get(rv, ci_date))
        ref = _norm_str(_get(rv, ci_ref))
        if not dt or not ref:
            res['skipped'] += 1
            res['errors'].append(
                f'Row {rn}: missing {"date" if not dt else "reference_no"} — skipped')
            continue

        net   = _norm_float(_get(rv, ci_net))
        vat   = _norm_float(_get(rv, ci_vat))
        gross = _norm_float(_get(rv, ci_gross)) or round(net + vat, 2)
        goods = _norm_float(_get(rv, ci_goods))
        svc   = _norm_float(_get(rv, ci_svc))

        data = {
            'date':          dt,
            'customer_name': _norm_str(_get(rv, ci_cust)),
            'reference_no':  ref,
            'tin':           _normalize_tin(_get(rv, ci_tin)),   # ← normalised
            'net_amount':    net,
            'output_vat':    vat,
            'gross_amount':  gross,
            'goods':         goods,
            'services':      svc,
            'particulars':   _norm_str(_get(rv, ci_part)),
        }

        if db_manager.add_sales_entry(data):
            res['imported'] += 1
        else:
            res['skipped'] += 1
            res['errors'].append(f'Row {rn}: insert failed — skipped')

    return res


def _import_pj(ws, db_manager) -> dict:
    res = _empty_result('Purchase Journal')
    col_map, data_start = _find_header_row(
        ws, ['date', 'payee name', 'reference no'])
    if data_start is None:
        res['errors'].append('Header row not found.')
        return res

    ci_date   = _col(col_map, 'date')
    ci_payee  = _col(col_map, 'payee name', 'payee')
    ci_ref    = _col(col_map, 'reference no', 'reference_no', 'ref no', 'ref')
    ci_tin    = _col(col_map, 'tin')
    ci_branch = _col(col_map, 'branch code', 'branch')
    ci_net    = _col(col_map, 'net amount', 'net')
    ci_vat    = _col(col_map, 'input vat', 'input_vat', 'vat')
    ci_gross  = _col(col_map, 'gross amount', 'gross')
    ci_acode  = _col(col_map, 'account code', 'account_code')
    ci_adesc  = _col(col_map, 'account description', 'account_description')
    ci_debit  = _col(col_map, 'debit')
    ci_credit = _col(col_map, 'credit')
    ci_part   = _col(col_map, 'particulars')

    for rn, rv in enumerate(
            ws.iter_rows(min_row=data_start, values_only=True), data_start):
        if all(v is None for v in rv):
            continue

        dt  = _norm_date(_get(rv, ci_date))
        ref = _norm_str(_get(rv, ci_ref))
        if not dt or not ref:
            res['skipped'] += 1
            res['errors'].append(
                f'Row {rn}: missing {"date" if not dt else "reference_no"} — skipped')
            continue

        net   = _norm_float(_get(rv, ci_net))
        vat   = _norm_float(_get(rv, ci_vat))
        gross = _norm_float(_get(rv, ci_gross)) or round(net + vat, 2)
        debit = _norm_float(_get(rv, ci_debit)) or net

        data = {
            'date':                dt,
            'payee_name':          _norm_str(_get(rv, ci_payee)),
            'reference_no':        ref,
            'tin':                 _normalize_tin(_get(rv, ci_tin)),  # ← normalised
            'branch_code':         _norm_str(_get(rv, ci_branch)),
            'net_amount':          net,
            'input_vat':           vat,
            'gross_amount':        gross,
            'account_code':        _norm_str(_get(rv, ci_acode)),
            'account_description': _norm_str(_get(rv, ci_adesc)),
            'debit':               debit,
            'credit':              _norm_float(_get(rv, ci_credit)),
            'particulars':         _norm_str(_get(rv, ci_part)),
        }

        if db_manager.add_purchase_entry(data):
            res['imported'] += 1
        else:
            res['skipped'] += 1
            res['errors'].append(f'Row {rn}: insert failed — skipped')

    return res


def _import_cdj(ws, db_manager) -> dict:
    res = _empty_result('Cash Disbursement Journal')
    col_map, data_start = _find_header_row(
        ws, ['date', 'reference no', 'account code'])
    if data_start is None:
        res['errors'].append('Header row not found.')
        return res

    ci_date   = _col(col_map, 'date')
    ci_ref    = _col(col_map, 'reference no', 'reference_no', 'ref no', 'ref')
    ci_part   = _col(col_map, 'particulars')
    ci_acode  = _col(col_map, 'account code', 'account_code')
    ci_adesc  = _col(col_map, 'account description', 'account_description')
    ci_debit  = _col(col_map, 'debit')
    ci_credit = _col(col_map, 'credit')

    for rn, rv in enumerate(
            ws.iter_rows(min_row=data_start, values_only=True), data_start):
        if all(v is None for v in rv):
            continue

        dt  = _norm_date(_get(rv, ci_date))
        ref = _norm_str(_get(rv, ci_ref))
        if not dt or not ref:
            res['skipped'] += 1
            res['errors'].append(
                f'Row {rn}: missing {"date" if not dt else "reference_no"} — skipped')
            continue

        data = {
            'date':                dt,
            'reference_no':        ref,
            'particulars':         _norm_str(_get(rv, ci_part)),
            'account_code':        _norm_str(_get(rv, ci_acode)),
            'account_description': _norm_str(_get(rv, ci_adesc)),
            'debit':               _norm_float(_get(rv, ci_debit)),
            'credit':              _norm_float(_get(rv, ci_credit)),
        }

        if db_manager.add_cash_disbursement_entry(data):
            res['imported'] += 1
        else:
            res['skipped'] += 1
            res['errors'].append(f'Row {rn}: insert failed — skipped')

    return res


def _import_crj(ws, db_manager) -> dict:
    res = _empty_result('Cash Receipts Journal')
    col_map, data_start = _find_header_row(
        ws, ['date', 'reference no', 'account code'])
    if data_start is None:
        res['errors'].append('Header row not found.')
        return res

    ci_date   = _col(col_map, 'date')
    ci_ref    = _col(col_map, 'reference no', 'reference_no', 'ref no', 'ref')
    ci_part   = _col(col_map, 'particulars')
    ci_acode  = _col(col_map, 'account code', 'account_code')
    ci_adesc  = _col(col_map, 'account description', 'account_description')
    ci_debit  = _col(col_map, 'debit')
    ci_credit = _col(col_map, 'credit')

    for rn, rv in enumerate(
            ws.iter_rows(min_row=data_start, values_only=True), data_start):
        if all(v is None for v in rv):
            continue

        dt  = _norm_date(_get(rv, ci_date))
        ref = _norm_str(_get(rv, ci_ref))
        if not dt or not ref:
            res['skipped'] += 1
            res['errors'].append(
                f'Row {rn}: missing {"date" if not dt else "reference_no"} — skipped')
            continue

        data = {
            'date':                dt,
            'reference_no':        ref,
            'particulars':         _norm_str(_get(rv, ci_part)),
            'account_code':        _norm_str(_get(rv, ci_acode)),
            'account_description': _norm_str(_get(rv, ci_adesc)),
            'debit':               _norm_float(_get(rv, ci_debit)),
            'credit':              _norm_float(_get(rv, ci_credit)),
        }

        if db_manager.add_cash_receipts_entry(data):
            res['imported'] += 1
        else:
            res['skipped'] += 1
            res['errors'].append(f'Row {rn}: insert failed — skipped')

    return res


# ---------------------------------------------------------------------------
# Master entry point
# ---------------------------------------------------------------------------

_IMPORTERS = {
    'coa':       _import_coa,
    'alphalist': _import_alphalist,
    'sj':        _import_sj,
    'pj':        _import_pj,
    'cdj':       _import_cdj,
    'crj':       _import_crj,
}

_DISPLAY_NAMES = {
    'coa':       'Chart of Accounts',
    'alphalist': 'Alphalist',
    'sj':        'Sales Journal',
    'pj':        'Purchase Journal',
    'cdj':       'Cash Disbursement Journal',
    'crj':       'Cash Receipts Journal',
}


def import_full_book(xlsx_path: str, db_manager,
                     progress_callback=None) -> list[dict]:
    if not _OPENPYXL_OK:
        raise ImportError('openpyxl is required.\nInstall with:  pip install openpyxl')

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        raise RuntimeError(
            f'Cannot open "{os.path.basename(xlsx_path)}".\n'
            f'Make sure it is a valid .xlsx file.\n\nDetail: {e}')

    results = []

    for sheet_name in wb.sheetnames:
        tag = _classify_sheet(sheet_name)
        if tag is None:
            continue

        display = _DISPLAY_NAMES.get(tag, sheet_name)
        if progress_callback:
            progress_callback(display)

        ws     = wb[sheet_name]
        result = _IMPORTERS[tag](ws, db_manager)
        result['sheet'] = display
        results.append(result)

    wb.close()
    return results


def build_summary_message(results: list[dict]) -> str:
    if not results:
        return ('No supported sheets were found in the workbook.\n\n'
                'Expected sheet names: COA, Alphalist, SJ_xx, PJ_xx, CDJ_xx, CRJ_xx')

    lines = ['Full Book Import — Summary\n', '─' * 40]
    total_imported = total_skipped = 0

    for r in results:
        lines.append(
            f"  {r['sheet']:<30}  "
            f"Imported: {r['imported']:>4}   "
            f"Skipped: {r['skipped']:>4}")
        total_imported += r['imported']
        total_skipped  += r['skipped']

    lines.append('─' * 40)
    lines.append(
        f"  {'TOTAL':<30}  "
        f"Imported: {total_imported:>4}   "
        f"Skipped: {total_skipped:>4}")

    all_errors = []
    for r in results:
        for e in r['errors'][:5]:
            all_errors.append(f"  [{r['sheet']}] {e}")
        if len(r['errors']) > 5:
            all_errors.append(
                f"  [{r['sheet']}] … and {len(r['errors']) - 5} more errors")

    if all_errors:
        lines.append('\nErrors / Skipped Details:')
        lines.extend(all_errors)

    return '\n'.join(lines)