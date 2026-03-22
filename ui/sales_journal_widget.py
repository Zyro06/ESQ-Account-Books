import os
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget,
                             QTableWidgetItem, QPushButton, QLabel, QLineEdit,
                             QDialog, QFormLayout, QDialogButtonBox, QMessageBox,
                             QHeaderView, QDateEdit, QDoubleSpinBox, QTextEdit,
                             QComboBox, QGroupBox, QShortcut, QFileDialog, QFrame)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QKeySequence, QFont, QColor
from database.db_manager import DatabaseManager, _numeric_suffix
from ui.search_utils import SearchFilter, add_month_combo
from resources.file_paths import get_import_dir, get_io_dir

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

_AR_NUM  = '1110'
_VAT_NUM = '2210'
_GOODS_NUM = '4010'
_SVC_NUM   = '4020'

_XLS_COLS = [
    ('Date','date'), ('Customer Name','customer_name'), ('Reference No','reference_no'),
    ('TIN','tin'), ('Particulars','particulars'),
    ('Account Description','account_description'), ('Account Code','account_code'),
    ('Debit','debit'), ('Credit','credit'),
]
_XLS_HEADERS = [h for h,_ in _XLS_COLS]
_XLS_KEYS    = [k for _,k in _XLS_COLS]


def _parse_date(s):
    for fmt in ("MM/dd/yyyy","M/d/yyyy","M/dd/yyyy","MM/d/yyyy","yyyy-MM-dd"):
        d = QDate.fromString(s, fmt)
        if d.isValid(): return d
    return QDate()

class _DateItem(QTableWidgetItem):
    def __init__(self, text):
        super().__init__(text)
        q = _parse_date(text)
        self._sort_key = q.toString("yyyy-MM-dd") if q.isValid() else text
    def __lt__(self, other):
        if isinstance(other, _DateItem): return self._sort_key < other._sort_key
        return super().__lt__(other)


class _LineDialog(QDialog):
    """Add/edit a single journal line."""
    def __init__(self, db_manager, parent=None, line_data=None):
        super().__init__(parent)
        self.db_manager  = db_manager
        self.account_map = {}
        self.setWindowTitle("Journal Line")
        self.setModal(True); self.setMinimumWidth(460)
        layout = QFormLayout(self)
        layout.setLabelAlignment(Qt.AlignRight)

        self.account_combo = QComboBox(); self.account_combo.setEditable(True)
        self._load_accounts()
        self.account_combo.currentTextChanged.connect(self._on_account_changed)
        layout.addRow("Account:", self.account_combo)

        self.code_input = QLineEdit(); self.code_input.setReadOnly(True)
        self.code_input.setStyleSheet("background:#f0f0f0; color:#666;")
        layout.addRow("Code:", self.code_input)

        self.debit_input  = QDoubleSpinBox(); self.debit_input.setMaximum(99999999.99); self.debit_input.setDecimals(2); self.debit_input.setGroupSeparatorShown(True)
        self.credit_input = QDoubleSpinBox(); self.credit_input.setMaximum(99999999.99); self.credit_input.setDecimals(2); self.credit_input.setGroupSeparatorShown(True)
        self.debit_input.valueChanged.connect(lambda v: self.credit_input.setValue(0) if v > 0 else None)
        self.credit_input.valueChanged.connect(lambda v: self.debit_input.setValue(0) if v > 0 else None)
        layout.addRow("Debit:",  self.debit_input)
        layout.addRow("Credit:", self.credit_input)

        if line_data:
            self.account_combo.setCurrentText(line_data.get('account_description',''))
            self.code_input.setText(line_data.get('account_code',''))
            self.debit_input.setValue(line_data.get('debit',0))
            self.credit_input.setValue(line_data.get('credit',0))

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self._validate); btns.rejected.connect(self.reject)
        layout.addRow(btns)

    def _load_accounts(self):
        self.account_combo.addItem("")
        for acct in self.db_manager.get_all_accounts():
            desc = acct['account_description']
            self.account_combo.addItem(desc)
            self.account_map[desc] = acct['account_code']

    def _on_account_changed(self, text):
        self.code_input.setText(self.account_map.get(text,''))

    def _validate(self):
        if not self.account_combo.currentText().strip():
            QMessageBox.warning(self,"Validation","Please select an account."); return
        if self.debit_input.value() == 0 and self.credit_input.value() == 0:
            QMessageBox.warning(self,"Validation","Enter a debit or credit amount."); return
        self.accept()

    def get_data(self):
        return {
            'account_description': self.account_combo.currentText().strip(),
            'account_code':        self.code_input.text().strip(),
            'debit':               self.debit_input.value(),
            'credit':              self.credit_input.value(),
        }


class SalesJournalDialog(QDialog):
    def __init__(self, db_manager, parent=None, entry_data=None, is_copy=False):
        super().__init__(parent)
        self.db_manager    = db_manager
        self.user_lines    = []
        self.alphalist_map = {}
        self._ar_code = ''; self._ar_desc = 'Accounts Receivable'

        title = ("Copy Entry" if is_copy else
                 "Add Sales Entry" if entry_data is None else "Edit Sales Entry")
        self.setWindowTitle(title)
        self.setModal(True); self.setMinimumWidth(820); self.setMinimumHeight(580)

        root = QVBoxLayout(self); root.setSpacing(10)

        # Header
        hdr = QFormLayout(); hdr.setLabelAlignment(Qt.AlignRight)
        self.date_input = QDateEdit(); self.date_input.setCalendarPopup(True); self.date_input.setDisplayFormat("MM/dd/yyyy")
        self.date_input.setDate(QDate.fromString(entry_data["date"],"MM/dd/yyyy") if entry_data else QDate.currentDate())
        hdr.addRow("Date:", self.date_input)

        self.customer_input = QComboBox(); self.customer_input.setEditable(True); self.customer_input.setMinimumWidth(300)
        self._load_customers()
        hdr.addRow("Customer Name:", self.customer_input)

        self.tin_input = QLineEdit(); self.tin_input.setReadOnly(True); self.tin_input.setStyleSheet("background:#f0f0f0;color:#666;")
        hdr.addRow("TIN:", self.tin_input)

        self.reference_input = QLineEdit()
        if entry_data: self.reference_input.setText(entry_data.get("reference_no",""))
        hdr.addRow("Reference No:", self.reference_input)

        self.particulars_input = QLineEdit()
        if entry_data: self.particulars_input.setText(entry_data.get("particulars","") or "")
        hdr.addRow("Particulars:", self.particulars_input)

        root.addLayout(hdr)
        sep = QFrame(); sep.setFrameShape(QFrame.HLine); sep.setFrameShadow(QFrame.Sunken); root.addWidget(sep)

        # Lines label + info
        lines_hdr = QHBoxLayout()
        lbl = QLabel("Journal Lines:"); lbl.setStyleSheet("font-weight:bold;")
        ar_info = QLabel("  (AR will be auto-computed as the balancing debit)")
        ar_info.setStyleSheet("color:#666; font-style:italic;")
        lines_hdr.addWidget(lbl); lines_hdr.addWidget(ar_info); lines_hdr.addStretch()
        root.addLayout(lines_hdr)

        # Lines table
        self.lines_table = QTableWidget()
        self.lines_table.setColumnCount(4)
        self.lines_table.setHorizontalHeaderLabels(["Account","Code","Debit","Credit"])
        self.lines_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.lines_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.lines_table.verticalHeader().setVisible(False)
        self.lines_table.setAlternatingRowColors(True)
        lh = self.lines_table.horizontalHeader()
        lh.setSectionResizeMode(0, QHeaderView.Stretch)
        lh.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        lh.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        lh.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        root.addWidget(self.lines_table)

        # Line buttons
        btn_row = QHBoxLayout()
        add_btn  = QPushButton("Add Line");    add_btn.clicked.connect(self._add_line)
        edit_btn = QPushButton("Edit Line");   edit_btn.clicked.connect(self._edit_line)
        del_btn  = QPushButton("Remove Line"); del_btn.clicked.connect(self._remove_line)
        btn_row.addWidget(add_btn); btn_row.addWidget(edit_btn); btn_row.addWidget(del_btn); btn_row.addStretch()
        root.addLayout(btn_row)

        # Totals
        tot_row = QHBoxLayout(); tot_row.addStretch()
        bf = QFont(); bf.setBold(True)
        self.total_dr_lbl = QLabel("Total Debit:  0.00")
        self.total_cr_lbl = QLabel("Total Credit: 0.00")
        self.balance_lbl  = QLabel("")
        self.ar_lbl       = QLabel("")
        self.total_dr_lbl.setFont(bf); self.total_cr_lbl.setFont(bf)
        tot_row.addWidget(self.total_dr_lbl); tot_row.addWidget(QLabel("  |  "))
        tot_row.addWidget(self.total_cr_lbl); tot_row.addWidget(QLabel("  "))
        tot_row.addWidget(self.balance_lbl)
        root.addLayout(tot_row)
        root.addWidget(self.ar_lbl)

        btns = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        btns.accepted.connect(self._on_save); btns.rejected.connect(self.reject)
        root.addWidget(btns)

        self.customer_input.currentTextChanged.connect(self._on_customer_changed)
        self._find_ar_account()

        if entry_data:
            self.customer_input.setCurrentText(entry_data.get("customer_name","") or "")
            self.tin_input.setText(str(entry_data.get("tin","") or ""))
            for ln in entry_data.get("lines",[]):
                if _numeric_suffix(ln.get("account_code","")) != _AR_NUM:
                    self.user_lines.append(dict(ln))
        self._refresh_lines()

    def _find_ar_account(self):
        for acct in self.db_manager.get_all_accounts():
            if _numeric_suffix(acct["account_code"]) == _AR_NUM:
                self._ar_code = acct["account_code"]
                self._ar_desc = acct["account_description"]
                break

    def _load_customers(self):
        self.customer_input.addItem("")
        for entry in self.db_manager.get_all_alphalist():
            if entry.get("entry_type","Customer") not in ("Customer","Customer&Vendor"): continue
            name = entry.get("company_name") or f"{entry.get('first_name','')} {entry.get('last_name','')}".strip()
            if name:
                self.customer_input.addItem(name)
                self.alphalist_map[name] = {"tin": entry.get("tin","")}

    def _on_customer_changed(self, name):
        info = self.alphalist_map.get(name,{})
        self.tin_input.setText(str(info.get("tin","")) if info else "")

    def _add_line(self):
        dlg = _LineDialog(self.db_manager, self)
        if dlg.exec_(): self.user_lines.append(dlg.get_data()); self._refresh_lines()

    def _edit_line(self):
        row = self.lines_table.currentRow()
        if row < 0 or row >= len(self.user_lines):
            QMessageBox.warning(self,"No Selection","Select a user line to edit."); return
        dlg = _LineDialog(self.db_manager, self, self.user_lines[row])
        if dlg.exec_(): self.user_lines[row] = dlg.get_data(); self._refresh_lines()

    def _remove_line(self):
        row = self.lines_table.currentRow()
        if row < 0 or row >= len(self.user_lines):
            QMessageBox.warning(self,"No Selection","Select a user line to remove."); return
        del self.user_lines[row]; self._refresh_lines()

    def _refresh_lines(self):
        total_cr = sum(l.get("credit",0) for l in self.user_lines)
        total_dr = sum(l.get("debit", 0) for l in self.user_lines)
        ar_amount = total_cr - total_dr  # AR = balancing debit

        all_lines = list(self.user_lines)
        ar_line = None
        if ar_amount > 0:
            ar_line = {"account_description": self._ar_desc, "account_code": self._ar_code,
                       "debit": ar_amount, "credit": 0, "_auto": True}
            all_lines.append(ar_line)

        self.lines_table.setRowCount(len(all_lines))
        for r, ln in enumerate(all_lines):
            is_auto = ln.get("_auto", False)
            color   = QColor(230,240,255) if is_auto else QColor(240,255,240)
            for c, val in enumerate([
                ln.get("account_description",""), ln.get("account_code",""),
                f"{ln.get('debit',0):,.2f}"  if ln.get("debit",0)  else "",
                f"{ln.get('credit',0):,.2f}" if ln.get("credit",0) else "",
            ]):
                item = QTableWidgetItem(val)
                item.setBackground(color)
                item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                item.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter if c>=2 else Qt.AlignLeft|Qt.AlignVCenter)
                self.lines_table.setItem(r, c, item)

        final_dr = total_dr + (ar_amount if ar_amount > 0 else 0)
        self.total_dr_lbl.setText(f"Total Debit:  {final_dr:,.2f}")
        self.total_cr_lbl.setText(f"Total Credit: {total_cr:,.2f}")
        balanced = abs(final_dr - total_cr) < 0.005 and final_dr > 0
        if balanced:
            self.balance_lbl.setText("Balanced"); self.balance_lbl.setStyleSheet("color:green;font-weight:bold;")
        elif all_lines:
            self.balance_lbl.setText("Check lines"); self.balance_lbl.setStyleSheet("color:orange;font-weight:bold;")
        else:
            self.balance_lbl.setText("")
        if ar_amount > 0:
            self.ar_lbl.setText(f"  AR (Balancing Debit): {ar_amount:,.2f}")
            self.ar_lbl.setStyleSheet("color:#2244aa; font-style:italic;")
        else:
            self.ar_lbl.setText("")

    def _on_save(self):
        if not self.customer_input.currentText().strip():
            QMessageBox.warning(self,"Validation","Customer Name is required."); return
        if not self.reference_input.text().strip():
            QMessageBox.warning(self,"Validation","Reference No. is required."); return
        if not self.user_lines:
            QMessageBox.warning(self,"Validation","Add at least one journal line."); return
        self.accept()

    def get_data(self):
        total_cr = sum(l.get("credit",0) for l in self.user_lines)
        total_dr = sum(l.get("debit", 0) for l in self.user_lines)
        ar_amount = total_cr - total_dr
        lines = list(self.user_lines)
        if ar_amount > 0:
            lines.append({"account_description": self._ar_desc, "account_code": self._ar_code,
                          "debit": ar_amount, "credit": 0})
        return {
            "date":          self.date_input.date().toString("MM/dd/yyyy"),
            "customer_name": self.customer_input.currentText().strip(),
            "reference_no":  self.reference_input.text().strip(),
            "tin":           self.tin_input.text().strip(),
            "particulars":   self.particulars_input.text().strip(),
            "lines":         lines,
        }


class _ViewSJDialog(QDialog):
    def __init__(self, parent, entry):
        super().__init__(parent)
        self.setWindowTitle(f"View — {entry['reference_no']}")
        self.setModal(True); self.setMinimumWidth(700); self.setMinimumHeight(420)
        root = QVBoxLayout(self)
        info = QLabel(
            f"<b>Date:</b> {entry['date']}  &nbsp; <b>Customer:</b> {entry['customer_name']}  &nbsp; "
            f"<b>Reference:</b> {entry['reference_no']}  &nbsp; <b>TIN:</b> {entry.get('tin','')}  &nbsp; "
            f"<b>Particulars:</b> {entry.get('particulars','')}")
        info.setWordWrap(True); root.addWidget(info)
        sep = QFrame(); sep.setFrameShape(QFrame.HLine); sep.setFrameShadow(QFrame.Sunken); root.addWidget(sep)
        tbl = QTableWidget(); tbl.setColumnCount(4)
        tbl.setHorizontalHeaderLabels(["Account","Code","Debit","Credit"])
        tbl.setEditTriggers(QTableWidget.NoEditTriggers); tbl.setAlternatingRowColors(True); tbl.verticalHeader().setVisible(False)
        th = tbl.horizontalHeader(); th.setSectionResizeMode(0,QHeaderView.Stretch); th.setSectionResizeMode(1,QHeaderView.ResizeToContents); th.setSectionResizeMode(2,QHeaderView.ResizeToContents); th.setSectionResizeMode(3,QHeaderView.ResizeToContents)
        lines = entry.get("lines",[])
        tbl.setRowCount(len(lines))
        td = tc = 0.0
        for r, ln in enumerate(lines):
            d = ln.get("debit",0); c = ln.get("credit",0)
            for col, val in enumerate([ln.get("account_description",""), ln.get("account_code",""), f"{d:,.2f}" if d else "", f"{c:,.2f}" if c else ""]):
                item = QTableWidgetItem(val); item.setFlags(Qt.ItemIsSelectable|Qt.ItemIsEnabled)
                if col>=2: item.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter)
                tbl.setItem(r, col, item)
            td += d; tc += c
        root.addWidget(tbl)
        bf = QFont(); bf.setBold(True)
        tot = QLabel(f"Total Debit: {td:,.2f}  |  Total Credit: {tc:,.2f}"); tot.setFont(bf)
        root.addWidget(tot)
        close = QPushButton("Close"); close.clicked.connect(self.accept); root.addWidget(close, alignment=Qt.AlignRight)


class SalesJournalWidget(QWidget):
    def __init__(self, db_manager: DatabaseManager):
        super().__init__()
        self.db_manager  = db_manager
        self.all_entries = []
        self._setup_ui(); self._setup_shortcuts(); self.load_data()

    def _setup_ui(self):
        layout = QVBoxLayout()
        title = QLabel("SALES JOURNAL"); title.setProperty("class","title"); title.setAlignment(Qt.AlignCenter); layout.addWidget(title)

        sg = QGroupBox("Search & Filter"); sl = QHBoxLayout()
        self.month_combo = add_month_combo(sl)
        sl.addWidget(QLabel("Search:")); self.search_input = QLineEdit(); self.search_input.setPlaceholderText("Search by customer, reference, TIN, or particulars..."); self.search_input.setClearButtonEnabled(True); sl.addWidget(self.search_input)
        sl.addWidget(QLabel("From:")); self.date_from = QDateEdit(); self.date_from.setCalendarPopup(True); self.date_from.setDisplayFormat("MM/dd/yyyy"); self.date_from.setDate(QDate(2000,1,1)); sl.addWidget(self.date_from)
        sl.addWidget(QLabel("To:")); self.date_to = QDateEdit(); self.date_to.setCalendarPopup(True); self.date_to.setDisplayFormat("MM/dd/yyyy"); self.date_to.setDate(QDate.currentDate()); sl.addWidget(self.date_to)
        self.clear_btn = QPushButton("Clear Filter"); self.clear_btn.clicked.connect(self._clear_filters); sl.addWidget(self.clear_btn)
        self.results_label = QLabel("Showing: 0 of 0"); sl.addWidget(self.results_label)
        sg.setLayout(sl); layout.addWidget(sg)

        br = QHBoxLayout()
        for label, slot in [("Add Entry",self._add_entry),("Edit Entry",self._edit_entry),("Copy Entry",self._copy_entry),("View Details",self._view_details)]:
            btn = QPushButton(label); btn.clicked.connect(slot); br.addWidget(btn)
        self.delete_btn = QPushButton("Delete Entry"); self.delete_btn.setProperty("class","danger"); self.delete_btn.clicked.connect(self._delete_entry); br.addWidget(self.delete_btn)
        self.import_btn = QPushButton("Import"); self.import_btn.clicked.connect(self._import_xls); br.addWidget(self.import_btn)
        self.export_btn = QPushButton("Export"); self.export_btn.clicked.connect(self._export_xls); br.addWidget(self.export_btn)
        br.addStretch()
        self.totals_label = QLabel("Totals: Gross: 0.00 | VAT: 0.00 | Goods: 0.00 | Services: 0.00"); self.totals_label.setProperty("class","total"); br.addWidget(self.totals_label)
        layout.addLayout(br)

        self.table = QTableWidget(); self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["Date","Customer Name","Reference No.","TIN","Lines","Total Gross"])
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectRows); self.table.setSelectionMode(QTableWidget.SingleSelection)
        self.table.setSortingEnabled(True); self.table.setAlternatingRowColors(True)
        self.table.doubleClicked.connect(self._view_details)
        layout.addWidget(self.table)

        self._search = SearchFilter(table=self.table, search_input=self.search_input, results_label=self.results_label, date_from=self.date_from, date_to=self.date_to, month_combo=self.month_combo, date_col=0)
        orig = self._search._run
        def _run_with_totals(): orig(); self._update_totals()
        self._search._timer.timeout.disconnect(); self._search._timer.timeout.connect(_run_with_totals)
        self.date_from.dateChanged.connect(_run_with_totals); self.date_to.dateChanged.connect(_run_with_totals); self.month_combo.currentIndexChanged.connect(_run_with_totals)
        self.setLayout(layout)

    def _setup_shortcuts(self):
        QShortcut(QKeySequence("Ctrl+N"),       self).activated.connect(self._add_entry)
        QShortcut(QKeySequence("Ctrl+E"),       self).activated.connect(self._edit_entry)
        QShortcut(QKeySequence("Ctrl+Shift+C"), self).activated.connect(self._copy_entry)
        QShortcut(QKeySequence("Ctrl+V"),       self).activated.connect(self._view_details)
        QShortcut(QKeySequence("Ctrl+D"),       self).activated.connect(self._delete_entry)
        QShortcut(QKeySequence("Ctrl+F"),       self).activated.connect(self.search_input.setFocus)
        QShortcut(QKeySequence("Ctrl+I"),       self).activated.connect(self._import_xls)
        QShortcut(QKeySequence("Ctrl+Shift+E"), self).activated.connect(self._export_xls)

    def load_data(self):
        self.all_entries = self.db_manager.get_sales_journal()
        self._populate_table()
        self._search.refresh(); self._update_totals()

    def _populate_table(self):
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(self.all_entries))
        for r, entry in enumerate(self.all_entries):
            gross = entry.get("gross_amount",0)
            date_item = _DateItem(entry.get("date",""))
            date_item.setData(Qt.UserRole, entry)
            self.table.setItem(r, 0, date_item)
            for c, text in [(1,entry.get("customer_name","")), (2,entry.get("reference_no","")), (3,str(entry.get("tin","") or ""))]:
                self.table.setItem(r, c, QTableWidgetItem(text))
            lines_item = QTableWidgetItem(str(len(entry.get("lines",[]))))
            lines_item.setTextAlignment(Qt.AlignCenter|Qt.AlignVCenter)
            self.table.setItem(r, 4, lines_item)
            gross_item = QTableWidgetItem(f"{gross:,.2f}"); gross_item.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter); gross_item.setFlags(Qt.ItemIsSelectable|Qt.ItemIsEnabled)
            self.table.setItem(r, 5, gross_item)
        self.table.setSortingEnabled(True)

    def _update_totals(self):
        gross = vat = goods = svc = 0.0
        for r in range(self.table.rowCount()):
            if self.table.isRowHidden(r): continue
            entry = self.table.item(r,0).data(Qt.UserRole)
            if entry:
                gross += entry.get("gross_amount",0); vat   += entry.get("output_vat",0)
                goods += entry.get("goods",0);        svc   += entry.get("services",0)
        self.totals_label.setText(f"Totals: Gross: {gross:,.2f} | VAT: {vat:,.2f} | Goods: {goods:,.2f} | Services: {svc:,.2f}")

    def _clear_filters(self):
        self.search_input.clear(); self.date_from.setDate(QDate(2000,1,1)); self.date_to.setDate(QDate.currentDate()); self.month_combo.setCurrentIndex(0)

    def _get_selected(self):
        row = self.table.currentRow()
        if row < 0: return None
        item = self.table.item(row,0)
        return item.data(Qt.UserRole) if item else None

    def _add_entry(self):
        dlg = SalesJournalDialog(self.db_manager, self)
        if dlg.exec_():
            if self.db_manager.add_sales_entry(dlg.get_data()):
                self.load_data(); QMessageBox.information(self,"Success","Sales entry added successfully!")
            else: QMessageBox.warning(self,"Error","Failed to add entry!")

    def _edit_entry(self):
        entry = self._get_selected()
        if not entry: QMessageBox.warning(self,"Warning","Please select an entry to edit."); return
        dlg = SalesJournalDialog(self.db_manager, self, entry)
        if dlg.exec_():
            if self.db_manager.update_sales_entry(entry["id"], dlg.get_data()):
                self.load_data(); QMessageBox.information(self,"Success","Entry updated successfully!")
            else: QMessageBox.warning(self,"Error","Failed to update entry!")

    def _copy_entry(self):
        entry = self._get_selected()
        if not entry: QMessageBox.warning(self,"Warning","Please select an entry to copy."); return
        dlg = SalesJournalDialog(self.db_manager, self, entry, is_copy=True)
        if dlg.exec_():
            if self.db_manager.add_sales_entry(dlg.get_data()):
                self.load_data(); QMessageBox.information(self,"Success","Entry copied successfully!")
            else: QMessageBox.warning(self,"Error","Failed to copy entry!")

    def _view_details(self):
        entry = self._get_selected()
        if not entry: QMessageBox.warning(self,"Warning","Please select an entry to view."); return
        _ViewSJDialog(self, entry).exec_()

    def _delete_entry(self):
        entry = self._get_selected()
        if not entry: QMessageBox.warning(self,"Warning","Please select an entry to delete."); return
        reply = QMessageBox.question(self,"Confirm Delete", f"Delete entry '{entry['reference_no']}'?", QMessageBox.Yes|QMessageBox.No)
        if reply == QMessageBox.Yes:
            if self.db_manager.delete_sales_entry(entry["id"]):
                self.load_data(); QMessageBox.information(self,"Success","Entry deleted!")
            else: QMessageBox.warning(self,"Error","Failed to delete.")

    def _export_xls(self):
        if not _OPENPYXL_OK: QMessageBox.critical(self,"Missing Library","openpyxl required."); return
        path, _ = QFileDialog.getSaveFileName(self, "Export Sales Journal", os.path.join(get_io_dir("Sales Journal"), "sales_journal.xlsx"), "Excel Files (*.xlsx)")
        if not path: return
        rows = []
        for r in range(self.table.rowCount()):
            if self.table.isRowHidden(r): continue
            entry = self.table.item(r,0).data(Qt.UserRole)
            for ln in entry.get("lines",[]):
                rows.append({"date": entry["date"], "customer_name": entry["customer_name"],
                             "reference_no": entry["reference_no"], "tin": entry.get("tin",""),
                             "particulars": entry.get("particulars",""),
                             "account_description": ln.get("account_description",""),
                             "account_code": ln.get("account_code",""),
                             "debit": f"{ln.get('debit',0):,.2f}" if ln.get("debit",0) else "",
                             "credit": f"{ln.get('credit',0):,.2f}" if ln.get("credit",0) else ""})
        n, err = _export_to_xls(rows, path, "Sales Journal")
        if err: QMessageBox.critical(self,"Export Failed",err)
        else:   QMessageBox.information(self,"Export Successful",f"{n} line(s) exported to:\n{path}")

    def _import_xls(self):
        if not _OPENPYXL_OK: QMessageBox.critical(self,"Missing Library","openpyxl required."); return
        path, _ = QFileDialog.getOpenFileName(self, "Import Sales Journal", get_import_dir(""), "Excel Files (*.xlsx *.xls)")
        if not path: return
        try: imported, skipped, errors = _import_from_xls(path, self.db_manager)
        except Exception as exc: QMessageBox.critical(self,"Import Failed",str(exc)); return
        self.load_data()
        msg = f"Import complete.\n  Imported: {imported}\n  Skipped: {skipped}"
        if errors: msg += "\n\nDetails:\n" + "\n".join(errors[:20])
        QMessageBox.information(self,"Import Summary",msg)


def _export_to_xls(rows, path, sheet_title):
    from datetime import datetime as _dt
    try:
        wb = Workbook(); ws = wb.active; ws.title = sheet_title[:31]
        hf=XLFont(name="Arial",bold=True,color="FFFFFF",size=11); hfl=PatternFill("solid",start_color="2F5496"); ha=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cf=XLFont(name="Arial",size=10); af=PatternFill("solid",start_color="DCE6F1"); th=Side(style="thin",color="B0B0B0"); bd=Border(left=th,right=th,top=th,bottom=th)
        HR=5
        ws.merge_cells(f"A2:{get_column_letter(len(_XLS_HEADERS))}2"); ws["A2"].value=sheet_title.upper(); ws["A2"].font=XLFont(name="Arial",bold=True,size=14); ws["A2"].alignment=Alignment(horizontal="left",vertical="center")
        ws.merge_cells(f"A3:{get_column_letter(len(_XLS_HEADERS))}3"); ws["A3"].value=f"For the Year {_dt.now().year}"; ws["A3"].font=XLFont(name="Arial",italic=True,size=11); ws["A3"].alignment=Alignment(horizontal="left",vertical="center")
        ws.row_dimensions[HR].height=28
        for ci,hdr in enumerate(_XLS_HEADERS,1):
            cell=ws.cell(row=HR,column=ci,value=hdr); cell.font=hf; cell.fill=hfl; cell.alignment=ha; cell.border=bd
        NUM={"debit","credit"}
        for ri,entry in enumerate(rows):
            ri2=6+ri; ws.row_dimensions[ri2].height=18; fill=af if ri%2==0 else None
            for ci,key in enumerate(_XLS_KEYS,1):
                val=entry.get(key,"") or ""; cell=ws.cell(row=ri2,column=ci,value=val); cell.font=cf; cell.border=bd
                cell.alignment=Alignment(horizontal="right" if key in NUM else "left",vertical="center")
                if fill: cell.fill=fill
        for ci,w in {1:12,2:28,3:16,4:16,5:28,6:28,7:14,8:14,9:14}.items():
            ws.column_dimensions[get_column_letter(ci)].width=w
        ws.freeze_panes="A6"; ws.auto_filter.ref=f"A{HR}:{get_column_letter(len(_XLS_HEADERS))}{HR}"
        wb.save(path); return len(rows),""
    except Exception as exc: return 0,str(exc)


def _import_from_xls(path, db_manager):
    HMAP = {h.lower():k for h,k in _XLS_COLS}
    try: wb = load_workbook(path,read_only=True,data_only=True)
    except: raise RuntimeError(f"Cannot open file.")
    ws=wb.active; col_index={}; data_start=None
    for r_idx,rv in enumerate(ws.iter_rows(min_row=1,max_row=10,values_only=True),1):
        if not any(v and str(v).strip().lower() in HMAP for v in rv): continue
        for ci,cv in enumerate(rv):
            if cv is None: continue
            key=HMAP.get(str(cv).strip().lower())
            if key: col_index[key]=ci
        data_start=r_idx+1; break
    if not col_index: raise ValueError("Header row not found.")
    def _val(rv,key):
        idx=col_index.get(key)
        if idx is None or idx>=len(rv): return ""
        v=rv[idx]; return str(v).strip() if v is not None else ""
    # Group rows by reference
    from collections import defaultdict
    groups = defaultdict(list)
    order  = []
    for rn,rv in enumerate(ws.iter_rows(min_row=data_start,values_only=True),data_start):
        if all(v is None for v in rv): continue
        ref=_val(rv,"reference_no"); date=_val(rv,"date")
        key=(date,ref)
        if key not in groups: order.append(key)
        groups[key].append(rv)
    imported=skipped=0; errors=[]
    for key in order:
        date,ref=key
        if not date or not ref: skipped+=1; continue
        first=groups[key][0]
        data = {
            "date": date, "customer_name": _val(first,"customer_name"),
            "reference_no": ref, "tin": _val(first,"tin"),
            "particulars": _val(first,"particulars"), "lines": []
        }
        for rv in groups[key]:
            ac=_val(rv,"account_code"); ad=_val(rv,"account_description")
            try: dr=float(_val(rv,"debit").replace(",","") or 0)
            except: dr=0
            try: cr=float(_val(rv,"credit").replace(",","") or 0)
            except: cr=0
            if ac and (dr>0 or cr>0):
                data["lines"].append({"account_code":ac,"account_description":ad,"debit":dr,"credit":cr})
        if not data["lines"]: skipped+=1; errors.append(f"Ref {ref}: no valid lines skipped"); continue
        if db_manager.add_sales_entry(data): imported+=1
        else: skipped+=1; errors.append(f"Ref {ref}: failed to insert")
    wb.close(); return imported,skipped,errors