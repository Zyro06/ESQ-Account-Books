from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget,
                             QTableWidgetItem, QPushButton, QLabel, QDateEdit,
                             QHeaderView, QGroupBox, QLineEdit, QFileDialog, QMessageBox)
from PyQt5.QtCore import Qt, QDate
from database.db_manager import DatabaseManager
from ui.search_utils import SearchFilter, add_month_combo
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False


class TrialBalanceWidget(QWidget):

    def __init__(self, db_manager: DatabaseManager):
        super().__init__()
        self.db_manager = db_manager
        self._setup_ui()
        self.load_data()

    def _setup_ui(self):
        layout = QVBoxLayout()

        title = QLabel("WORKING TRIAL BALANCE")
        title.setProperty("class", "title")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # ── Search & Filter ────────────────────────────────────────────────
        filter_group  = QGroupBox("Search & Filter")
        filter_layout = QHBoxLayout()

        self.month_combo = add_month_combo(filter_layout)

        filter_layout.addWidget(QLabel("Search:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search by account code or description...")
        self.search_input.setClearButtonEnabled(True)
        filter_layout.addWidget(self.search_input)

        filter_layout.addWidget(QLabel("From:"))
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDisplayFormat("MM/dd/yyyy")
        self.date_from.setDate(QDate(2000, 1, 1))
        filter_layout.addWidget(self.date_from)

        filter_layout.addWidget(QLabel("To:"))
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDisplayFormat("MM/dd/yyyy")
        self.date_to.setDate(QDate.currentDate())
        filter_layout.addWidget(self.date_to)

        self.generate_btn = QPushButton("Generate")
        self.generate_btn.clicked.connect(self.load_data)
        filter_layout.addWidget(self.generate_btn)

        self.clear_btn = QPushButton("Clear Filter")
        self.clear_btn.clicked.connect(self._clear_filter)
        filter_layout.addWidget(self.clear_btn)

        self.export_btn = QPushButton("Export")
        self.export_btn.clicked.connect(self._export_xls)
        filter_layout.addWidget(self.export_btn)

        self.results_label = QLabel("Showing: 0 of 0")
        filter_layout.addWidget(self.results_label)

        filter_layout.addStretch()
        filter_group.setLayout(filter_layout)
        layout.addWidget(filter_group)

        # ── Table ──────────────────────────────────────────────────────────
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(
            ["Account Code", "Account Description", "Amount"])
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setAlternatingRowColors(True)
        layout.addWidget(self.table)

        # ── Totals ─────────────────────────────────────────────────────────
        self.totals_label = QLabel("")
        self.totals_label.setProperty("class", "total")
        layout.addWidget(self.totals_label)

        self.setLayout(layout)

        # SearchFilter — text search only (no date col on TB rows)
        self._search = SearchFilter(
            table         = self.table,
            search_input  = self.search_input,
            results_label = self.results_label,
        )
        # Hook totals update after search filter runs
        orig_run = self._search._run
        def _run_with_totals():
            orig_run()
            self._update_totals()
        self._search._timer.timeout.disconnect()
        self._search._timer.timeout.connect(_run_with_totals)

        # Month combo regenerates data
        self.month_combo.currentIndexChanged.connect(self._on_month_changed)

    # ------------------------------------------------------------------ helpers

    def _clear_filter(self):
        self.search_input.clear()
        self.month_combo.setCurrentIndex(0)
        self.date_from.setDate(QDate(2000, 1, 1))
        self.date_to.setDate(QDate.currentDate())
        self.load_data()

    def _on_month_changed(self):
        self.load_data()

    # ------------------------------------------------------------------ data

    def load_data(self):
        month_idx = self.month_combo.currentIndex()

        if month_idx > 0:
            year = self.date_to.date().year()
            import calendar
            last_day  = calendar.monthrange(year, month_idx)[1]
            date_from = QDate(year, month_idx, 1).toString("MM/dd/yyyy")
            date_to   = QDate(year, month_idx, last_day).toString("MM/dd/yyyy")
        else:
            date_from = self.date_from.date().toString("MM/dd/yyyy")
            date_to   = self.date_to.date().toString("MM/dd/yyyy")

        trial_balance = self.db_manager.get_trial_balance(date_from, date_to)
        self.table.setRowCount(len(trial_balance))

        for row, entry in enumerate(trial_balance):
            amount = entry['amount']
            items  = [entry['account_code'], entry['account_description'], f"{amount:,.2f}"]
            for col, item_text in enumerate(items):
                item = QTableWidgetItem(str(item_text))
                item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                if col == 2:
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    if amount > 0:
                        item.setForeground(Qt.darkGreen)
                    elif amount < 0:
                        item.setForeground(Qt.darkRed)
                self.table.setItem(row, col, item)

        self._search.refresh()
        self._update_totals()

    def _update_totals(self):
        total_debit = total_credit = 0.0
        for row in range(self.table.rowCount()):
            if self.table.isRowHidden(row):
                continue
            try:
                amount = float(self.table.item(row, 2).text().replace(',', ''))
                if amount > 0:
                    total_debit  += amount
                else:
                    total_credit += abs(amount)
            except (AttributeError, ValueError):
                pass
        self.totals_label.setText(
            f"Total Debit: {total_debit:,.2f} | Total Credit: {total_credit:,.2f}")

    # ------------------------------------------------------------------ Export

    def _export_xls(self):
        if not _OPENPYXL_OK:
            QMessageBox.critical(self, "Missing Library",
                                 "openpyxl is required.\nInstall with: pip install openpyxl")
            return

        from resources.file_paths import get_io_dir
        from datetime import datetime as _dt
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Trial Balance",
            os.path.join(get_io_dir("Trial Balance"), "trial_balance.xlsx"),
            "Excel Files (*.xlsx)")
        if not path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Working Trial Balance"

            hdr_font  = XLFont(name='Arial', bold=True, color='FFFFFF', size=11)
            hdr_fill  = PatternFill('solid', start_color='2F5496')
            hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_font = XLFont(name='Arial', size=10)
            alt_fill  = PatternFill('solid', start_color='DCE6F1')
            pos_font  = XLFont(name='Arial', size=10, color='006100')  # dark green
            neg_font  = XLFont(name='Arial', size=10, color='9C0006')  # dark red
            thin      = Side(style='thin', color='B0B0B0')
            border    = Border(left=thin, right=thin, top=thin, bottom=thin)
            r_align   = Alignment(horizontal='right', vertical='center')
            l_align   = Alignment(horizontal='left',  vertical='center')

            # Title
            ws.merge_cells('A2:C2')
            ws['A2'].value = 'WORKING TRIAL BALANCE'
            ws['A2'].font  = XLFont(name='Arial', bold=True, size=14)
            ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[2].height = 22

            ws.merge_cells('A3:C3')
            ws['A3'].value = f'For the Year {_dt.now().year}'
            ws['A3'].font  = XLFont(name='Arial', italic=True, size=11)
            ws['A3'].alignment = Alignment(horizontal='left', vertical='center')

            HR = 5
            headers = ['Account Code', 'Account Description', 'Amount']
            ws.row_dimensions[HR].height = 28
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=HR, column=ci, value=h)
                cell.font = hdr_font; cell.fill = hdr_fill
                cell.alignment = hdr_align; cell.border = border

            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 45
            ws.column_dimensions['C'].width = 16
            ws.freeze_panes = 'A6'

            ri = 0
            total_debit = total_credit = 0.0
            for row in range(self.table.rowCount()):
                if self.table.isRowHidden(row):
                    continue
                row_idx = 6 + ri
                fill = alt_fill if ri % 2 == 0 else None
                ws.row_dimensions[row_idx].height = 18
                code   = self.table.item(row, 0).text() if self.table.item(row, 0) else ''
                desc   = self.table.item(row, 1).text() if self.table.item(row, 1) else ''
                amt_s  = self.table.item(row, 2).text() if self.table.item(row, 2) else '0'
                amount = float(amt_s.replace(',', ''))

                for ci, (val, align) in enumerate([
                    (code,   l_align),
                    (desc,   l_align),
                    (amt_s,  r_align),
                ], 1):
                    cell = ws.cell(row=row_idx, column=ci, value=val)
                    cell.border = border; cell.alignment = align
                    if ci == 3:
                        cell.font = pos_font if amount > 0 else (neg_font if amount < 0 else cell_font)
                    else:
                        cell.font = cell_font
                    if fill:
                        cell.fill = fill

                if amount > 0:
                    total_debit  += amount
                else:
                    total_credit += abs(amount)
                ri += 1

            # Totals row
            tot_row = 6 + ri
            ws.row_dimensions[tot_row].height = 20
            tot_fill = PatternFill('solid', start_color='D9E1F2')
            tot_font = XLFont(name='Arial', bold=True, size=10)
            for ci, val in enumerate(
                    ['', f'Total Debit: {total_debit:,.2f}  |  Total Credit: {total_credit:,.2f}', ''], 1):
                cell = ws.cell(row=tot_row, column=ci, value=val)
                cell.font = tot_font; cell.fill = tot_fill; cell.border = border
                cell.alignment = l_align

            ws.auto_filter.ref = f'A{HR}:C{HR}'
            wb.save(path)
            QMessageBox.information(self, "Export Successful", f"Exported to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", str(e))